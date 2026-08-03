"""
app.py — Flask REST API
Meat Products of India — Billing & Inventory Management App
"""

import os
import re
import json
import shutil
import sqlite3
import urllib.parse
from functools import wraps
from datetime import datetime, date, timedelta
from flask import Flask, jsonify, request, render_template, send_file, session, render_template_string, redirect
from flask_cors import CORS
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from werkzeug.security import generate_password_hash, check_password_hash
from database import get_db, init_db, dict_row, dict_rows, post_ledger_entry
from license_manager import get_license_info, activate_subscription
from license_sync import sync_with_cloud_server, notify_cloud_payment, re_register_with_cloud
from cloud_backup import start_cloud_backup_scheduler, run_cloud_backup_job

import sys
if getattr(sys, 'frozen', False):
    bundle_dir = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    app = Flask(__name__, template_folder=os.path.join(bundle_dir, 'templates'), static_folder=os.path.join(bundle_dir, 'static'))
else:
    app = Flask(__name__)

app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

@app.after_request
def add_header(response):
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

if not os.environ.get('APP_SECRET_KEY'):
    print("WARNING: using insecure default secret key — set APP_SECRET_KEY env var for production")
    app.secret_key = 'mpi_secret_key_2025_change_in_production_!@#'
else:
    app.secret_key = os.environ.get('APP_SECRET_KEY')

CORS(app, supports_credentials=True)

# ─── Role Permissions ────────────────────────────────────────────────────────

ROLE_LABELS = {
    'admin':         'Developer Superuser',
    'md':            'Managing Director',
    'manager':       'Store Manager',
    'accountant':    'Accountant',
    'counter_staff': 'Counter Staff',
    'tester':        'Tester Staff (Sandbox)',
}

# Pages each role can access (used by frontend)
ROLE_PAGES = {
    'admin':         ['dashboard','billing','bills','inventory','stock-in','purchase-orders',
                      'categories','customers','suppliers','expenses','accounts','reports','settings','users'],
    'md':            ['dashboard','billing','bills','inventory','stock-in','purchase-orders',
                      'categories','customers','suppliers','expenses','accounts','reports','settings','users'],
    'manager':       ['dashboard','billing','bills','inventory','stock-in','purchase-orders',
                      'categories','customers','suppliers','expenses','accounts','reports','users'],
    'accountant':    ['dashboard','billing','bills','customers','expenses','accounts','reports','inventory'],
    'counter_staff': ['billing','bills','customers'],
    'tester':        ['billing','bills','customers'],
}

# ─── Auth Helpers ────────────────────────────────────────────────────────────

def current_user():
    return {
        'id':        session.get('user_id'),
        'username':  session.get('username'),
        'full_name': session.get('full_name'),
        'role':      session.get('user_role'),
    } if 'user_id' in session else None


def require_auth(f):
    """Decorator: requires a valid session."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'status': 'error', 'message': 'Authentication required'}), 401
        if session.get('must_change_password'):
            if request.endpoint not in ('change_password', 'logout', 'me'):
                return jsonify({
                    'status': 'error',
                    'must_change_password': True,
                    'message': 'Password change required before accessing system features'
                }), 403
        return f(*args, **kwargs)
    return decorated


def require_role(*roles):
    """Decorator: requires one of the specified roles."""
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if 'user_id' not in session:
                return jsonify({'status': 'error', 'message': 'Authentication required'}), 401
            if session.get('must_change_password'):
                return jsonify({
                    'status': 'error',
                    'must_change_password': True,
                    'message': 'Password change required before accessing system features'
                }), 403
            user_role = session.get('user_role')
            if user_role not in ['admin', 'md'] and user_role not in roles:
                return jsonify({'status': 'error', 'message': 'Insufficient permissions for this action'}), 403
            return f(*args, **kwargs)
        return decorated
    return decorator


def get_user_permissions(user_id):
    """Fetch all permission codes granted to the user, respecting role hierarchy (superiors inherit subordinate permissions)."""
    if not user_id:
        return set()
    conn = get_db()
    user = conn.execute('SELECT id, role, role_id FROM users WHERE id=?', (user_id,)).fetchone()
    if not user:
        conn.close()
        return set()

    role_id = user['role_id']
    if not role_id:
        legacy_map = {'admin': 1, 'md': 1, 'manager': 2, 'accountant': 3, 'counter_staff': 4, 'tester': 4}
        role_id = legacy_map.get(user['role'], 4)

    # Superior roles inherit all permissions of subordinate roles
    role_ids = [role_id]
    if user['role'] in ('admin', 'md') or role_id == 1:
        conn.close()
        return {'*'}
    elif user['role'] == 'manager' or role_id == 2:
        role_ids.extend([3, 4, 5])
    elif user['role'] == 'accountant' or role_id == 3:
        role_ids.append(4)

    placeholders = ','.join('?' for _ in role_ids)
    rows = conn.execute(f'''
        SELECT DISTINCT p.code FROM permissions p
        JOIN role_permissions rp ON p.id = rp.permission_id
        WHERE rp.role_id IN ({placeholders})
    ''', role_ids).fetchall()
    conn.close()

    perms = {r['code'] for r in rows}
    # Ensure billing / counter staff roles always possess customer creation & view permissions
    if user['role'] in ('counter_staff', 'tester', 'billing_staff') or role_id == 4:
        perms.update(['customers.view', 'customers.manage', 'customers.create', 'billing.create', 'billing.view', 'billing.hold', 'billing.payment'])
    return perms


def log_permission_audit(user_id, username, permission_code, route, method, allowed):
    """Record an entry to audit_log table."""
    try:
        conn = get_db()
        conn.execute('''
            INSERT INTO audit_log (user_id, username, permission_code, route, method, allowed)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (user_id, username, permission_code, route, method, 1 if allowed else 0))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[Audit Log Error] {e}")


def require_permission(*codes):
    """
    Decorator enforcing granular role-based permissions.
    1. Checks if current user's role has ANY of the specified permission codes (or '*').
    2. Logs denied attempts to audit_log and returns 403 error/redirect.
    3. Audits allowed sensitive actions (settings.gst_toggle, billing.void_bill, inventory.edit_price).
    """
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            is_ajax = request.is_json or request.path.startswith('/api/') or request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            if 'user_id' not in session:
                if is_ajax:
                    return jsonify({'status': 'error', 'message': 'Authentication required'}), 401
                return redirect('/#login')

            if session.get('must_change_password'):
                if request.endpoint not in ('change_password', 'logout', 'me'):
                    return jsonify({
                        'status': 'error',
                        'must_change_password': True,
                        'message': 'Password change required before accessing system features'
                    }), 403

            user_id = session.get('user_id')
            username = session.get('username', 'unknown')
            route = request.path
            method = request.method

            user_perms = get_user_permissions(user_id)
            is_allowed = ('*' in user_perms) or any(c in user_perms for c in codes)
            code = codes[0] if codes else 'access'

            # Special discount cap check for billing.give_discount
            eval_code = code
            if is_allowed and 'billing.create' in codes:
                d = request.get_json(silent=True) or {}
                disc = float(d.get('discount_percent', 0))
                if disc > 0 and 'billing.give_discount' not in user_perms and '*' not in user_perms:
                    conn = get_db()
                    cap_row = conn.execute("SELECT value FROM shop_settings WHERE key='max_discount_staff'").fetchone()
                    conn.close()
                    try:
                        max_disc = float(cap_row['value']) if cap_row and cap_row['value'] else 10.0
                    except (ValueError, TypeError):
                        max_disc = 10.0

                    if disc > max_disc:
                        is_allowed = False
                        eval_code = 'billing.give_discount'

            if not is_allowed:
                log_permission_audit(user_id, username, eval_code, route, method, allowed=0)
                if is_ajax:
                    return jsonify({
                        'status': 'error',
                        'message': f'Permission denied: requires "{eval_code}" permission'
                    }), 403
                return render_template_string('<h1>403 Forbidden</h1><p>You do not have permission ({{ code }}) to access this page.</p>', code=eval_code), 403

            # Audit ALLOWED sensitive actions
            if eval_code in ('settings.gst_toggle', 'billing.void_bill', 'inventory.edit_price'):
                log_permission_audit(user_id, username, eval_code, route, method, allowed=1)

            return f(*args, **kwargs)
        return decorated
    return decorator


# ─── Helpers ────────────────────────────────────────────────────────────────

def ok(data=None, message="Success"):
    return jsonify({"status": "ok", "message": message, "data": data})

def err(message="Error", code=400):
    return jsonify({"status": "error", "message": message}), code

def log_activity(action, description=None, table_name=None, record_id=None):
    """Record a user action to the activity_log table for audit trail."""
    if 'user_id' not in session:
        return
    try:
        conn = get_db()
        conn.execute(
            '''INSERT INTO activity_log (username, full_name, role, action, description, table_name, record_id, ip_address)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
            (
                session.get('username', 'unknown'),
                session.get('full_name', 'Unknown User'),
                session.get('user_role', 'unknown'),
                action,
                description,
                table_name,
                record_id,
                request.remote_addr
            )
        )
        conn.commit()
        conn.close()
    except Exception:
        pass  # Never let audit logging break the main request

def is_counter_staff():
    return session.get('user_role') in ('counter_staff', 'tester')

def wants_excel():
    return request.args.get('export', '').strip().lower() == 'excel'

def export_to_excel(sheets, filename):
    wb = openpyxl.Workbook()
    # Remove default worksheet
    wb.remove(wb.active)

    header_font = Font(name='Inter', size=11, bold=True, color='0F172A')
    header_fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
    header_align = Alignment(horizontal='left', vertical='center')
    data_font = Font(name='Inter', size=10, color='0F172A')

    for idx, sheet_item in enumerate(sheets):
        raw_name = sheet_item.get('sheet_name', f'Sheet{idx+1}')
        # Sanitize sheet_name: 31-char limit, strip invalid chars : \ / ? * [ ]
        clean_name = re.sub(r'[\:\\/\?\*\[\]]', '', str(raw_name))[:31].strip() or f'Sheet{idx+1}'

        ws = wb.create_sheet(title=clean_name)

        headers = sheet_item.get('headers', [])
        rows = sheet_item.get('rows', [])

        # Write Headers
        if headers:
            ws.append(headers)
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            ws.freeze_panes = 'A2'

        # Write Data Rows
        start_row = 2 if headers else 1
        for r_offset, row_data in enumerate(rows):
            r_num = start_row + r_offset
            ws.append(row_data)
            for c_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=r_num, column=c_idx)
                cell.font = data_font

                # Format numeric columns as numbers
                if isinstance(val, float):
                    cell.value = float(val)
                    cell.number_format = '#,##0.00'
                elif isinstance(val, int) and not isinstance(val, bool):
                    cell.value = int(val)
                    cell.number_format = '#,##0'

        # Calculate max string length per column & set auto-sized column widths (capped at 50)
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                v_str = str(cell.value) if cell.value is not None else ''
                max_len = max(max_len, len(v_str))
            ws.column_dimensions[col_letter].width = max(min(max_len + 4, 50), 12)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Sanitize filename
    clean_filename = re.sub(r'[^\w\-]', '_', str(filename)).strip('_')
    if not clean_filename.lower().endswith('.xlsx'):
        clean_filename += '.xlsx'

    response = app.response_class(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response.headers['Content-Disposition'] = f'attachment; filename="{clean_filename}"'
    return response

def get_setting(key, conn=None):
    close = False
    if conn is None:
        conn = get_db(); close = True
    row = conn.execute('SELECT value FROM shop_settings WHERE key=?', (key,)).fetchone()
    if close: conn.close()
    return row['value'] if row else None

def next_bill_no(conn):
    if session.get('user_role') == 'tester':
        n = int(get_setting('next_test_no', conn) or 1)
        bill_no = f"TEST-{n:05d}"
        conn.execute("UPDATE shop_settings SET value=? WHERE key='next_test_no'", (str(n + 1),))
        return bill_no

    n = int(get_setting('next_bill_no', conn) or 1)
    prefix = get_setting('bill_prefix', conn) or 'MPI'
    bill_no = f"{prefix}-{n:05d}"
    conn.execute("UPDATE shop_settings SET value=? WHERE key='next_bill_no'", (str(n + 1),))
    return bill_no

def next_po_no(conn):
    n = int(get_setting('next_po_no', conn) or 1)
    po_no = f"PO-{n:05d}"
    conn.execute("UPDATE shop_settings SET value=? WHERE key='next_po_no'", (str(n + 1),))
    return po_no

def next_cn_no(conn):
    n = int(get_setting('next_cn_no', conn) or 1)
    cn_no = f"CN-{n:05d}"
    conn.execute("INSERT INTO shop_settings (key, value) VALUES ('next_cn_no', ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value", (str(n + 1),))
    return cn_no

def next_conversion_no(conn):
    n = int(get_setting('next_conversion_no', conn) or 1)
    cnv_no = f"CNV-{n:05d}"
    conn.execute("INSERT INTO shop_settings (key, value) VALUES ('next_conversion_no', ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value", (str(n + 1),))
    return cnv_no


# ─── Subscription & 12-Digit Online Activation Endpoints ────────────────────

@app.route('/api/license/status', methods=['GET'])
def license_status():
    info = get_license_info()
    return ok(info)

@app.route('/api/license/system-info', methods=['GET'])
def license_system_info():
    """Returns outlet_code and machine_id for display in the title bar / UI."""
    from license_manager import get_machine_id as _si_mid
    machine_id = _si_mid()
    conn = get_db()
    oc_row  = conn.execute("SELECT value FROM shop_settings WHERE key='outlet_code'").fetchone()
    mid_row = conn.execute("SELECT value FROM shop_settings WHERE key='system_machine_id'").fetchone()
    name_row = conn.execute("SELECT value FROM shop_settings WHERE key='outlet_name'").fetchone()
    city_row = conn.execute("SELECT value FROM shop_settings WHERE key='outlet_city'").fetchone()
    conn.close()
    outlet_code = oc_row['value']  if oc_row  else None
    outlet_name = name_row['value'] if name_row else None
    outlet_city = city_row['value'] if city_row else None
    return ok({
        'machine_id':     machine_id,
        'machine_id_short': machine_id[:8],
        'outlet_code':    outlet_code,
        'outlet_name':    outlet_name,
        'outlet_city':    outlet_city,
        'registered':     bool(outlet_code),
    })

@app.route('/api/license/activate', methods=['POST'])
@require_auth
def license_activate():
    d = request.get_json() or {}
    key = (d.get('key') or '').strip()
    if not key:
        return err("Please enter a 12-digit activation key")

    success, msg = activate_subscription(key)
    if not success:
        return err(msg, 400)

    info = get_license_info()
    return ok(info, message=msg)

@app.route('/api/license/sync-cloud', methods=['POST'])
@require_auth
def license_sync_cloud():
    # Check if this outlet was re-registered offline and needs to push to server first
    try:
        _conn = get_db()
        _rereg_row = _conn.execute("SELECT value FROM shop_settings WHERE key='outlet_needs_reregister'").fetchone()
        _is_pending = _rereg_row and str(_rereg_row['value']).strip() == '1'
        _conn.close()
        if _is_pending:
            re_register_with_cloud()
    except Exception:
        pass

    success, msg = sync_with_cloud_server()
    info = get_license_info()
    return ok(info, message=msg)

@app.route('/api/license/notify-payment', methods=['POST'])
@require_auth
def license_notify_payment():
    d = request.get_json() or {}
    utr = (d.get('utr_number') or '').strip()
    success, msg = notify_cloud_payment(utr)
    if success:
        return ok(message=msg)
    return err(msg, 400)

def deduct_fefo_stock(conn, product_id, qty_needed):
    qty_left = float(qty_needed)
    if qty_left <= 0:
        return 0.0
    batches = conn.execute('''
        SELECT sb.id, sb.quantity_remaining, sb.unit_price, sb.unit_cost, p.purchase_price
        FROM stock_batches sb
        JOIN products p ON sb.product_id = p.id
        WHERE sb.product_id = ? AND sb.quantity_remaining > 0
        ORDER BY CASE WHEN sb.expiry_date IS NULL OR sb.expiry_date='' THEN 1 ELSE 0 END ASC,
                 sb.expiry_date ASC, sb.id ASC
    ''', (product_id,)).fetchall()

    total_cost_deducted = 0.0
    for b in batches:
        b_id = b['id']
        b_qty = float(b['quantity_remaining'])
        take_qty = min(b_qty, qty_left)

        effective_unit_cost = float(b['unit_cost'] if b['unit_cost'] is not None else (b['unit_price'] if b['unit_price'] is not None else (b['purchase_price'] or 0)))
        total_cost_deducted += take_qty * effective_unit_cost

        conn.execute(
            'UPDATE stock_batches SET quantity_remaining = quantity_remaining - ? WHERE id = ?',
            (take_qty, b_id)
        )
        qty_left -= take_qty
        if qty_left <= 0:
            break

    avg_unit_cost = (total_cost_deducted / float(qty_needed)) if float(qty_needed) > 0 else 0.0
    return avg_unit_cost


# Base unit for inventory storage is purchase_unit. Stock-in quantities are recorded internally in purchase_unit terms.
# Stock-out (billing) quantities are supplied in sale_unit terms and converted to purchase_unit via conversion_factor before updating stock.
def update_stock(conn, product_id, delta, tx_type, unit_price=0, ref=None,
                 supplier_id=None, expiry_date=None, notes=None, status='approved',
                 created_by=None, approved_by=None, batch_no=None, unit_cost=None):
    delta = float(delta)
    cur = conn.execute(
        '''INSERT INTO stock_transactions
           (product_id, type, quantity, unit_price, reference_id, supplier_id, expiry_date, notes, status, created_by, approved_by)
           VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
        (product_id, tx_type, abs(delta), unit_price, ref, supplier_id, expiry_date, notes, status, created_by, approved_by)
    )
    tx_id = cur.lastrowid
    deducted_cost = 0.0

    if status == 'approved':
        conn.execute(
            'UPDATE products SET current_stock = current_stock + ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?',
            (delta, product_id)
        )
        if delta > 0 and tx_type in ('in', 'adjustment', 'conversion_in'):
            b_no = batch_no or (f"RET-{ref}" if ref and "Reversal" in (notes or "") else f"BATCH-{tx_id:05d}")
            b_cost = unit_cost if unit_cost is not None else unit_price
            conn.execute(
                '''INSERT INTO stock_batches
                   (product_id, batch_no, quantity_remaining, unit_price, unit_cost, expiry_date, supplier_id, stock_transaction_id)
                   VALUES (?,?,?,?,?,?,?,?)''',
                (product_id, b_no, abs(delta), unit_price, b_cost, expiry_date, supplier_id, tx_id)
            )
        elif delta < 0:
            deducted_cost = deduct_fefo_stock(conn, product_id, abs(delta))

        # Check for auto-reorder low stock alert
        prod = conn.execute(
            'SELECT name, min_stock, current_stock, purchase_unit, active FROM products WHERE id=?',
            (product_id,)
        ).fetchone()
        if prod and prod['active'] and prod['current_stock'] <= prod['min_stock']:
            pname = prod['name']
            pstock = round(prod['current_stock'], 3)
            pmin = prod['min_stock']
            punit = prod['purchase_unit'] or 'kg'
            
            recent_alert = conn.execute('''
                SELECT id FROM notifications
                WHERE title LIKE '%Low Stock%'
                  AND message LIKE ?
                  AND read = 0
                  AND created_at >= datetime('now', '-1 day')
            ''', (f"%{pname}%",)).fetchone()
            
            if not recent_alert:
                title = "⚠️ Low Stock Alert"
                msg = f"⚠️ Low Stock: {pname} is at {pstock} {punit}, at or below reorder point of {pmin}."
                conn.execute('INSERT INTO notifications (target_role, title, message) VALUES (?,?,?)',
                             ('manager', title, msg))
                conn.execute('INSERT INTO notifications (target_role, title, message) VALUES (?,?,?)',
                             ('admin', title, msg))

    return deducted_cost if delta < 0 else tx_id

# ─── Root & Jinja Helpers ───────────────────────────────────────────────────

@app.context_processor
def utility_processor():
    def can(permission_code):
        if 'user_id' not in session:
            return False
        perms = get_user_permissions(session['user_id'])
        return (permission_code in perms) or ('*' in perms)

    user_perms = sorted(list(get_user_permissions(session.get('user_id')))) if 'user_id' in session else []
    return dict(can=can, user_permissions=user_perms)

@app.route('/')
def index():
    is_desktop = os.environ.get('FLASK_DESKTOP') == '1'
    return render_template('index.html', is_desktop=is_desktop)

@app.route('/invoice/<int:bill_id>')
def invoice_page(bill_id):
    return render_template('invoice_print.html', bill_id=bill_id)

@app.route('/invoice/<int:bill_id>/thermal')
def invoice_thermal_page(bill_id):
    width = request.args.get('width', '80')
    if width not in ('58', '80'):
        width = '80'
    return render_template('invoice_thermal.html', bill_id=bill_id, width=width)


# ─── Auth ────────────────────────────────────────────────────────────────────

@app.route('/api/auth/login', methods=['POST'])
def login():
    try:
        d = request.get_json()
        if d is None:
            return err("Invalid or missing JSON payload")
        username = (d.get('username') or '').strip()
        password = d.get('password') or ''
        if not username or not password:
            return err("Username and password required")

        # Ensure DB tables exist before attempting login
        try:
            from database import init_db
            init_db()
        except Exception:
            pass

        # ── Re-Registration & Cloud Sync Check (Background Thread) ─────────────
        try:
            import threading
            from license_sync import sync_with_cloud_server as _login_sync
            threading.Thread(target=_login_sync, daemon=True).start()
        except Exception:
            pass

        try:
            from license_manager import get_license_info as _get_lic
            _lic = _get_lic()
            if _lic.get('status') == 'needs_reregister':
                return err(
                    "⚠️ This outlet was deleted from the central server. "
                    "Please re-register this system using the 'Register Managing Director' button on the login screen.",
                    403
                )
        except Exception:
            pass

        conn = get_db()
        now = datetime.now()
        now_str = now.isoformat()

        attempt = conn.execute('SELECT * FROM login_attempts WHERE username=? COLLATE NOCASE', (username,)).fetchone()
        if attempt and attempt['locked_until']:
            try:
                locked_until = datetime.fromisoformat(attempt['locked_until'])
                if now < locked_until:
                    mins_left = max(1, int((locked_until - now).total_seconds() / 60) + 1)
                    conn.close()
                    return err(f"Account locked due to 5 consecutive failed login attempts. Please try again after {mins_left} minutes.", 400)
            except Exception:
                pass

        user = conn.execute(
            'SELECT * FROM users WHERE username=? COLLATE NOCASE AND active=1', (username,)
        ).fetchone()

        if not user or not check_password_hash(user['password_hash'], password):
            if attempt:
                try:
                    last_failed = datetime.fromisoformat(attempt['last_failed_at']) if attempt['last_failed_at'] else now
                except Exception:
                    last_failed = now

                if now - last_failed < timedelta(minutes=10):
                    new_count = (attempt['failed_count'] or 0) + 1
                else:
                    new_count = 1

                if new_count >= 5:
                    lock_until_str = (now + timedelta(minutes=15)).isoformat()
                    conn.execute(
                        'UPDATE login_attempts SET failed_count=?, last_failed_at=?, locked_until=? WHERE username=? COLLATE NOCASE',
                        (new_count, now_str, lock_until_str, username)
                    )
                else:
                    conn.execute(
                        'UPDATE login_attempts SET failed_count=?, last_failed_at=?, locked_until=NULL WHERE username=? COLLATE NOCASE',
                        (new_count, now_str, username)
                    )
            else:
                conn.execute(
                    'INSERT INTO login_attempts (username, failed_count, last_failed_at) VALUES (?, 1, ?)',
                    (username, now_str)
                )

            conn.commit()
            conn.close()
            return err("Invalid username or password", 401)

        conn.execute('DELETE FROM login_attempts WHERE username=? COLLATE NOCASE', (username,))
        conn.execute('UPDATE users SET last_login=CURRENT_TIMESTAMP WHERE id=?', (user['id'],))
        conn.commit()
        conn.close()

        user_dict = dict(user)
        must_change = bool(user_dict.get('must_change_password', 0))

        # ── Hardware-Binding Cross-Check (non-admin roles only) ────────────────────
        if user['role'] != 'admin':
            from license_manager import get_machine_id as _login_mid
            from database import get_db as _ldb
            current_mid = _login_mid()
            lconn = _ldb()
            oc_row  = lconn.execute("SELECT value FROM shop_settings WHERE key='outlet_code'").fetchone()
            mid_row = lconn.execute("SELECT value FROM shop_settings WHERE key='system_machine_id'").fetchone()
            lconn.close()
            machine_outlet_code = (oc_row['value'] or '').strip().upper()  if oc_row and oc_row['value']  else None
            machine_sys_id      = (mid_row['value'] or '').strip().upper() if mid_row and mid_row['value'] else None

            user_oc_val  = user_dict.get('outlet_code')
            user_mid_val = user_dict.get('machine_id')
            user_oc  = (user_oc_val or '').strip().upper()  if user_oc_val  else None
            user_mid = (user_mid_val or '').strip().upper() if user_mid_val else None

            if machine_outlet_code and user_oc and user_oc != machine_outlet_code:
                return err(
                    f"❌ Access Denied. Your account belongs to outlet {user_oc}. "
                    f"This machine is outlet {machine_outlet_code}. "
                    f"Please use the computer assigned to your outlet.", 403
                )
            if machine_sys_id and user_mid and user_mid != current_mid:
                return err(
                    f"❌ Access Denied. Your account is registered to a different computer. "
                    f"Contact your Managing Director if you believe this is an error.", 403
                )

        session.clear()
        session['user_id']              = user['id']
        session['username']             = user['username']
        session['full_name']            = user['full_name']
        session['user_role']            = user['role']
        session['outlet_code']          = user_dict.get('outlet_code') or ''
        session['must_change_password'] = must_change
        session.permanent               = True

        resp_data = {
            'id':                   user['id'],
            'username':             user['username'],
            'full_name':            user['full_name'],
            'role':                 user['role'],
            'role_label':           ROLE_LABELS.get(user['role'], user['role']),
            'pages':                ROLE_PAGES.get(user['role'], []),
            'permissions':          sorted(list(get_user_permissions(user['id']))),
            'outlet_code':          user_dict.get('outlet_code') or '',
            'machine_id':           (user_dict.get('machine_id') or '')[:8],
            'must_change_password': must_change,
        }
        return ok(resp_data, "Login successful")
    except Exception as e:
        import traceback
        traceback.print_exc()
        return err(f"Login failure: {str(e)}", 500)

@app.route('/api/auth/register-md', methods=['POST'])
def register_md():
    try:
        from database import init_db
        init_db()

        # ── First-Run Gate Check ──────────────────────────────────────────────────
        # Allow fresh registration if: no owner exists yet, OR if the outlet was
        # deleted by the developer and needs to re-register (outlet_needs_reregister=1).
        conn = get_db()
        needs_reregister_row = conn.execute(
            "SELECT value FROM shop_settings WHERE key='outlet_needs_reregister'"
        ).fetchone()
        is_reregister = needs_reregister_row and str(needs_reregister_row['value']).strip() == '1'

        existing_owner = conn.execute(
            "SELECT id FROM users WHERE (role = 'md' OR role = 'admin') AND username != 'sudo' LIMIT 1"
        ).fetchone()
        conn.close()

        if existing_owner and not is_reregister:
            return err("An owner account has already been registered on this system. Contact support to transfer ownership.", 403)

        # If re-registration: wipe old MD/admin accounts and all local users so system starts fresh
        if is_reregister and existing_owner:
            conn = get_db()
            conn.execute("DELETE FROM users WHERE role IN ('md', 'admin') AND username != 'sudo'")
            conn.execute("DELETE FROM users WHERE role IN ('manager','accountant','counter_staff','tester')")
            conn.execute("INSERT OR REPLACE INTO shop_settings (key, value) VALUES ('outlet_needs_reregister', '0')")
            conn.commit()
            conn.close()

        import json as _json
        import urllib.request as _urllib_req

        d = request.get_json() or {}
        username   = (d.get('username')   or '').strip()
        password   = d.get('password')    or ''
        full_name  = (d.get('full_name')  or '').strip()
        group_name = (d.get('group_name') or '').strip()

        # Outlet address fields
        outlet_name  = (d.get('outlet_name')  or '').strip()
        outlet_phone = (d.get('outlet_phone') or '').strip()
        addr1        = (d.get('addr1')        or '').strip()
        addr2        = (d.get('addr2')        or '').strip()
        city         = (d.get('city')         or '').strip()
        state        = (d.get('state')        or '').strip()
        pincode      = (d.get('pincode')      or '').strip()

        if not username:  return err("Username required")
        if not full_name: return err("Managing Director Full Name required")
        if len(password) < 6: return err("Password must be at least 6 characters")
        if not outlet_name: return err("Branch / Outlet Name is required")
        if not city:        return err("City is required")
        if not state:       return err("State is required")
        if pincode and (not pincode.isdigit() or len(pincode) != 6):
            return err("Pincode must be exactly 6 digits")

        from license_manager import get_machine_id as _get_mid
        from license_sync import get_cloud_server_url as _get_cloud_url
        machine_id = _get_mid()

        # ── Call cloud server to assign/retrieve outlet code ──────────────────
        outlet_code = None
        cloud_error = None
        try:
            cloud_url  = _get_cloud_url()
            full_addr  = ', '.join(filter(None, [addr1, addr2, city, state, pincode]))
            payload_bytes = _json.dumps({
                'machine_id':   machine_id,
                'md_username':  username,
                'md_fullname':  full_name,
                'group_name':   group_name,
                'outlet_name':  outlet_name,
                'outlet_phone': outlet_phone,
                'address':      full_addr,
                'city':         city,
                'state':        state,
                'pincode':      pincode,
            }).encode('utf-8')
            req = _urllib_req.Request(
                f"{cloud_url}/api/v1/outlet/register",
                data=payload_bytes,
                headers={'Content-Type': 'application/json', 'User-Agent': 'MPI-Billing-App/1.0'},
                method='POST'
            )
            with _urllib_req.urlopen(req, timeout=10) as resp:
                resp_data = _json.loads(resp.read().decode())
                if resp_data.get('status') == 'ok':
                    outlet_code = resp_data.get('outlet_code')
        except Exception as e:
            cloud_error = str(e)
            # Fallback: generate code locally if cloud is unreachable
            prefix = ''.join(c for c in outlet_name.upper() if c.isalpha())[:2] or 'XX'
            if len(prefix) < 2: prefix = (prefix + 'XX')[:2]
            outlet_code = f"{prefix}01"

        conn = get_db()

        # Save outlet details + outlet_code + machine_id to shop_settings
        full_address = ', '.join(filter(None, [addr1, addr2, city, state, pincode]))
        outlet_settings = {
            'shop_name':       outlet_name,
            'md_group_name':   group_name or outlet_name,
            'outlet_name':     outlet_name,
            'outlet_phone':    outlet_phone,
            'outlet_city':     city,
            'outlet_state':    state,
            'outlet_pincode':  pincode,
            'shop_address':    full_address,
            'shop_phone':      outlet_phone,
            'outlet_code':     outlet_code,
            'system_machine_id': machine_id,
            'outlet_needs_reregister': '0',
            'outlet_revoked':  '0',
        }
        for k, v in outlet_settings.items():
            conn.execute(
                "INSERT INTO shop_settings (key, value) VALUES (?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                (k, v)
            )

        existing = conn.execute('SELECT id FROM users WHERE username=? COLLATE NOCASE', (username,)).fetchone()
        if existing:
            conn.execute(
                'UPDATE users SET password_hash=?, full_name=?, role="md", active=1 WHERE id=?',
                (generate_password_hash(password), full_name, existing['id'])
            )
            conn.commit(); conn.close()
            session['user_id']   = existing['id']
            session['username']  = username
            session['full_name'] = full_name
            session['user_role'] = 'md'
            try:
                log_activity('MD_REGISTER', f"MD account updated. Outlet: '{outlet_name}' ({outlet_code}), Machine: {machine_id[:8]}", 'users', existing['id'])
            except Exception:
                pass
            # Sync newly created MD user & outlet details to central developer portal immediately
            try:
                from license_sync import sync_with_cloud_server as _reg_sync
                _reg_sync()
            except Exception:
                pass
            msg = f"MD account updated! Outlet Code: {outlet_code}"
            if cloud_error: msg += f" (offline fallback — sync when online)"
            return ok({'outlet_code': outlet_code, 'machine_id': machine_id}, message=msg)
        else:
            c = conn.execute(
                'INSERT INTO users (username, password_hash, full_name, role) VALUES (?,?,?,"md")',
                (username, generate_password_hash(password), full_name)
            )
            uid = c.lastrowid
            conn.commit(); conn.close()
            session['user_id']   = uid
            session['username']  = username
            session['full_name'] = full_name
            session['user_role'] = 'md'
            try:
                log_activity('MD_REGISTER', f"MD registered. Outlet: '{outlet_name}' ({outlet_code}), Machine: {machine_id[:8]}", 'users', uid)
            except Exception:
                pass
            # Sync newly created MD user & outlet details to central developer portal immediately
            try:
                from license_sync import sync_with_cloud_server as _reg_sync
                _reg_sync()
            except Exception:
                pass
            msg = f"Managing Director registered! Outlet Code: {outlet_code}"
            if cloud_error: msg += f" (offline fallback — sync when online)"
            return ok({'outlet_code': outlet_code, 'machine_id': machine_id}, message=msg), 201

    except Exception as exc:
        import traceback
        traceback.print_exc()
        return err(f"Registration failed: {str(exc)}", 500)

@app.route('/api/auth/logout', methods=['POST'])
def logout():
    # If logging out from Tester account, clean up all test bills & restore stock
    if session.get('user_role') == 'tester':
        conn = get_db()
        test_bills = conn.execute('SELECT id, bill_no FROM bills WHERE is_test=1 OR bill_no LIKE "TEST-%"').fetchall()
        for tb in test_bills:
            bid = tb['id']
            items = conn.execute('SELECT product_id, quantity FROM bill_items WHERE bill_id=?', (bid,)).fetchall()
            for it in items:
                if it['product_id']:
                    conn.execute('UPDATE products SET current_stock = current_stock + ? WHERE id=?', (it['quantity'], it['product_id']))
            conn.execute('DELETE FROM stock_transactions WHERE reference_id=?', (tb['bill_no'],))
            conn.execute('DELETE FROM bill_items WHERE bill_id=?', (bid,))
            conn.execute('DELETE FROM bills WHERE id=?', (bid,))
        conn.commit()
        conn.close()

    session.clear()
    return ok(message="Logged out")

# ─── Background Cloud License Sync Loop ─────────────────────────────────────
_cloud_sync_thread_started = False

def _background_cloud_sync_loop():
    import time
    time.sleep(2)
    while True:
        try:
            sync_with_cloud_server()
        except Exception:
            pass
        time.sleep(60)

def start_background_cloud_sync():
    global _cloud_sync_thread_started
    if not _cloud_sync_thread_started:
        _cloud_sync_thread_started = True
        import threading
        t = threading.Thread(target=_background_cloud_sync_loop, daemon=True)
        t.start()

try:
    start_background_cloud_sync()
except Exception as _e:
    print(f"Background cloud sync thread init notice: {_e}")

@app.route('/api/auth/me', methods=['GET'])
def me():
    if 'user_id' not in session:
        return err("Not authenticated", 401)
    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE id=?', (session['user_id'],)).fetchone()
    conn.close()
    if not user or not user['active']:
        session.clear()
        return err("Session expired", 401)

    # ⚡ Auto-sync with central license server on page load / session check!
    try:
        import threading
        threading.Thread(target=sync_with_cloud_server, daemon=True).start()
    except Exception:
        pass

    must_change = bool(dict(user).get('must_change_password', 0))
    return ok({
        'id':                   user['id'],
        'username':             user['username'],
        'full_name':            user['full_name'],
        'role':                 user['role'],
        'role_label':           ROLE_LABELS.get(user['role'], user['role']),
        'pages':                ROLE_PAGES.get(user['role'], []),
        'permissions':          sorted(list(get_user_permissions(user['id']))),
        'last_login':           user['last_login'],
        'must_change_password': must_change,
    })

@app.route('/api/auth/change-password', methods=['POST'])
@require_auth
def change_password():
    d = request.get_json()
    if d is None:
        return err("Invalid or missing JSON payload")
    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE id=?', (session['user_id'],)).fetchone()
    if not check_password_hash(user['password_hash'], d.get('current_password', '')):
        conn.close()
        return err("Current password is incorrect")
    new_pwd = d.get('new_password', '')
    if len(new_pwd) < 6:
        conn.close()
        return err("New password must be at least 6 characters")
    conn.execute('UPDATE users SET password_hash=?, must_change_password=0 WHERE id=?',
                 (generate_password_hash(new_pwd), session['user_id']))
    conn.commit(); conn.close()
    session['must_change_password'] = False
    return ok(data={'must_change_password': False}, message="Password changed successfully")

# ─── User Management & Role Administration ────────────────────────────────────

@app.route('/api/auth/users', methods=['GET'])
@require_permission('users.view')
def list_users():
    conn = get_db()
    current_role = session.get('user_role')
    if current_role == 'admin':
        rows = conn.execute(
            'SELECT id, username, full_name, role, active, last_login, created_at, employee_id FROM users ORDER BY role, full_name'
        ).fetchall()
    else:
        # Completely hide developer superuser (admin role) from MD and Manager
        rows = conn.execute(
            "SELECT id, username, full_name, role, active, last_login, created_at, employee_id FROM users WHERE role != 'admin' ORDER BY role, full_name"
        ).fetchall()
    conn.close()
    result = [dict(r) for r in rows]
    for r in result:
        r['role_label'] = ROLE_LABELS.get(r['role'], r['role'])
    return ok(result)

@app.route('/api/auth/users', methods=['POST'])
@require_permission('users.manage')
def create_user():
    d = request.get_json()
    if d is None:
        return err("Invalid or missing JSON payload")
    username = (d.get('username') or '').strip()
    password = d.get('password') or ''
    full_name = (d.get('full_name') or '').strip()
    employee_id = (d.get('employee_id') or '').strip().upper()
    role = d.get('role', '')
    if not username: return err("Username required")
    if not full_name: return err("Full name required")
    if not employee_id: return err("Employee ID is mandatory to identify employees uniquely")
    if len(password) < 6: return err("Password must be at least 6 characters")
    if role not in ROLE_LABELS: return err("Invalid role")

    # Manager restriction: cannot create Developer or Managing Director accounts
    if session.get('user_role') == 'manager' and role in ('admin', 'md'):
        return err("Managers are not authorized to create Developer or Managing Director accounts", 403)

    # Stamp this machine's outlet_code and machine_id onto the new user
    from license_manager import get_machine_id as _cu_mid
    current_machine_id = _cu_mid()
    conn = get_db()

    # Check unique employee ID
    dup = conn.execute('SELECT username FROM users WHERE employee_id=? COLLATE NOCASE', (employee_id,)).fetchone()
    if dup:
        conn.close()
        return err(f"Employee ID '{employee_id}' is already assigned to @{dup['username']}")

    oc_row = conn.execute("SELECT value FROM shop_settings WHERE key='outlet_code'").fetchone()
    machine_outlet_code = oc_row['value'].strip() if oc_row else None

    try:
        c = conn.execute(
            'INSERT INTO users (username, password_hash, full_name, role, outlet_code, machine_id, employee_id) VALUES (?,?,?,?,?,?,?)',
            (username, generate_password_hash(password), full_name, role, machine_outlet_code, current_machine_id, employee_id)
        )
        conn.commit()
        row = conn.execute(
            'SELECT id, username, full_name, role, active, last_login, created_at, outlet_code, employee_id FROM users WHERE id=?',
            (c.lastrowid,)
        ).fetchone()
        conn.close()
        result = dict(row)
        result['role_label'] = ROLE_LABELS.get(result['role'])
        outlet_info = f" for outlet {machine_outlet_code}" if machine_outlet_code else ""
        log_activity('CREATE_USER', f"Created user @{username} ({full_name}) with Emp ID {employee_id} as {ROLE_LABELS.get(role, role)}{outlet_info}", 'users', c.lastrowid)
        return ok(result, "User created"), 201
    except sqlite3.IntegrityError:
        conn.close()
        return err("Username already exists")

@app.route('/api/auth/users/<int:uid>', methods=['PUT'])
@require_permission('users.manage')
def update_user(uid):
    d = request.get_json()
    if d is None:
        return err("Invalid or missing JSON payload")
    conn = get_db()

    target_user = conn.execute('SELECT * FROM users WHERE id=?', (uid,)).fetchone()
    if not target_user:
        conn.close()
        return err("User not found", 404)

    # Developer (admin) accounts cannot be modified by non-developer roles (MD, Manager)
    if target_user['role'] == 'admin' and session.get('user_role') != 'admin':
        conn.close()
        return err("Managing Directors and Store Managers are not authorized to modify Developer accounts", 403)

    # Manager restriction: cannot edit or reset password for Developers or Managing Directors
    if session.get('user_role') == 'manager':
        if target_user['role'] in ('admin', 'md') or d.get('role') in ('admin', 'md'):
            conn.close()
            return err("Managers are not authorized to modify Developer or Managing Director accounts", 403)

    full_name = (d.get('full_name') or '').strip()
    role = d.get('role', target_user['role'])
    active = int(d.get('active', target_user['active']))
    if not full_name: conn.close(); return err("Full name required")
    if role not in ROLE_LABELS: conn.close(); return err("Invalid role")

    # Prevent self-demotion or self-deactivation
    if uid == session['user_id'] and role != session['user_role']:
        conn.close(); return err("Cannot change your own role")
    if uid == session['user_id'] and not active:
        conn.close(); return err("Cannot deactivate your own account")

    conn.execute(
        'UPDATE users SET full_name=?, role=?, active=? WHERE id=?',
        (full_name, role, active, uid)
    )

    # Optional password reset
    new_pwd = d.get('new_password', '')
    if new_pwd:
        if len(new_pwd) < 6: conn.close(); return err("Password must be at least 6 characters")
        conn.execute('UPDATE users SET password_hash=? WHERE id=?',
                     (generate_password_hash(new_pwd), uid))

        # Audit notification if Manager changed someone's password
        if session.get('user_role') == 'manager':
            msg = f"Manager @{session.get('username')} ({session.get('full_name')}) changed password for @{target_user['username']} ({target_user['full_name']} — {ROLE_LABELS.get(target_user['role'])})"
            conn.execute(
                'INSERT INTO notifications (target_role, title, message) VALUES (?,?,?)',
                ('admin', '🔑 Password Changed by Manager', msg)
            )

    conn.commit()
    row = conn.execute(
        'SELECT id, username, full_name, role, active, last_login, created_at FROM users WHERE id=?', (uid,)
    ).fetchone()
    conn.close()
    result = dict(row)
    result['role_label'] = ROLE_LABELS.get(result['role'])
    return ok(result)

@app.route('/api/auth/users/<int:uid>', methods=['DELETE'])
@require_permission('users.manage')
def delete_user(uid):
    if uid == session['user_id']:
        return err("Cannot delete your own account")

    conn = get_db()
    target_user = conn.execute('SELECT * FROM users WHERE id=?', (uid,)).fetchone()
    if not target_user:
        conn.close()
        return err("User not found", 404)

    # Developer (admin) accounts cannot be deleted by non-developer roles (MD, Manager)
    if target_user['role'] == 'admin' and session.get('user_role') != 'admin':
        conn.close()
        return err("Managing Directors and Store Managers are not authorized to delete Developer accounts", 403)

    if session.get('user_role') == 'manager' and target_user['role'] in ('admin', 'md'):
        conn.close()
        return err("Managers cannot delete Developer or Managing Director accounts", 403)

    conn.execute('DELETE FROM users WHERE id=?', (uid,))
    conn.commit(); conn.close()
    log_activity('DELETE_USER', f"Deleted user @{target_user['username']} ({target_user['full_name']} — {ROLE_LABELS.get(target_user['role'], target_user['role'])})", 'users', uid)
    return ok(message="User deleted")

# ─── Notifications (Managing Director Alerts) ────────────────────────────────

@app.route('/api/notifications', methods=['GET'])
@require_permission('notifications.view')
def get_notifications():
    conn = get_db()
    rows = conn.execute(
        'SELECT * FROM notifications WHERE target_role IN ("admin","md") ORDER BY created_at DESC LIMIT 20'
    ).fetchall()
    unread_count = conn.execute(
        'SELECT COUNT(*) FROM notifications WHERE target_role IN ("admin","md") AND read=0'
    ).fetchone()[0]
    conn.close()
    return ok({'notifications': dict_rows(rows), 'unread_count': unread_count})

@app.route('/api/notifications/read', methods=['POST'])
@require_permission('notifications.view')
def mark_notifications_read():
    conn = get_db()
    conn.execute('UPDATE notifications SET read=1 WHERE target_role="admin"')
    conn.commit(); conn.close()
    return ok(message="Notifications marked as read")


# ─── Settings ───────────────────────────────────────────────────────────────

@app.route('/api/settings', methods=['GET'])
@require_permission('settings.view')
def get_settings():
    conn = get_db()
    rows = conn.execute('SELECT key, value FROM shop_settings').fetchall()
    conn.close()
    return ok({r['key']: r['value'] for r in rows})

@app.route('/api/settings', methods=['POST'])
@require_permission('settings.manage')
def update_settings():
    data = request.get_json()
    if data is None or not isinstance(data, dict):
        return err("Invalid or missing JSON payload")
    conn = get_db()
    changed = []
    for k, v in data.items():
        conn.execute(
            'INSERT INTO shop_settings (key, value) VALUES (?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value',
            (k, str(v))
        )
        changed.append(k)
    conn.commit(); conn.close()
    # Log if GST was toggled
    if 'gst_enabled' in changed:
        gst_val = data.get('gst_enabled', '')
        log_activity('TOGGLE_GST', f"GST billing {'enabled' if str(gst_val)=='true' else 'disabled'}", 'shop_settings')
    log_activity('UPDATE_SETTINGS', f"Updated settings: {', '.join(changed)}", 'shop_settings')
    return ok(message="Settings saved")

@app.route('/api/settings/toggle-gst', methods=['POST'])
@require_permission('settings.gst_toggle')
def toggle_gst():
    d = request.get_json() or {}
    enabled = d.get('enabled', True)
    conn = get_db()
    conn.execute(
        "INSERT INTO shop_settings (key, value) VALUES ('gst_enabled',?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        ('true' if enabled else 'false',)
    )
    conn.commit(); conn.close()
    log_activity('TOGGLE_GST', f"GST billing {'enabled' if enabled else 'disabled'}", 'shop_settings')
    return ok(message=f"GST {'enabled' if enabled else 'disabled'} successfully")

# ─── Activity Audit Log ──────────────────────────────────────────────────────

@app.route('/api/activity-log', methods=['GET'])
@require_permission('activity.view')
def get_activity_log():
    page     = max(int(request.args.get('page', 1)), 1)
    per_page = int(request.args.get('per_page', 50))
    role_f   = request.args.get('role', '')
    user_f   = request.args.get('username', '')
    action_f = request.args.get('action', '')
    offset   = (page - 1) * per_page

    filters = []
    params  = []
    if role_f:   filters.append("role = ?");     params.append(role_f)
    if user_f:   filters.append("username LIKE ?"); params.append(f'%{user_f}%')
    if action_f: filters.append("action = ?");   params.append(action_f)

    where = ('WHERE ' + ' AND '.join(filters)) if filters else ''
    conn = get_db()
    total = conn.execute(f'SELECT COUNT(*) FROM activity_log {where}', params).fetchone()[0]
    rows  = conn.execute(
        f'SELECT * FROM activity_log {where} ORDER BY id DESC LIMIT ? OFFSET ?',
        params + [per_page, offset]
    ).fetchall()
    conn.close()
    return ok({
        'total': total, 'page': page, 'per_page': per_page,
        'pages': (total + per_page - 1) // per_page,
        'logs': [dict(r) for r in rows]
    })

# ─── Categories ─────────────────────────────────────────────────────────────

@app.route('/api/categories', methods=['GET'])
@require_permission('inventory.view')
def list_categories():
    conn = get_db()
    rows = conn.execute('SELECT * FROM categories ORDER BY name').fetchall()
    conn.close()
    cat_list = dict_rows(rows)

    if request.args.get('tree', '').lower() == 'true':
        nodes_map = {c['id']: {**c, 'children': []} for c in cat_list}
        root_nodes = []
        for c in cat_list:
            cid = c['id']
            pid = c.get('parent_category_id')
            if pid and pid in nodes_map and pid != cid:
                nodes_map[pid]['children'].append(nodes_map[cid])
            else:
                root_nodes.append(nodes_map[cid])
        return ok(root_nodes)

    return ok(cat_list)

@app.route('/api/categories', methods=['POST'])
@require_permission('inventory.create')
def create_category():
    d = request.get_json()
    if d is None:
        return err("Invalid or missing JSON payload")
    if not d.get('name'):
        return err("Category name is required")
    conn = get_db()
    try:
        c = conn.execute(
            'INSERT INTO categories (name, gst_rate, hsn_code, description, parent_category_id) VALUES (?,?,?,?,?)',
            (d['name'], d.get('gst_rate', 0), d.get('hsn_code', ''), d.get('description', ''), d.get('parent_category_id'))
        )
        conn.commit()
        row = conn.execute('SELECT * FROM categories WHERE id=?', (c.lastrowid,)).fetchone()
        conn.close()
        return ok(dict_row(row), "Category created"), 201
    except sqlite3.IntegrityError:
        conn.close()
        return err("Category name already exists")

@app.route('/api/categories/<int:cat_id>', methods=['PUT'])
@require_permission('inventory.edit')
def update_category(cat_id):
    d = request.get_json()
    if d is None:
        return err("Invalid or missing JSON payload")
    if not d.get('name'):
        return err("Category name is required")

    conn = get_db()
    cat = conn.execute('SELECT * FROM categories WHERE id=?', (cat_id,)).fetchone()
    if not cat:
        conn.close()
        return err("Category not found", 404)

    p_id = d.get('parent_category_id')
    if p_id and int(p_id) == cat_id:
        conn.close()
        return err("Category cannot be its own parent")

    try:
        conn.execute(
            'UPDATE categories SET name=?, gst_rate=?, hsn_code=?, description=?, parent_category_id=? WHERE id=?',
            (d['name'], d.get('gst_rate', 0), d.get('hsn_code', ''), d.get('description', ''), p_id, cat_id)
        )
        conn.commit()
        row = conn.execute('SELECT * FROM categories WHERE id=?', (cat_id,)).fetchone()
        conn.close()
        return ok(dict_row(row))
    except sqlite3.IntegrityError:
        conn.close()
        return err("Category name already exists")

@app.route('/api/categories/<int:cat_id>', methods=['DELETE'])
@require_permission('inventory.delete')
def delete_category(cat_id):
    conn = get_db()
    conn.execute('DELETE FROM categories WHERE id=?', (cat_id,))
    conn.commit(); conn.close()
    return ok(message="Category deleted")

# ─── Products ───────────────────────────────────────────────────────────────

@app.route('/api/products', methods=['GET'])
@require_permission('inventory.view')
def list_products():
    conn = get_db()
    active_only = request.args.get('active', 'true').lower() == 'true'
    q = request.args.get('q', '').strip()
    cat_id = request.args.get('category_id', '')
    product_type_param = request.args.get('product_type', '').strip().lower()

    sql = '''
        SELECT p.*, c.name AS category_name, c.gst_rate AS cat_gst_rate
        FROM products p
        LEFT JOIN categories c ON p.category_id = c.id
        WHERE 1=1
    '''
    params = []
    if active_only:
        sql += ' AND p.active = 1'
    if q:
        sql += ' AND (p.name LIKE ? OR p.barcode LIKE ? OR p.code LIKE ? OR p.brand LIKE ?)'
        params += [f'%{q}%', f'%{q}%', f'%{q}%', f'%{q}%']
    if cat_id:
        sql += ' AND p.category_id = ?'
        params.append(cat_id)
    if product_type_param in ('perishable', 'general'):
        sql += ' AND p.product_type = ?'
        params.append(product_type_param)

    sql += ' ORDER BY p.name'
    rows = conn.execute(sql, params).fetchall()
    conn.close()

    if wants_excel():
        if is_counter_staff():
            return err("Counter Staff are not authorized to export product data", 403)
        excel_rows = []
        for p in dict_rows(rows):
            excel_rows.append([
                p.get('code') or '',
                p.get('name') or '',
                p.get('category_name') or '',
                p.get('product_type') or 'perishable',
                float(p.get('current_stock') or 0),
                p.get('purchase_unit') or 'kg',
                p.get('sale_unit') or 'kg',
                float(p.get('purchase_price') or 0),
                float(p.get('selling_price') or 0),
                float(p.get('mrp') or 0),
                float(p.get('gst_rate') or 0),
                p.get('hsn_code') or '',
                float(p.get('min_stock') or 0)
            ])
        sheets = [{"sheet_name": "Products", "headers": ["Code","Name","Category","Product Type","Current Stock","Purchase Unit","Sale Unit","Purchase Price (₹)","Selling Price (₹)","MRP (₹)","GST Rate (%)","HSN Code","Min Stock"], "rows": excel_rows}]
        return export_to_excel(sheets, "products_list")

    product_list = dict_rows(rows)
    # Counter staff must not see confidential buying prices
    if is_counter_staff():
        for p in product_list:
            p.pop('purchase_price', None)
            p.pop('mrp', None)
            p.pop('min_stock', None)
            p.pop('reorder_qty', None)
    return ok(product_list)

@app.route('/api/products/<int:pid>', methods=['GET'])
@require_permission('inventory.view')
def get_product(pid):
    conn = get_db()
    row = conn.execute(
        '''SELECT p.*, c.name AS category_name FROM products p
           LEFT JOIN categories c ON p.category_id=c.id WHERE p.id=?''', (pid,)
    ).fetchone()
    conn.close()
    if not row: return err("Product not found", 404)
    p = dict_row(row)
    if is_counter_staff():
        p.pop('purchase_price', None)
        p.pop('mrp', None)
        p.pop('min_stock', None)
        p.pop('reorder_qty', None)
    return ok(p)

@app.route('/api/products/barcode/<barcode>', methods=['GET'])
@require_permission('inventory.view')
def get_product_by_barcode(barcode):
    b = (barcode or '').strip()
    if not b:
        return err("Barcode required", 400)

    conn = get_db()
    row = conn.execute('''
        SELECT p.*, c.name AS category_name
        FROM products p
        LEFT JOIN categories c ON p.category_id = c.id
        WHERE p.barcode = ? AND p.active = 1
    ''', (b,)).fetchone()
    conn.close()

    if not row:
        return err(f"Product not found for barcode: {b}", 404)

    return ok(dict_row(row))

@app.route('/api/products', methods=['POST'])
@require_permission('inventory.create')
def create_product():
    d = request.get_json()
    if not d.get('name'): return err("Product name is required")
    code = (d.get('code') or '').strip().upper()
    if not code or len(code) != 4 or not code.isalnum():
        return err("Product code must be exactly 4 alphanumeric characters (e.g. CHIC, MUTC)")

    ptype = (d.get('product_type') or 'perishable').strip().lower()
    if ptype not in ('perishable', 'general'):
        return err("product_type must be either 'perishable' or 'general'")

    p_unit = (d.get('purchase_unit') or d.get('unit') or ('unit' if ptype == 'general' else 'kg')).strip()
    s_unit = (d.get('sale_unit') or d.get('unit') or p_unit or ('unit' if ptype == 'general' else 'kg')).strip()
    if p_unit == s_unit:
        conv_factor = 1.0
    else:
        try:
            conv_factor = float(d.get('conversion_factor', 1.0))
            if conv_factor <= 0:
                return err("Conversion factor must be greater than 0")
        except (ValueError, TypeError):
            return err("Invalid conversion_factor")

    mrp_val = float(d.get('mrp')) if d.get('mrp') is not None else None
    inc_tax = int(d.get('is_price_inclusive_of_tax', 1 if ptype == 'general' else 0))
    brand_val = (d.get('brand') or '').strip() or None
    pack_size_val = (d.get('pack_size') or '').strip() or None
    lead_time = int(d.get('reorder_lead_time_days', 7 if ptype == 'general' else 1))

    conn = get_db()
    try:
        c = conn.execute(
            '''INSERT INTO products
               (category_id, name, code, hsn_code, unit, purchase_unit, sale_unit, conversion_factor,
                purchase_price, selling_price, gst_rate, min_stock, current_stock, barcode,
                product_type, mrp, is_price_inclusive_of_tax, brand, pack_size, reorder_lead_time_days)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (d.get('category_id'), d['name'], code, d.get('hsn_code', ''),
             s_unit, p_unit, s_unit, conv_factor,
             d.get('purchase_price', 0), d.get('selling_price', 0),
             d.get('gst_rate', 0), d.get('min_stock', 1), 0,
             d.get('barcode') or None,
             ptype, mrp_val, inc_tax, brand_val, pack_size_val, lead_time)
        )
        pid = c.lastrowid
        role = session.get('user_role')
        username = session.get('username')
        st_status = 'approved' if role in ('admin', 'md', 'manager') else 'pending_verification'
        st_approver = username if role in ('admin', 'md', 'manager') else None

        if d.get('current_stock', 0) > 0:
            update_stock(conn, pid, d['current_stock'], 'in', d.get('purchase_price', 0),
                         notes='Opening stock', status=st_status, created_by=username, approved_by=st_approver)

        conn.commit()
        row = conn.execute(
            'SELECT p.*, c.name AS category_name FROM products p LEFT JOIN categories c ON p.category_id=c.id WHERE p.id=?',
            (pid,)
        ).fetchone()
        conn.close()
        log_activity('ADD_PRODUCT', f"Added product '{d['name']}' (code: {code})", 'products', pid)
        return ok(dict_row(row), "Product created"), 201
    except sqlite3.IntegrityError as e:
        conn.close()
        if "UNIQUE constraint failed: products.code" in str(e):
            return err(f"Product code '{code}' already exists. Please choose a unique 4-letter code.")
        return err(str(e))

@app.route('/api/products/<int:pid>', methods=['PUT'])
@require_permission('inventory.edit')
def update_product(pid):
    d = request.get_json()
    code = (d.get('code') or '').strip().upper()
    if not code or len(code) != 4 or not code.isalnum():
        return err("Product code must be exactly 4 alphanumeric characters (e.g. CHIC, MUTC)")

    ptype = (d.get('product_type') or 'perishable').strip().lower()
    if ptype not in ('perishable', 'general'):
        return err("product_type must be either 'perishable' or 'general'")

    p_unit = (d.get('purchase_unit') or d.get('unit') or ('unit' if ptype == 'general' else 'kg')).strip()
    s_unit = (d.get('sale_unit') or d.get('unit') or p_unit or ('unit' if ptype == 'general' else 'kg')).strip()
    if p_unit == s_unit:
        conv_factor = 1.0
    else:
        try:
            conv_factor = float(d.get('conversion_factor', 1.0))
            if conv_factor <= 0:
                return err("Conversion factor must be greater than 0")
        except (ValueError, TypeError):
            return err("Invalid conversion_factor")

    mrp_val = float(d.get('mrp')) if d.get('mrp') is not None else None
    inc_tax = int(d.get('is_price_inclusive_of_tax', 1 if ptype == 'general' else 0))
    brand_val = (d.get('brand') or '').strip() or None
    pack_size_val = (d.get('pack_size') or '').strip() or None
    lead_time = int(d.get('reorder_lead_time_days', 7 if ptype == 'general' else 1))

    conn = get_db()
    old_prod = conn.execute('SELECT selling_price, purchase_price FROM products WHERE id=?', (pid,)).fetchone()
    if old_prod:
        new_sp = float(d.get('selling_price', old_prod['selling_price']))
        new_pp = float(d.get('purchase_price', old_prod['purchase_price']))
        if new_sp != old_prod['selling_price'] or new_pp != old_prod['purchase_price']:
            log_permission_audit(session.get('user_id'), session.get('username'), 'inventory.edit_price', request.path, request.method, allowed=1)

    try:
        conn.execute(
            '''UPDATE products SET
               category_id=?, name=?, code=?, hsn_code=?, unit=?, purchase_unit=?, sale_unit=?,
               conversion_factor=?, purchase_price=?, selling_price=?, gst_rate=?, min_stock=?,
               barcode=?, product_type=?, mrp=?, is_price_inclusive_of_tax=?, brand=?, pack_size=?,
               reorder_lead_time_days=?, active=?, updated_at=CURRENT_TIMESTAMP
               WHERE id=?''',
            (d.get('category_id'), d['name'], code, d.get('hsn_code', ''),
             s_unit, p_unit, s_unit, conv_factor,
             d.get('purchase_price', 0), d.get('selling_price', 0), d.get('gst_rate', 0),
             d.get('min_stock', 1), d.get('barcode') or None,
             ptype, mrp_val, inc_tax, brand_val, pack_size_val, lead_time,
             d.get('active', 1), pid)
        )
        conn.commit()
        row = conn.execute(
            'SELECT p.*, c.name AS category_name FROM products p LEFT JOIN categories c ON p.category_id=c.id WHERE p.id=?',
            (pid,)
        ).fetchone()
        conn.close()
        return ok(dict_row(row))
    except sqlite3.IntegrityError as e:
        conn.close()
        if "UNIQUE constraint failed: products.code" in str(e):
            return err(f"Product code '{code}' already exists. Please choose a unique 4-letter code.")
        return err(str(e))

@app.route('/api/products/<int:pid>', methods=['DELETE'])
@require_permission('inventory.delete')
def delete_product(pid):
    conn = get_db()
    conn.execute('UPDATE products SET active=0 WHERE id=?', (pid,))
    conn.commit(); conn.close()
    return ok(message="Product deactivated")

@app.route('/api/products/<int:pid>/batches', methods=['GET'])
@require_auth
def get_product_batches(pid):
    conn = get_db()
    prod = conn.execute('SELECT id, name FROM products WHERE id=?', (pid,)).fetchone()
    if not prod:
        conn.close()
        return err("Product not found", 404)
    rows = conn.execute('''
        SELECT sb.*, s.name AS supplier_name
        FROM stock_batches sb
        LEFT JOIN suppliers s ON sb.supplier_id = s.id
        WHERE sb.product_id = ? AND sb.quantity_remaining > 0
        ORDER BY CASE WHEN sb.expiry_date IS NULL OR sb.expiry_date='' THEN 1 ELSE 0 END ASC,
                 sb.expiry_date ASC, sb.id ASC
    ''', (pid,)).fetchall()
    conn.close()
    return ok(dict_rows(rows))

@app.route('/api/products/low-stock', methods=['GET'])
@require_auth
def low_stock():
    ptype = request.args.get('product_type', '').strip().lower()
    conn = get_db()
    sql = '''SELECT p.*, c.name AS category_name FROM products p
             LEFT JOIN categories c ON p.category_id=c.id
             WHERE p.active=1 AND p.current_stock <= p.min_stock'''
    params = []
    if ptype in ('perishable', 'general'):
        sql += ' AND p.product_type = ?'
        params.append(ptype)
    sql += ' ORDER BY p.current_stock ASC'
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    result = dict_rows(rows)
    for r in result:
        conv = float(r.get('conversion_factor') or 1.0)
        r['current_stock_sale_unit'] = round(float(r['current_stock']) * conv, 3)
    return ok(result)


# ─── Stock Conversions / Processing Journals ──────────────────────────────

@app.route('/api/stock/conversions', methods=['POST'])
@require_permission('stock.conversions')
def create_stock_conversion():
    d = request.get_json()
    if d is None:
        return err("Invalid or missing JSON payload")

    input_pid = d.get('input_product_id')
    try:
        input_qty = float(d.get('input_quantity', 0))
    except (ValueError, TypeError):
        input_qty = 0.0

    outputs = d.get('outputs', [])
    notes = (d.get('notes') or '').strip()

    if not input_pid or input_qty <= 0:
        return err("input_product_id and valid input_quantity > 0 are required")
    if not isinstance(outputs, list) or not outputs:
        return err("At least one output product is required in outputs")

    conn = get_db()
    input_prod = conn.execute('SELECT * FROM products WHERE id=? AND active=1', (input_pid,)).fetchone()
    if not input_prod:
        conn.close()
        return err("Input product not found or inactive", 404)

    if float(input_prod['current_stock'] or 0) < input_qty:
        conn.close()
        unit_label = input_prod['purchase_unit'] or 'units'
        return err(f"Insufficient stock for {input_prod['name']}. Available: {round(float(input_prod['current_stock']), 3)} {unit_label}")

    cnv_no = next_conversion_no(conn)
    cnv_date = d.get('conversion_date', str(date.today()))

    total_output_qty = 0.0
    validated_outputs = []

    for out in outputs:
        out_pid = out.get('output_product_id')
        try:
            out_qty = float(out.get('output_quantity', 0))
        except (ValueError, TypeError):
            out_qty = 0.0

        if not out_pid or out_qty <= 0:
            conn.close()
            return err("Each output item must have output_product_id and output_quantity > 0")

        out_prod = conn.execute('SELECT * FROM products WHERE id=? AND active=1', (out_pid,)).fetchone()
        if not out_prod:
            conn.close()
            return err(f"Output product ID {out_pid} not found or inactive", 404)

        input_unit = (input_prod['purchase_unit'] or 'kg').strip().lower()
        out_unit = (out_prod['purchase_unit'] or 'kg').strip().lower()

        equiv_qty = out_qty
        if input_unit != out_unit:
            conv = float(out_prod['conversion_factor'] or 1.0)
            if conv > 0 and input_unit in ('kg', 'liter', 'l') and out_unit in ('g', 'gram', 'ml'):
                equiv_qty = out_qty / 1000.0
            elif conv > 0 and input_unit in ('g', 'gram', 'ml') and out_unit in ('kg', 'liter', 'l'):
                equiv_qty = out_qty * 1000.0

        total_output_qty += equiv_qty
        validated_outputs.append((out_prod, out_qty))

    loss_qty = max(round(input_qty - total_output_qty, 3), 0.0)
    yield_pct = round((total_output_qty / input_qty) * 100.0, 2) if input_qty > 0 else 0.0

    total_input_cost = input_qty * float(input_prod['purchase_price'] or 0)

    update_stock(
        conn, input_pid, -input_qty, 'conversion_out',
        unit_price=float(input_prod['purchase_price'] or 0),
        ref=cnv_no, notes=f"Stock conversion input for {cnv_no}",
        created_by=session.get('username')
    )

    if loss_qty > 0:
        conn.execute('''
            INSERT INTO stock_transactions
            (product_id, type, quantity, unit_price, reference_id, notes, status, created_by)
            VALUES (?, 'wastage', ?, ?, ?, ?, 'approved', ?)
        ''', (input_pid, loss_qty, float(input_prod['purchase_price'] or 0), cnv_no, f"Processing loss for conversion {cnv_no}", session.get('username')))

    c = conn.execute('''
        INSERT INTO stock_conversions
        (conversion_no, conversion_date, input_product_id, input_quantity, yield_percent, loss_quantity, notes, created_by)
        VALUES (?,?,?,?,?,?,?,?)
    ''', (cnv_no, cnv_date, input_pid, input_qty, yield_pct, loss_qty, notes, session.get('username')))
    cnv_id = c.lastrowid

    for out_prod, out_qty in validated_outputs:
        out_pid = out_prod['id']
        input_unit = (input_prod['purchase_unit'] or 'kg').strip().lower()
        out_unit = (out_prod['purchase_unit'] or 'kg').strip().lower()

        equiv_qty = out_qty
        if input_unit != out_unit:
            conv = float(out_prod['conversion_factor'] or 1.0)
            if conv > 0 and input_unit in ('kg', 'liter', 'l') and out_unit in ('g', 'gram', 'ml'):
                equiv_qty = out_qty / 1000.0
            elif conv > 0 and input_unit in ('g', 'gram', 'ml') and out_unit in ('kg', 'liter', 'l'):
                equiv_qty = out_qty * 1000.0

        cost_share = (equiv_qty / total_output_qty) * total_input_cost if total_output_qty > 0 else 0.0
        alloc_unit_cost = round(cost_share / out_qty, 4) if out_qty > 0 else 0.0

        conn.execute('''
            INSERT INTO stock_conversion_outputs (conversion_id, output_product_id, output_quantity, allocated_unit_cost)
            VALUES (?,?,?,?)
        ''', (cnv_id, out_pid, out_qty, alloc_unit_cost))

        out_expiry = None
        shelf_life = out_prod.get('shelf_life_days') if hasattr(out_prod, 'keys') and 'shelf_life_days' in out_prod.keys() else None
        if shelf_life and int(shelf_life) > 0:
            try:
                base_dt = datetime.strptime(cnv_date[:10], '%Y-%m-%d')
                out_expiry = (base_dt + timedelta(days=int(shelf_life))).strftime('%Y-%m-%d')
            except Exception:
                out_expiry = None

        update_stock(
            conn, out_pid, out_qty, 'conversion_in',
            unit_price=alloc_unit_cost, unit_cost=alloc_unit_cost,
            ref=cnv_no, expiry_date=out_expiry,
            notes=f"Stock conversion output for {cnv_no}",
            created_by=session.get('username')
        )

    loss_val = round(loss_qty * float(input_prod['purchase_price'] or 0), 2)
    if loss_val > 0:
        loss_acc = conn.execute("SELECT id FROM ledger_accounts WHERE name='Inventory Loss/Processing'").fetchone()
        if not loss_acc:
            conn.execute('''
                INSERT INTO ledger_accounts (name, account_group, account_type, is_system)
                VALUES ('Inventory Loss/Processing', 'Expense', 'Direct Expense', 1)
            ''')
        entries = [
            {'account_name': 'Inventory Loss/Processing', 'debit': loss_val, 'credit': 0, 'narration': f"Processing loss for conversion {cnv_no}"},
            {'account_name': 'Purchase Account', 'debit': 0, 'credit': loss_val, 'narration': f"Inventory reduction for conversion loss {cnv_no}"}
        ]
        try:
            post_ledger_entry(
                conn,
                voucher_type='journal',
                voucher_no=cnv_no,
                voucher_date=cnv_date[:10],
                entries=entries,
                reference_table='stock_conversions',
                reference_id=cnv_id,
                created_by=session.get('username')
            )
        except Exception:
            pass

    conn.commit()

    cnv_row = conn.execute('''
        SELECT sc.*, p.name AS input_product_name, p.code AS input_product_code, p.purchase_unit AS input_unit
        FROM stock_conversions sc
        JOIN products p ON sc.input_product_id = p.id
        WHERE sc.id = ?
    ''', (cnv_id,)).fetchone()

    out_rows = conn.execute('''
        SELECT sco.*, p.name AS output_product_name, p.code AS output_product_code, p.purchase_unit AS output_unit
        FROM stock_conversion_outputs sco
        JOIN products p ON sco.output_product_id = p.id
        WHERE sco.conversion_id = ?
    ''', (cnv_id,)).fetchall()
    conn.close()

    res = dict_row(cnv_row)
    res['outputs'] = dict_rows(out_rows)
    return ok(res, f"Stock conversion {cnv_no} created successfully"), 201


@app.route('/api/stock/conversions', methods=['GET'])
@require_permission('inventory.view')
def list_stock_conversions():
    date_from = request.args.get('from', '')
    date_to   = request.args.get('to', '')
    input_pid = request.args.get('input_product_id', '')

    conn = get_db()
    sql = '''
        SELECT sc.*, p.name AS input_product_name, p.code AS input_product_code, p.purchase_unit AS input_unit
        FROM stock_conversions sc
        JOIN products p ON sc.input_product_id = p.id
        WHERE 1=1
    '''
    params = []
    if date_from:
        sql += ' AND date(sc.conversion_date) >= date(?)'
        params.append(date_from)
    if date_to:
        sql += ' AND date(sc.conversion_date) <= date(?)'
        params.append(date_to)
    if input_pid:
        sql += ' AND sc.input_product_id = ?'
        params.append(input_pid)

    sql += ' ORDER BY sc.conversion_date DESC, sc.id DESC'
    rows = conn.execute(sql, params).fetchall()

    conversions = dict_rows(rows)
    for cnv in conversions:
        out_rows = conn.execute('''
            SELECT sco.*, p.name AS output_product_name, p.code AS output_product_code, p.purchase_unit AS output_unit
            FROM stock_conversion_outputs sco
            JOIN products p ON sco.output_product_id = p.id
            WHERE sco.conversion_id = ?
        ''', (cnv['id'],)).fetchall()
        cnv['outputs'] = dict_rows(out_rows)

    conn.close()

    if wants_excel():
        excel_rows = []
        for cnv in conversions:
            in_qty = float(cnv.get('input_quantity') or 0)
            total_out_qty = sum(float(o.get('output_quantity') or 0) for o in cnv.get('outputs', []))
            loss_qty = round(in_qty - total_out_qty, 4)
            yield_pct = round((total_out_qty / in_qty * 100), 2) if in_qty > 0 else 0.0
            outputs_summary = ', '.join(
                f"{o.get('output_product_name','')} ({float(o.get('output_quantity') or 0)} {o.get('output_unit','')})" 
                for o in cnv.get('outputs', [])
            )
            excel_rows.append([
                cnv.get('conversion_no', ''),
                cnv.get('conversion_date', ''),
                cnv.get('input_product_name', ''),
                in_qty,
                cnv.get('input_unit', ''),
                outputs_summary,
                yield_pct,
                loss_qty,
                cnv.get('notes', '') or ''
            ])
        sheets = [{
            "sheet_name": "Stock Conversions",
            "headers": ["Conversion No", "Date", "Input Product", "Input Qty", "Input Unit", "Output Products", "Yield (%)", "Loss Qty", "Notes"],
            "rows": excel_rows
        }]
        f_from = date_from or "all"
        f_to   = date_to   or "all"
        return export_to_excel(sheets, f"stock_conversions_{f_from}_to_{f_to}")

    return ok(conversions)


@app.route('/api/stock/conversions/<int:cid>', methods=['GET'])
@require_permission('inventory.view')
def get_stock_conversion(cid):
    conn = get_db()
    row = conn.execute('''
        SELECT sc.*, p.name AS input_product_name, p.code AS input_product_code, p.purchase_unit AS input_unit
        FROM stock_conversions sc
        JOIN products p ON sc.input_product_id = p.id
        WHERE sc.id = ?
    ''', (cid,)).fetchone()

    if not row:
        conn.close()
        return err("Stock conversion not found", 404)

    out_rows = conn.execute('''
        SELECT sco.*, p.name AS output_product_name, p.code AS output_product_code, p.purchase_unit AS output_unit
        FROM stock_conversion_outputs sco
        JOIN products p ON sco.output_product_id = p.id
        WHERE sco.conversion_id = ?
    ''', (cid,)).fetchall()
    conn.close()

    res = dict_row(row)
    res['outputs'] = dict_rows(out_rows)
    return ok(res)


# ─── Stock Conversion Templates & Yield Variance ─────────────────────────

@app.route('/api/conversion-templates', methods=['GET'])
@require_permission('inventory.view')
def list_conversion_templates():
    input_pid = request.args.get('input_product_id', '')
    conn = get_db()
    sql = '''
        SELECT ct.*, p.name AS input_product_name, p.code AS input_product_code, p.purchase_unit AS input_unit
        FROM conversion_templates ct
        JOIN products p ON ct.input_product_id = p.id
        WHERE ct.is_active = 1
    '''
    params = []
    if input_pid:
        sql += ' AND ct.input_product_id = ?'
        params.append(input_pid)
    sql += ' ORDER BY ct.name ASC'

    rows = conn.execute(sql, params).fetchall()
    templates = dict_rows(rows)

    for t in templates:
        items = conn.execute('''
            SELECT cti.*, p.name AS output_product_name, p.code AS output_product_code, p.purchase_unit AS output_unit
            FROM conversion_template_items cti
            JOIN products p ON cti.output_product_id = p.id
            WHERE cti.template_id = ?
        ''', (t['id'],)).fetchall()
        t['items'] = dict_rows(items)

    conn.close()
    return ok(templates)


@app.route('/api/conversion-templates/<int:tid>', methods=['GET'])
@require_permission('inventory.view')
def get_conversion_template(tid):
    conn = get_db()
    t = conn.execute('''
        SELECT ct.*, p.name AS input_product_name, p.code AS input_product_code, p.purchase_unit AS input_unit
        FROM conversion_templates ct
        JOIN products p ON ct.input_product_id = p.id
        WHERE ct.id = ? AND ct.is_active = 1
    ''', (tid,)).fetchone()

    if not t:
        conn.close()
        return err("Conversion template not found", 404)

    items = conn.execute('''
        SELECT cti.*, p.name AS output_product_name, p.code AS output_product_code, p.purchase_unit AS output_unit
        FROM conversion_template_items cti
        JOIN products p ON cti.output_product_id = p.id
        WHERE cti.template_id = ?
    ''', (tid,)).fetchall()
    conn.close()

    res = dict_row(t)
    res['items'] = dict_rows(items)
    return ok(res)


@app.route('/api/conversion-templates', methods=['POST'])
@require_permission('stock.conversions')
def create_conversion_template():
    d = request.get_json()
    if d is None:
        return err("Invalid or missing JSON payload")

    name = (d.get('name') or '').strip()
    input_pid = d.get('input_product_id')
    items = d.get('items', [])

    if not name:
        return err("Template name is required")
    if not input_pid:
        return err("input_product_id is required")
    if not isinstance(items, list) or not items:
        return err("At least one template item is required in items")

    conn = get_db()
    input_prod = conn.execute('SELECT id FROM products WHERE id=? AND active=1', (input_pid,)).fetchone()
    if not input_prod:
        conn.close()
        return err("Input product not found or inactive", 404)

    c = conn.execute('''
        INSERT INTO conversion_templates (name, input_product_id, is_active)
        VALUES (?,?,1)
    ''', (name, input_pid))
    tid = c.lastrowid

    for it in items:
        out_pid = it.get('output_product_id')
        yield_pct = float(it.get('expected_yield_percent', 0))
        if out_pid:
            conn.execute('''
                INSERT INTO conversion_template_items (template_id, output_product_id, expected_yield_percent)
                VALUES (?,?,?)
            ''', (tid, out_pid, yield_pct))

    conn.commit()

    t = conn.execute('''
        SELECT ct.*, p.name AS input_product_name, p.code AS input_product_code
        FROM conversion_templates ct JOIN products p ON ct.input_product_id = p.id
        WHERE ct.id = ?
    ''', (tid,)).fetchone()
    t_items = conn.execute('''
        SELECT cti.*, p.name AS output_product_name, p.code AS output_product_code
        FROM conversion_template_items cti JOIN products p ON cti.output_product_id = p.id
        WHERE cti.template_id = ?
    ''', (tid,)).fetchall()
    conn.close()

    res = dict_row(t)
    res['items'] = dict_rows(t_items)
    return ok(res, "Conversion template created"), 201


@app.route('/api/conversion-templates/<int:tid>', methods=['PUT'])
@require_permission('stock.conversions')
def update_conversion_template(tid):
    d = request.get_json()
    if d is None:
        return err("Invalid or missing JSON payload")

    conn = get_db()
    t = conn.execute('SELECT * FROM conversion_templates WHERE id=? AND is_active=1', (tid,)).fetchone()
    if not t:
        conn.close()
        return err("Conversion template not found", 404)

    name = (d.get('name') or t['name']).strip()
    input_pid = d.get('input_product_id', t['input_product_id'])
    items = d.get('items')

    conn.execute('UPDATE conversion_templates SET name=?, input_product_id=? WHERE id=?', (name, input_pid, tid))

    if items is not None and isinstance(items, list):
        conn.execute('DELETE FROM conversion_template_items WHERE template_id=?', (tid,))
        for it in items:
            out_pid = it.get('output_product_id')
            yield_pct = float(it.get('expected_yield_percent', 0))
            if out_pid:
                conn.execute('''
                    INSERT INTO conversion_template_items (template_id, output_product_id, expected_yield_percent)
                    VALUES (?,?,?)
                ''', (tid, out_pid, yield_pct))

    conn.commit()

    res_t = conn.execute('''
        SELECT ct.*, p.name AS input_product_name, p.code AS input_product_code
        FROM conversion_templates ct JOIN products p ON ct.input_product_id = p.id
        WHERE ct.id = ?
    ''', (tid,)).fetchone()
    t_items = conn.execute('''
        SELECT cti.*, p.name AS output_product_name, p.code AS output_product_code
        FROM conversion_template_items cti JOIN products p ON cti.output_product_id = p.id
        WHERE cti.template_id = ?
    ''', (tid,)).fetchall()
    conn.close()

    res = dict_row(res_t)
    res['items'] = dict_rows(t_items)
    return ok(res, "Conversion template updated")


@app.route('/api/conversion-templates/<int:tid>', methods=['DELETE'])
@require_permission('stock.conversions')
def delete_conversion_template(tid):
    conn = get_db()
    t = conn.execute('SELECT * FROM conversion_templates WHERE id=? AND is_active=1', (tid,)).fetchone()
    if not t:
        conn.close()
        return err("Conversion template not found", 404)

    conn.execute('UPDATE conversion_templates SET is_active=0 WHERE id=?', (tid,))
    conn.commit()
    conn.close()
    return ok(message="Conversion template deleted")


@app.route('/api/conversion-templates/<int:tid>/prefill', methods=['GET'])
@require_permission('inventory.view')
def prefill_conversion_template(tid):
    try:
        input_qty = float(request.args.get('input_quantity', 0))
    except (ValueError, TypeError):
        input_qty = 0.0

    conn = get_db()
    t = conn.execute('''
        SELECT ct.*, p.name AS input_product_name, p.code AS input_product_code, p.purchase_unit AS input_unit
        FROM conversion_templates ct JOIN products p ON ct.input_product_id = p.id
        WHERE ct.id = ? AND ct.is_active = 1
    ''', (tid,)).fetchone()

    if not t:
        conn.close()
        return err("Conversion template not found", 404)

    items = conn.execute('''
        SELECT cti.*, p.name AS output_product_name, p.code AS output_product_code, p.purchase_unit AS output_unit
        FROM conversion_template_items cti JOIN products p ON cti.output_product_id = p.id
        WHERE cti.template_id = ?
    ''', (tid,)).fetchall()
    conn.close()

    suggested_outputs = []
    for it in items:
        y_pct = float(it['expected_yield_percent'] or 0)
        sugg_qty = round(input_qty * y_pct / 100.0, 3)
        suggested_outputs.append({
            'output_product_id': it['output_product_id'],
            'output_product_name': it['output_product_name'],
            'output_product_code': it['output_product_code'],
            'output_unit': it['output_unit'],
            'expected_yield_percent': y_pct,
            'suggested_output_quantity': sugg_qty
        })

    res = dict_row(t)
    res['suggested_outputs'] = suggested_outputs
    return ok(res)


@app.route('/api/reports/conversion-yield', methods=['GET'])
@require_permission('reports.view')
def conversion_yield_report():
    date_from = request.args.get('from', str(date.today().replace(day=1)))
    date_to   = request.args.get('to', str(date.today()))
    input_pid = request.args.get('input_product_id', '')
    try:
        threshold = float(request.args.get('variance_threshold', 5.0))
    except (ValueError, TypeError):
        threshold = 5.0

    conn = get_db()
    sql = '''
        SELECT sc.*, p.name AS input_product_name, p.code AS input_product_code, p.purchase_unit AS input_unit
        FROM stock_conversions sc
        JOIN products p ON sc.input_product_id = p.id
        WHERE date(sc.conversion_date) BETWEEN ? AND ?
    '''
    params = [date_from, date_to]
    if input_pid:
        sql += ' AND sc.input_product_id = ?'
        params.append(input_pid)

    sql += ' ORDER BY sc.conversion_date DESC, sc.id DESC'
    cnv_rows = conn.execute(sql, params).fetchall()

    conversions = dict_rows(cnv_rows)
    for cnv in conversions:
        template = conn.execute('''
            SELECT id, name FROM conversion_templates
            WHERE input_product_id = ? AND is_active = 1
            ORDER BY id DESC LIMIT 1
        ''', (cnv['input_product_id'],)).fetchone()

        template_items_map = {}
        if template:
            t_items = conn.execute('''
                SELECT output_product_id, expected_yield_percent
                FROM conversion_template_items WHERE template_id = ?
            ''', (template['id'],)).fetchall()
            template_items_map = {ti['output_product_id']: float(ti['expected_yield_percent'] or 0) for ti in t_items}

        out_rows = conn.execute('''
            SELECT sco.*, p.name AS output_product_name, p.code AS output_product_code, p.purchase_unit AS output_unit,
                   p.conversion_factor
            FROM stock_conversion_outputs sco
            JOIN products p ON sco.output_product_id = p.id
            WHERE sco.conversion_id = ?
        ''', (cnv['id'],)).fetchall()

        out_list = []
        has_variance_flag = False
        in_qty = float(cnv['input_quantity'] or 0)

        for out in out_rows:
            out_dict = dict_row(out)
            o_qty = float(out['output_quantity'] or 0)

            input_unit = (cnv['input_unit'] or 'kg').strip().lower()
            out_unit = (out['output_unit'] or 'kg').strip().lower()
            equiv_qty = o_qty
            if input_unit != out_unit:
                conv = float(out['conversion_factor'] or 1.0)
                if conv > 0 and input_unit in ('kg', 'liter', 'l') and out_unit in ('g', 'gram', 'ml'):
                    equiv_qty = o_qty / 1000.0
                elif conv > 0 and input_unit in ('g', 'gram', 'ml') and out_unit in ('kg', 'liter', 'l'):
                    equiv_qty = o_qty * 1000.0

            actual_yield_pct = round((equiv_qty / in_qty) * 100.0, 2) if in_qty > 0 else 0.0
            expected_yield_pct = template_items_map.get(out['output_product_id'], 0.0)
            variance = round(actual_yield_pct - expected_yield_pct, 2)
            is_flagged = abs(variance) >= threshold if template else False

            if is_flagged:
                has_variance_flag = True

            out_dict['actual_yield_percent'] = actual_yield_pct
            out_dict['expected_yield_percent'] = expected_yield_pct
            out_dict['variance_percent'] = variance
            out_dict['is_flagged'] = is_flagged
            out_list.append(out_dict)

        cnv['outputs'] = out_list
        cnv['template_name'] = template['name'] if template else None
        cnv['has_variance_flag'] = has_variance_flag

    conn.close()

    if wants_excel():
        rows = []
        for c in conversions:
            in_qty = float(c.get('input_quantity') or 0)
            for o in c.get('outputs', []):
                rows.append([
                    c.get('conversion_date', ''),
                    c.get('conversion_no', ''),
                    c.get('input_product_name', ''),
                    in_qty,
                    o.get('output_product_name', ''),
                    float(o.get('output_quantity') or 0),
                    float(o.get('actual_yield_percent') or 0),
                    float(o.get('expected_yield_percent') or 0),
                    float(o.get('variance_percent') or 0),
                    "FLAGGED" if o.get('is_flagged') else "NORMAL"
                ])
        sheets = [{
            "sheet_name": "Conversion Yield",
            "headers": ["Conversion Date", "Conversion No", "Input Product", "Input Qty", "Output Product", "Output Qty", "Actual Yield (%)", "Expected Yield (%)", "Variance (%)", "Status"],
            "rows": rows
        }]
        f_from = date_from or "start"
        f_to = date_to or "end"
        return export_to_excel(sheets, f"conversion_yield_{f_from}_to_{f_to}")

    return ok({
        "conversions": conversions,
        "variance_threshold": threshold,
        "from": date_from,
        "to": date_to
    })

# ─── Customers ──────────────────────────────────────────────────────────────

@app.route('/api/customers', methods=['GET'])
@require_permission('customers.view')
def list_customers():
    conn = get_db()
    q = request.args.get('q', '').strip()
    sql = '''
        SELECT c.*, COALESCE(SUM(b.amount_due), 0) AS total_dues
        FROM customers c
        LEFT JOIN bills b ON c.id = b.customer_id AND b.status != 'cancelled' AND b.amount_due > 0
        WHERE 1=1
    '''
    params = []
    if q:
        sql += ' AND (c.name LIKE ? OR c.phone LIKE ?)'
        params += [f'%{q}%', f'%{q}%']
    sql += ' GROUP BY c.id ORDER BY c.name'
    rows = conn.execute(sql, params).fetchall()
    conn.close()

    if wants_excel():
        excel_rows = []
        for r in dict_rows(rows):
            excel_rows.append([
                r['name'] or '',
                r['phone'] or '',
                r['gstin'] or '',
                r['state_code'] or '',
                r['email'] or '',
                r['address'] or '',
                float(r['total_dues'] or 0)
            ])
        sheets = [{
            "sheet_name": "Customers",
            "headers": ["Name", "Phone", "GSTIN", "State Code", "Email", "Address", "Total Dues (₹)"],
            "rows": excel_rows
        }]
        return export_to_excel(sheets, "customers_list")

    return ok(dict_rows(rows))

@app.route('/api/customers/<int:cid>', methods=['GET'])
@require_permission('customers.view')
def get_customer(cid):
    conn = get_db()
    row = conn.execute('SELECT * FROM customers WHERE id=?', (cid,)).fetchone()
    if not row: conn.close(); return err("Not found", 404)
    bills = conn.execute(
        'SELECT id, bill_no, date, grand_total, payment_mode, status FROM bills WHERE customer_id=? ORDER BY date DESC LIMIT 20',
        (cid,)
    ).fetchall()
    conn.close()
    result = dict_row(row)
    result['bills'] = dict_rows(bills)
    return ok(result)

@app.route('/api/customers/<int:cid>/dues', methods=['GET'])
@require_permission('customers.view')
def get_customer_dues(cid):
    conn = get_db()
    cust = conn.execute('SELECT * FROM customers WHERE id=?', (cid,)).fetchone()
    if not cust:
        conn.close()
        return err("Customer not found", 404)

    bills = conn.execute(
        'SELECT * FROM bills WHERE customer_id=? AND status != "cancelled" AND amount_due > 0 ORDER BY date DESC',
        (cid,)
    ).fetchall()
    total_due = conn.execute(
        'SELECT COALESCE(SUM(amount_due), 0) FROM bills WHERE customer_id=? AND status != "cancelled" AND amount_due > 0',
        (cid,)
    ).fetchone()[0]
    conn.close()

    res = dict_row(cust)
    res['due_bills'] = dict_rows(bills)
    res['total_due'] = round(total_due, 2)
    return ok(res)

@app.route('/api/customers', methods=['POST'])
@require_auth
def create_customer():
    d = request.get_json()
    if d is None:
        return err("Invalid or missing JSON payload")
    if not d.get('name'): return err("Name required")
    phone = (d.get('phone') or '').strip()
    if phone and (not phone.isdigit() or len(phone) != 10):
        return err("Phone number must be exactly 10 digits")
    conn = get_db()
    c = conn.execute(
        'INSERT INTO customers (name, phone, email, address, gstin, state_code) VALUES (?,?,?,?,?,?)',
        (d['name'], phone, d.get('email',''), d.get('address',''), d.get('gstin',''), d.get('state_code',''))
    )
    conn.commit()
    row = conn.execute('SELECT * FROM customers WHERE id=?', (c.lastrowid,)).fetchone()
    conn.close()
    return ok(dict_row(row)), 201

@app.route('/api/customers/<int:cid>', methods=['PUT'])
@require_permission('customers.manage')
def update_customer(cid):
    d = request.get_json()
    if d is None:
        return err("Invalid or missing JSON payload")
    if not d.get('name'): return err("Name required")
    phone = (d.get('phone') or '').strip()
    if phone and (not phone.isdigit() or len(phone) != 10):
        return err("Phone number must be exactly 10 digits")
    conn = get_db()
    conn.execute(
        'UPDATE customers SET name=?, phone=?, email=?, address=?, gstin=?, state_code=? WHERE id=?',
        (d['name'], phone, d.get('email',''), d.get('address',''), d.get('gstin',''), d.get('state_code',''), cid)
    )
    conn.commit()
    row = conn.execute('SELECT * FROM customers WHERE id=?', (cid,)).fetchone()
    conn.close()
    return ok(dict_row(row))

@app.route('/api/customers/<int:cid>', methods=['DELETE'])
@require_permission('customers.manage')
def delete_customer(cid):
    conn = get_db()
    conn.execute('DELETE FROM customers WHERE id=?', (cid,))
    conn.commit(); conn.close()
    return ok(message="Customer deleted")

# ─── Suppliers ──────────────────────────────────────────────────────────────

@app.route('/api/suppliers', methods=['GET'])
@require_permission('suppliers.view')
def list_suppliers():
    conn = get_db()
    q = request.args.get('q', '').strip()
    sql = '''
        SELECT s.*, COALESCE(SUM(po.amount_due), 0) AS total_payables
        FROM suppliers s
        LEFT JOIN purchase_orders po ON s.id = po.supplier_id AND po.amount_due > 0 AND (po.status IS NULL OR po.status != 'cancelled')
        WHERE 1=1
    '''
    params = []
    if q:
        sql += ' AND (s.name LIKE ? OR s.phone LIKE ?)'
        params += [f'%{q}%', f'%{q}%']
    sql += ' GROUP BY s.id ORDER BY s.name'
    rows = conn.execute(sql, params).fetchall()
    conn.close()

    if wants_excel():
        excel_rows = []
        for r in dict_rows(rows):
            excel_rows.append([
                r['name'] or '',
                r['contact_person'] or '',
                r['phone'] or '',
                r['gstin'] or '',
                r['email'] or '',
                r['address'] or '',
                float(r['total_payables'] or 0)
            ])
        sheets = [{
            "sheet_name": "Suppliers",
            "headers": ["Name", "Contact Person", "Phone", "GSTIN", "Email", "Address", "Total Payables (₹)"],
            "rows": excel_rows
        }]
        return export_to_excel(sheets, "suppliers_list")

    return ok(dict_rows(rows))

@app.route('/api/suppliers/<int:sid>', methods=['GET'])
@require_permission('suppliers.view')
def get_supplier(sid):
    conn = get_db()
    row = conn.execute('SELECT * FROM suppliers WHERE id=?', (sid,)).fetchone()
    conn.close()
    if not row: return err("Not found", 404)
    return ok(dict_row(row))

@app.route('/api/suppliers/<int:sid>/ledger', methods=['GET'])
@require_permission('suppliers.view')
def supplier_ledger(sid):
    conn = get_db()
    supplier = conn.execute('SELECT * FROM suppliers WHERE id=?', (sid,)).fetchone()
    if not supplier:
        conn.close()
        return err("Supplier not found", 404)

    pos = conn.execute(
        'SELECT * FROM purchase_orders WHERE supplier_id=? AND (status IS NULL OR status != "cancelled") ORDER BY date ASC',
        (sid,)
    ).fetchall()

    payments = conn.execute(
        '''SELECT pop.*, po.po_no
           FROM po_payments pop
           JOIN purchase_orders po ON pop.order_id = po.id
           WHERE po.supplier_id = ? AND (po.status IS NULL OR po.status != "cancelled")
           ORDER BY pop.paid_at ASC''',
        (sid,)
    ).fetchall()
    conn.close()

    entries = []
    for p in pos:
        entries.append({
            'id': f"PO-{p['id']}",
            'date': p['date'],
            'type': 'purchase_order',
            'reference_no': p['po_no'],
            'description': f"Purchase Order #{p['po_no']}",
            'debit': round(float(p['total']), 2),
            'credit': 0.0
        })

    for pop in payments:
        entries.append({
            'id': f"PMT-{pop['id']}",
            'date': pop['paid_at'],
            'type': 'payment',
            'reference_no': pop['po_no'],
            'description': f"Payment via {pop['payment_mode']} ({pop['notes'] or ''})".strip(),
            'debit': 0.0,
            'credit': round(float(pop['amount']), 2)
        })

    entries.sort(key=lambda x: str(x['date']))

    running_balance = 0.0
    for entry in entries:
        running_balance += (entry['debit'] - entry['credit'])
        entry['running_balance'] = round(running_balance, 2)

    total_purchased = round(sum(e['debit'] for e in entries), 2)
    total_paid = round(sum(e['credit'] for e in entries), 2)
    outstanding_payable = round(total_purchased - total_paid, 2)

    return ok({
        "supplier": dict_row(supplier),
        "summary": {
            "total_purchased": total_purchased,
            "total_paid": total_paid,
            "outstanding_payable": outstanding_payable
        },
        "ledger": entries
    })

@app.route('/api/suppliers', methods=['POST'])
@require_permission('suppliers.manage')
def create_supplier():
    d = request.get_json()
    if d is None:
        return err("Invalid or missing JSON payload")
    if not d.get('name'): return err("Name required")
    phone = (d.get('phone') or '').strip()
    if phone and (not phone.isdigit() or len(phone) != 10):
        return err("Phone number must be exactly 10 digits")
    conn = get_db()
    c = conn.execute(
        'INSERT INTO suppliers (name, contact_person, phone, email, address, gstin) VALUES (?,?,?,?,?,?)',
        (d['name'], d.get('contact_person',''), phone,
         d.get('email',''), d.get('address',''), d.get('gstin',''))
    )
    conn.commit()
    row = conn.execute('SELECT * FROM suppliers WHERE id=?', (c.lastrowid,)).fetchone()
    conn.close()
    return ok(dict_row(row)), 201

@app.route('/api/suppliers/<int:sid>', methods=['PUT'])
@require_permission('suppliers.manage')
def update_supplier(sid):
    d = request.get_json()
    if d is None:
        return err("Invalid or missing JSON payload")
    if not d.get('name'): return err("Name required")
    phone = (d.get('phone') or '').strip()
    if phone and (not phone.isdigit() or len(phone) != 10):
        return err("Phone number must be exactly 10 digits")
    conn = get_db()
    conn.execute(
        'UPDATE suppliers SET name=?, contact_person=?, phone=?, email=?, address=?, gstin=? WHERE id=?',
        (d['name'], d.get('contact_person',''), phone,
         d.get('email',''), d.get('address',''), d.get('gstin',''), sid)
    )
    conn.commit()
    row = conn.execute('SELECT * FROM suppliers WHERE id=?', (sid,)).fetchone()
    conn.close()
    return ok(dict_row(row))

@app.route('/api/suppliers/<int:sid>', methods=['DELETE'])
@require_permission('suppliers.manage')
def delete_supplier(sid):
    conn = get_db()
    conn.execute('DELETE FROM suppliers WHERE id=?', (sid,))
    conn.commit(); conn.close()
    return ok(message="Supplier deleted")

# ─── Stock ──────────────────────────────────────────────────────────────────

@app.route('/api/stock/transactions', methods=['GET'])
@require_permission('inventory.view')
def stock_transactions():
    conn = get_db()
    pid = request.args.get('product_id')
    tx_type = request.args.get('type')
    date_from = request.args.get('from')
    date_to   = request.args.get('to')
    limit = int(request.args.get('limit', 500))
    sql = '''
        SELECT st.*, p.name AS product_name, s.name AS supplier_name
        FROM stock_transactions st
        LEFT JOIN products p ON st.product_id=p.id
        LEFT JOIN suppliers s ON st.supplier_id=s.id
        WHERE 1=1
    '''
    params = []
    if pid:
        sql += ' AND st.product_id=?'; params.append(pid)
    if tx_type:
        sql += ' AND st.type=?'; params.append(tx_type)
    if date_from:
        sql += ' AND date(st.date) >= ?'; params.append(date_from)
    if date_to:
        sql += ' AND date(st.date) <= ?'; params.append(date_to)
    sql += ' ORDER BY st.date DESC LIMIT ?'
    params.append(limit)
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return ok(dict_rows(rows))


@app.route('/api/stock/summary', methods=['GET'])
@require_permission('inventory.view')
def stock_summary():
    """
    Returns all products with period-based inward / outward totals for the
    Inventory Stock Summary Register.
    Query params: from=YYYY-MM-DD  to=YYYY-MM-DD  (defaults = current month)
    """
    from datetime import date
    today = date.today()
    date_from = request.args.get('from', today.strftime('%Y-%m-01'))
    date_to   = request.args.get('to',   today.strftime('%Y-%m-%d'))

    conn = get_db()

    # All active products with category (correct column: active, not is_active)
    products = conn.execute('''
        SELECT p.*, c.name AS category_name
        FROM products p
        LEFT JOIN categories c ON p.category_id = c.id
        WHERE p.active = 1
        ORDER BY c.name, p.name
    ''').fetchall()

    # Determine valid approved statuses for this DB
    # stock_transactions.status can be: 'approved', 'pending', NULL — include approved + NULL
    movements = conn.execute('''
        SELECT product_id, type, SUM(quantity) AS total_qty
        FROM stock_transactions
        WHERE date(date) BETWEEN ? AND ?
          AND (status = 'approved' OR status IS NULL OR status = '')
        GROUP BY product_id, type
    ''', (date_from, date_to)).fetchall()
    conn.close()

    # Build lookup dict keyed by product_id
    move_map = {}
    for m in movements:
        pid = m['product_id']
        if pid not in move_map:
            move_map[pid] = {'in': 0.0, 'out': 0.0, 'wastage': 0.0}
        t = m['type'] or 'in'
        move_map[pid][t] = float(m['total_qty'] or 0)

    result = []
    for p in products:
        pd = dict_row(p)
        pm = move_map.get(p['id'], {})
        inward  = pm.get('in', 0.0)
        outward = pm.get('out', 0.0) + pm.get('wastage', 0.0)
        closing = float(p['current_stock'] or 0)
        opening = closing - inward + outward   # derived
        value   = closing * float(p['purchase_price'] or 0)
        pd['period_from'] = date_from
        pd['period_to']   = date_to
        pd['inward']      = round(inward,  3)
        pd['outward']     = round(outward, 3)
        pd['closing']     = round(closing, 3)
        pd['opening']     = round(opening, 3)
        pd['stock_value'] = round(value,   2)
        result.append(pd)

    return ok(result)

@app.route('/api/stock/in', methods=['POST'])
@require_permission('stock.in')
def stock_in():
    d = request.get_json()
    if d is None:
        return err("Invalid or missing JSON payload")
    if not d.get('product_id'): return err("product_id required")
    try:
        qty = float(d.get('quantity', 0))
    except (ValueError, TypeError):
        qty = 0.0
    if qty <= 0: return err("Valid quantity required")

    role = session.get('user_role')
    username = session.get('username')
    status = 'approved' if role in ('admin', 'manager') else 'pending_verification'
    approved_by = username if role in ('admin', 'manager') else None

    conn = get_db()
    update_stock(conn, d['product_id'], qty, 'in',
                 float(d.get('unit_price', 0)), d.get('reference'),
                 d.get('supplier_id'), d.get('expiry_date'), d.get('notes'),
                 status=status, created_by=username, approved_by=approved_by)

    if status == 'pending_verification':
        prod = conn.execute('SELECT name FROM products WHERE id=?', (d['product_id'],)).fetchone()
        pname = prod['name'] if prod else 'Product'
        msg = f"Accountant @{username} submitted Stock-In of {qty} units for {pname}. Pending Manager / MD verification."
        conn.execute('INSERT INTO notifications (target_role, title, message) VALUES (?,?,?)',
                     ('manager', '📦 Stock-In Pending Verification', msg))
        conn.execute('INSERT INTO notifications (target_role, title, message) VALUES (?,?,?)',
                     ('admin', '📦 Stock-In Pending Verification', msg))
        conn.commit()
        conn.close()
        return ok(message="Stock entry submitted! Pending approval by Manager or Managing Director.")

    conn.commit()
    row = conn.execute('SELECT * FROM products WHERE id=?', (d['product_id'],)).fetchone()
    conn.close()
    return ok(dict_row(row), "Stock updated and approved")

@app.route('/api/stock/pending', methods=['GET'])
@require_permission('stock.verify')
def list_pending_stock():
    conn = get_db()
    rows = conn.execute('''
        SELECT st.*, p.name AS product_name, p.unit, p.code AS product_code, s.name AS supplier_name
        FROM stock_transactions st
        JOIN products p ON st.product_id = p.id
        LEFT JOIN suppliers s ON st.supplier_id = s.id
        WHERE st.status = 'pending_verification'
        ORDER BY st.date DESC
    ''').fetchall()
    conn.close()
    return ok(dict_rows(rows))

@app.route('/api/stock/verify/<int:tx_id>', methods=['POST'])
@require_permission('stock.verify')
def verify_stock(tx_id):
    d = request.get_json() or {}
    action = d.get('action')
    if action not in ('approve', 'reject'):
        return err("Action must be 'approve' or 'reject'")

    conn = get_db()
    tx = conn.execute('SELECT * FROM stock_transactions WHERE id=?', (tx_id,)).fetchone()
    if not tx: conn.close(); return err("Stock transaction not found", 404)
    if tx['status'] != 'pending_verification':
        conn.close(); return err("Transaction is already processed")

    approver = session.get('username')
    prod = conn.execute('SELECT name FROM products WHERE id=?', (tx['product_id'],)).fetchone()
    pname = prod['name'] if prod else 'Product'

    if action == 'approve':
        conn.execute('UPDATE stock_transactions SET status="approved", approved_by=? WHERE id=?', (approver, tx_id))
        conn.execute('UPDATE products SET current_stock = current_stock + ?, updated_at = CURRENT_TIMESTAMP WHERE id=?',
                     (tx['quantity'], tx['product_id']))
        if tx['type'] in ('in', 'adjustment') and tx['quantity'] > 0:
            b_no = f"BATCH-{tx_id:05d}"
            conn.execute(
                '''INSERT INTO stock_batches
                   (product_id, batch_no, quantity_remaining, unit_price, expiry_date, supplier_id, stock_transaction_id)
                   VALUES (?,?,?,?,?,?,?)''',
                (tx['product_id'], b_no, tx['quantity'], tx['unit_price'], tx['expiry_date'], tx['supplier_id'], tx_id)
            )
        msg = f"Stock-In for {pname} ({tx['quantity']} units) approved by @{approver}."
    else:
        conn.execute('UPDATE stock_transactions SET status="rejected", approved_by=? WHERE id=?', (approver, tx_id))
        msg = f"Stock-In for {pname} ({tx['quantity']} units) rejected by @{approver}."

    conn.commit(); conn.close()
    return ok(message=msg)

@app.route('/api/stock/wastage', methods=['POST'])
@require_permission('stock.wastage')
def stock_wastage():
    d = request.get_json()
    if d is None:
        return err("Invalid or missing JSON payload")
    if not d.get('product_id'): return err("product_id required")
    try:
        qty = float(d.get('quantity', 0))
    except (ValueError, TypeError):
        qty = 0.0
    if qty <= 0: return err("Valid quantity required")
    
    notes = (d.get('notes') or d.get('reason') or '').strip()
    if not notes:
        return err("Wastage notes / reason is required")

    conn = get_db()
    prod = conn.execute(
        'SELECT current_stock, conversion_factor, purchase_unit, sale_unit, name FROM products WHERE id=?',
        (d['product_id'],)
    ).fetchone()
    if not prod:
        conn.close()
        return err("Product not found", 404)

    conv = float(prod['conversion_factor'] or 1.0)
    unit_mode = d.get('unit_mode', 'sale_unit')
    if unit_mode == 'purchase_unit':
        qty_purchase = round(qty, 4)
        qty_sale = round(qty * conv, 3)
    else:
        qty_sale = round(qty, 3)
        qty_purchase = round(qty / conv, 4)

    if prod['current_stock'] < qty_purchase:
        conn.close()
        avail_sale = round(prod['current_stock'] * conv, 3)
        s_unit = prod['sale_unit'] or 'units'
        return err(f"Insufficient stock for wastage entry of {prod['name']}. Available: {avail_sale} {s_unit}")

    update_stock(conn, d['product_id'], -qty_purchase, 'wastage', notes=notes)
    conn.commit()
    row = conn.execute('SELECT * FROM products WHERE id=?', (d['product_id'],)).fetchone()
    conn.close()
    return ok(dict_row(row), f"Wastage of {qty_sale} {prod['sale_unit'] or 'units'} recorded")

@app.route('/api/stock/adjustment', methods=['POST'])
@require_permission('stock.adjustment')
def stock_adjustment():
    d = request.get_json()
    if d is None:
        return err("Invalid or missing JSON payload")
    if not d.get('product_id'): return err("product_id required")
    if 'new_quantity' not in d: return err("new_quantity required")
    try:
        new_qty = float(d['new_quantity'])
    except (ValueError, TypeError):
        return err("Valid new_quantity required")

    conn = get_db()
    prod = conn.execute('SELECT current_stock FROM products WHERE id=?', (d['product_id'],)).fetchone()
    if not prod: conn.close(); return err("Product not found", 404)
    delta = new_qty - prod['current_stock']
    update_stock(conn, d['product_id'], delta, 'adjustment', notes=d.get('notes', 'Manual adjustment'))
    conn.commit()
    row = conn.execute('SELECT * FROM products WHERE id=?', (d['product_id'],)).fetchone()
# ─── Hold & Recall Bills ───────────────────────────────────────────────────

@app.route('/billing/hold', methods=['POST'])
@app.route('/api/billing/hold', methods=['POST'])
@require_permission('billing.hold')
def hold_bill():
    d = request.get_json() or {}
    items = d.get('items') or d.get('cart') or []
    if not items or len(items) == 0:
        return err("Cart is empty. Add products before holding.", 400)

    terminal_id = d.get('terminal_id', 'POS-1')
    cashier_id = session.get('user_id')
    customer_id = d.get('customer_id') or (d.get('customer', {}).get('id') if isinstance(d.get('customer'), dict) else None)
    customer_name = d.get('customer_name') or (d.get('customer', {}).get('name') if isinstance(d.get('customer'), dict) else None)
    discount_pct = float(d.get('discount_percent') or d.get('discountPct') or 0)
    payment_mode = d.get('payment_mode') or d.get('paymentMode') or 'cash'
    notes = d.get('notes', '')
    total = float(d.get('total') or d.get('total_amount') or 0)

    items_json = json.dumps(items)

    conn = get_db()
    c = conn.execute(
        '''INSERT INTO held_bills
           (terminal_id, cashier_user_id, customer_id, customer_name, items_json,
            discount_percent, payment_mode, total_amount, notes, status)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'held')''',
        (terminal_id, cashier_id, customer_id, customer_name, items_json,
         discount_pct, payment_mode, total, notes)
    )
    held_id = c.lastrowid
    ref_code = f"HOLD-{held_id:04d}"
    conn.execute('UPDATE held_bills SET reference_code = ? WHERE id = ?', (ref_code, held_id))
    conn.commit()
    conn.close()

    # Clear server-side cart if tracked in session
    session.pop('cart', None)

    log_activity('hold_bill', f"Held bill {ref_code} ({len(items)} items, Total ₹{total})", 'held_bills', held_id)

    return ok({
        'id': held_id,
        'reference_code': ref_code,
        'reference': ref_code,
        'terminal_id': terminal_id,
        'total': total,
        'item_count': len(items)
    }, f"Bill held successfully as {ref_code}")


@app.route('/billing/held', methods=['GET'])
@app.route('/api/billing/held', methods=['GET'])
@require_permission('billing.hold')
def list_held_bills():
    terminal_id = request.args.get('terminal_id')
    conn = get_db()
    sql = '''
        SELECT hb.*, u.full_name AS cashier_name, c.name AS customer_master_name, c.phone AS customer_phone
        FROM held_bills hb
        LEFT JOIN users u ON hb.cashier_user_id = u.id
        LEFT JOIN customers c ON hb.customer_id = c.id
        WHERE hb.status = 'held'
    '''
    params = []
    if terminal_id:
        sql += ' AND hb.terminal_id = ?'
        params.append(terminal_id)

    sql += ' ORDER BY hb.id DESC'
    rows = conn.execute(sql, params).fetchall()
    conn.close()

    held_bills = []
    for r in rows:
        hb = dict_row(r)
        try:
            items = json.loads(hb['items_json']) if hb.get('items_json') else []
        except Exception:
            items = []
        item_count = sum(float(i.get('quantity', 1)) for i in items) if isinstance(items, list) else 0
        hb['items'] = items
        hb['item_count'] = item_count
        hb['reference'] = hb.get('reference_code') or f"HOLD-{hb['id']:04d}"
        hb['time_held'] = hb.get('created_at')
        hb['total'] = float(hb.get('total_amount') or 0)
        held_bills.append(hb)

    return ok(held_bills)


@app.route('/billing/recall/<int:bill_id>', methods=['POST'])
@app.route('/api/billing/recall/<int:bill_id>', methods=['POST'])
@require_permission('billing.hold')
def recall_held_bill(bill_id):
    conn = get_db()
    bill = conn.execute('SELECT * FROM held_bills WHERE id = ? AND status = "held"', (bill_id,)).fetchone()
    if not bill:
        conn.close()
        return err("Held bill not found or already recalled", 404)

    conn.execute('UPDATE held_bills SET status = "recalled" WHERE id = ?', (bill_id,))
    conn.commit()
    conn.close()

    b_dict = dict_row(bill)
    try:
        items = json.loads(b_dict['items_json'])
    except Exception:
        items = []

    session['cart'] = items
    ref = b_dict.get('reference_code') or f"HOLD-{bill_id:04d}"
    log_activity('recall_held_bill', f"Recalled held bill {ref}", 'held_bills', bill_id)

    return ok({
        'id': bill_id,
        'reference_code': ref,
        'reference': ref,
        'cart': items,
        'items': items,
        'customer_id': b_dict.get('customer_id'),
        'customer_name': b_dict.get('customer_name'),
        'discount_percent': float(b_dict.get('discount_percent') or 0),
        'payment_mode': b_dict.get('payment_mode') or 'cash',
        'notes': b_dict.get('notes') or '',
        'total': float(b_dict.get('total_amount') or 0)
    }, f"Bill {ref} recalled successfully")


@app.route('/billing/held/<int:bill_id>', methods=['DELETE'])
@app.route('/api/billing/held/<int:bill_id>', methods=['DELETE'])
@require_permission('billing.delete_held')
def delete_held_bill(bill_id):
    conn = get_db()
    bill = conn.execute('SELECT * FROM held_bills WHERE id = ?', (bill_id,)).fetchone()
    if not bill:
        conn.close()
        return err("Held bill not found", 404)

    ref = bill['reference_code'] or f"HOLD-{bill_id:04d}"
    conn.execute('DELETE FROM held_bills WHERE id = ?', (bill_id,))
    conn.commit()
    conn.close()

    log_activity('delete_held_bill', f"Deleted held bill {ref}", 'held_bills', bill_id)
    return ok(message=f"Held bill {ref} deleted successfully")


# ─── Bills ──────────────────────────────────────────────────────────────────

@app.route('/api/bills', methods=['GET'])
@require_permission('billing.view')
def list_bills():
    conn = get_db()
    date_from = request.args.get('from', '')
    date_to   = request.args.get('to', '')
    search    = request.args.get('q', '')
    status_param = request.args.get('status', '').strip().lower()
    limit     = int(request.args.get('limit', 50))
    offset    = int(request.args.get('offset', 0))

    if session.get('user_role') == 'tester':
        sql = 'SELECT * FROM bills WHERE is_test=1'
    else:
        sql = 'SELECT * FROM bills WHERE (is_test IS NULL OR is_test=0)'

    params = []
    if status_param:
        sql += ' AND LOWER(status) = ?'; params.append(status_param)
    else:
        sql += ' AND status != "cancelled"'

    if date_from:
        sql += ' AND date(date) >= ?'; params.append(date_from)
    if date_to:
        sql += ' AND date(date) <= ?'; params.append(date_to)
    if search:
        sql += ' AND (bill_no LIKE ? OR customer_name LIKE ? OR customer_phone LIKE ?)'
        params += [f'%{search}%', f'%{search}%', f'%{search}%']

    sql_order = ' ORDER BY date DESC, id DESC'

    if wants_excel():
        rows = conn.execute(sql + sql_order, params).fetchall()
        conn.close()
        excel_rows = []
        for b in rows:
            excel_rows.append([
                b['bill_no'],
                b['date'],
                b['customer_name'] or 'Walk-in',
                b['customer_phone'] or '',
                float(b['subtotal'] or 0),
                float(b['cgst'] or 0),
                float(b['sgst'] or 0),
                float(b['igst'] or 0),
                float(b['discount_amount'] or 0),
                float(b['grand_total'] or 0),
                float(b['amount_paid'] or 0),
                float(b['amount_due'] or 0),
                b['payment_mode'] or '',
                b['status'] or ''
            ])
        sheets = [{
            "sheet_name": "Bills",
            "headers": ["Bill No", "Date", "Customer Name", "Customer Phone", "Subtotal (₹)", "CGST (₹)", "SGST (₹)", "IGST (₹)", "Discount (₹)", "Grand Total (₹)", "Amount Paid (₹)", "Amount Due (₹)", "Payment Mode", "Status"],
            "rows": excel_rows
        }]
        return export_to_excel(sheets, "bills_list")

    total = conn.execute('SELECT COUNT(*) FROM bills WHERE status != "cancelled"').fetchone()[0]
    rows = conn.execute(sql + sql_order + ' LIMIT ? OFFSET ?', params + [limit, offset]).fetchall()
    conn.close()
    return ok({"bills": dict_rows(rows), "total": total})

@app.route('/api/bills/<int:bid>', methods=['GET'])
@require_permission('billing.view')
def get_bill(bid):
    conn = get_db()
    bill = conn.execute('SELECT * FROM bills WHERE id=?', (bid,)).fetchone()
    if not bill: conn.close(); return err("Bill not found", 404)
    items    = conn.execute('SELECT * FROM bill_items WHERE bill_id=?', (bid,)).fetchall()
    payments = conn.execute('SELECT * FROM bill_payments WHERE bill_id=? ORDER BY paid_at ASC', (bid,)).fetchall()
    settings = {r['key']: r['value'] for r in conn.execute('SELECT key,value FROM shop_settings').fetchall()}
    conn.close()
    result = dict_row(bill)
    result['items']    = dict_rows(items)
    result['payments'] = dict_rows(payments)
    result['settings'] = settings
    return ok(result)

@app.route('/api/bills/<int:bid>/share-link', methods=['GET'])
@require_permission('billing.view')
def bill_share_link(bid):
    conn = get_db()
    bill = conn.execute('SELECT * FROM bills WHERE id=?', (bid,)).fetchone()
    if not bill:
        conn.close()
        return err("Bill not found", 404)

    phone = (bill['customer_phone'] or '').strip()
    digits = ''.join(c for c in phone if c.isdigit())
    if not digits:
        conn.close()
        return err("No phone number on this bill — add one before sharing.", 400)

    if len(digits) == 10:
        phone_formatted = "91" + digits
    else:
        phone_formatted = digits

    items = conn.execute('SELECT product_name, quantity, unit, amount FROM bill_items WHERE bill_id=?', (bid,)).fetchall()
    shop_name = get_setting('shop_name', conn) or 'Meat Products of India'
    conn.close()

    invoice_url = f"{request.host_url.rstrip('/')}/invoice/{bill['id']}"

    item_lines = []
    for it in items:
        unit_str = f" {it['unit']}" if it['unit'] else ""
        item_lines.append(f"• {it['product_name']} x {it['quantity']}{unit_str}")
    items_summary = "\n".join(item_lines)

    msg = (
        f"*{shop_name}*\n"
        f"Bill No: {bill['bill_no']}\n"
        f"Date: {bill['date']}\n\n"
        f"Items:\n{items_summary}\n\n"
        f"Grand Total: ₹{bill['grand_total']:.2f}\n\n"
        f"View invoice: {invoice_url}\n\n"
        f"Thank you for shopping with us!"
    )

    wa_url = f"https://wa.me/{phone_formatted}?text={urllib.parse.quote(msg)}"
    return ok({
        "whatsapp_url": wa_url,
        "phone": phone_formatted,
        "message": msg
    })

@app.route('/api/bills', methods=['POST'])
@require_permission('billing.create')
def create_bill():
    lic = get_license_info()
    if lic.get('is_locked'):
        return err("Subscription expired and grace period ended. Please activate software with a 12-digit key to create bills.", 403)

    d = request.get_json()
    if d is None:
        return err("Invalid or missing JSON payload")
    items = d.get('items', [])
    if not isinstance(items, list) or not items: return err("Bill must have at least one item")
    conn = get_db()
    bill_no = next_bill_no(conn)
    
    # Determine place of supply and inter-state status
    shop_state_code = (get_setting('shop_state_code', conn) or '').strip()
    cust_state_code = (d.get('place_of_supply') or d.get('state_code') or '').strip()
    if not cust_state_code and d.get('customer_id'):
        cust = conn.execute('SELECT state_code FROM customers WHERE id=?', (d['customer_id'],)).fetchone()
        if cust and cust['state_code']:
            cust_state_code = cust['state_code'].strip()
            
    place_of_supply = cust_state_code
    is_interstate = 1 if (shop_state_code and place_of_supply and shop_state_code != place_of_supply) else 0

    subtotal = 0; cgst_total = 0; sgst_total = 0; igst_total = 0
    # Validate stock
    for it in items:
        if it.get('product_id'):
            prod = conn.execute(
                'SELECT current_stock, conversion_factor, sale_unit, name FROM products WHERE id=?', (it['product_id'],)
            ).fetchone()
            if prod:
                conv = float(prod['conversion_factor'] or 1.0)
                qty_sale = float(it['quantity'])
                qty_purchase = round(qty_sale / conv, 4)
                if prod['current_stock'] < qty_purchase:
                    conn.close()
                    avail_sale_units = round(prod['current_stock'] * conv, 3)
                    unit_label = prod['sale_unit'] or 'units'
                    return err(f"Insufficient stock for {prod['name']}. Available: {avail_sale_units} {unit_label}")
    discount_pct = float(d.get('discount_percent', 0))
    raw_subtotal = sum(float(it['quantity']) * float(it['unit_price']) for it in items)
    discount_amt = round(raw_subtotal * discount_pct / 100, 2)
    for it in items:
        qty = float(it['quantity']); price = float(it['unit_price']); gst_rate = float(it.get('gst_rate', 0))
        prod_row = None
        if it.get('product_id'):
            prod_row = conn.execute('SELECT product_type, is_price_inclusive_of_tax FROM products WHERE id=?', (it['product_id'],)).fetchone()
        is_inc_tax = (prod_row and prod_row['product_type'] == 'general' and int(prod_row['is_price_inclusive_of_tax'] or 0) == 1)

        if is_inc_tax:
            item_gross = round(qty * price * (1 - discount_pct / 100), 2)
            if gst_rate > 0:
                item_taxable = round(item_gross / (1.0 + gst_rate / 100.0), 2)
                total_tax = round(item_gross - item_taxable, 2)
            else:
                item_taxable = item_gross
                total_tax = 0.0

            if is_interstate:
                cgst_amt = 0.0
                sgst_amt = 0.0
                igst_amt = total_tax
            else:
                cgst_amt = round(total_tax / 2.0, 2)
                sgst_amt = round(total_tax - cgst_amt, 2)
                igst_amt = 0.0
        else:
            item_taxable = round(qty * price * (1 - discount_pct / 100), 2)
            if is_interstate:
                cgst_amt = 0.0
                sgst_amt = 0.0
                igst_amt = round(item_taxable * gst_rate / 100, 2)
            else:
                cgst_amt = round(item_taxable * gst_rate / 2 / 100, 2)
                sgst_amt = round(item_taxable * gst_rate / 2 / 100, 2)
                igst_amt = 0.0

        cgst_total += cgst_amt
        sgst_total += sgst_amt
        igst_total += igst_amt
        subtotal   += item_taxable
    subtotal = round(subtotal, 2)
    cgst_total = round(cgst_total, 2)
    sgst_total = round(sgst_total, 2)
    igst_total = round(igst_total, 2)
    grand_total  = round(subtotal + cgst_total + sgst_total + igst_total, 2)

    if 'amount_paid' in d and d['amount_paid'] is not None:
        raw_paid = float(d['amount_paid'])
    else:
        raw_paid = grand_total

    if raw_paid >= grand_total:
        amount_paid = grand_total
        amount_due = 0.0
        change_amount = round(raw_paid - grand_total, 2)
        status = 'paid'
    elif raw_paid <= 0:
        amount_paid = 0.0
        amount_due = grand_total
        change_amount = 0.0
        status = 'due'
    else:
        amount_paid = round(raw_paid, 2)
        amount_due = round(grand_total - amount_paid, 2)
        change_amount = 0.0
        status = 'partial'

    is_test = 1 if session.get('user_role') == 'tester' else 0
    c = conn.execute(
        '''INSERT INTO bills
           (bill_no, customer_id, customer_name, customer_phone, customer_gstin,
            place_of_supply, is_interstate,
            subtotal, discount_percent, discount_amount, cgst, sgst, igst, grand_total,
            amount_paid, amount_due, change_amount, payment_mode, notes, is_test, status)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (bill_no, d.get('customer_id'), d.get('customer_name', 'Walk-in Customer'),
         d.get('customer_phone', ''), d.get('customer_gstin', ''),
         place_of_supply, is_interstate,
         round(subtotal, 2), discount_pct, discount_amt,
         round(cgst_total, 2), round(sgst_total, 2), round(igst_total, 2), grand_total,
         amount_paid, amount_due, max(change_amount, 0), d.get('payment_mode', 'cash'),
         d.get('notes', ''), is_test, status)
    )
    bill_id = c.lastrowid

    if amount_paid > 0:
        conn.execute(
            '''INSERT INTO bill_payments (bill_id, amount, payment_mode, received_by, notes)
               VALUES (?,?,?,?,?)''',
            (bill_id, amount_paid, d.get('payment_mode', 'cash'), session.get('username'), 'Initial payment at billing')
        )

    for it in items:
        qty = float(it['quantity']); price = float(it['unit_price']); gst_rate = float(it.get('gst_rate', 0))
        prod_row = None
        if it.get('product_id'):
            prod_row = conn.execute('SELECT product_type, is_price_inclusive_of_tax FROM products WHERE id=?', (it['product_id'],)).fetchone()
        is_inc_tax = (prod_row and prod_row['product_type'] == 'general' and int(prod_row['is_price_inclusive_of_tax'] or 0) == 1)

        if is_inc_tax:
            item_gross = round(qty * price * (1 - discount_pct / 100), 2)
            if gst_rate > 0:
                item_taxable = round(item_gross / (1.0 + gst_rate / 100.0), 2)
                total_tax = round(item_gross - item_taxable, 2)
            else:
                item_taxable = item_gross
                total_tax = 0.0

            if is_interstate:
                cgst_amt = 0.0
                sgst_amt = 0.0
                igst_amt = total_tax
            else:
                cgst_amt = round(total_tax / 2.0, 2)
                sgst_amt = round(total_tax - cgst_amt, 2)
                igst_amt = 0.0

            amount = item_gross
        else:
            item_taxable = round(qty * price * (1 - discount_pct / 100), 2)
            if is_interstate:
                cgst_amt = 0.0
                sgst_amt = 0.0
                igst_amt = round(item_taxable * gst_rate / 100, 2)
            else:
                cgst_amt = round(item_taxable * gst_rate / 2 / 100, 2)
                sgst_amt = round(item_taxable * gst_rate / 2 / 100, 2)
                igst_amt = 0.0

            amount = round(item_taxable + cgst_amt + sgst_amt + igst_amt, 2)

        item_cost_price = None
        prod = None
        if it.get('product_id'):
            prod = conn.execute('SELECT name, conversion_factor, purchase_price FROM products WHERE id=?', (it['product_id'],)).fetchone()
            conv = float(prod['conversion_factor'] or 1.0) if prod else 1.0
            qty_purchase = round(qty / conv, 4)
            unit_cost_purchase = update_stock(conn, it['product_id'], -qty_purchase, 'out', price, bill_no)
            if unit_cost_purchase and float(unit_cost_purchase) > 0:
                item_cost_price = round(float(unit_cost_purchase) * conv, 4)
            else:
                item_cost_price = round(float(prod['purchase_price'] or 0) * conv, 4) if prod else None

        prod_name = (it.get('product_name') or (prod['name'] if prod else 'Item')).strip()

        conn.execute(
            '''INSERT INTO bill_items
               (bill_id, product_id, product_name, hsn_code, unit, quantity,
                unit_price, gst_rate, discount, taxable_amt, cgst_amt, sgst_amt, igst_amt, amount, cost_price)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (bill_id, it.get('product_id'), prod_name, it.get('hsn_code', ''),
             it.get('unit', 'kg'), qty, price, gst_rate, discount_pct,
             item_taxable, cgst_amt, sgst_amt, igst_amt, amount, item_cost_price)
        )

    # ── Double-Entry Ledger Posting for Sales Bill ───────────────────────────
    pmode = (d.get('payment_mode') or 'cash').strip().lower()
    pay_account = 'Cash' if pmode == 'cash' else 'Bank'
    bill_date = str(date.today())

    ledger_entries = []
    if amount_paid > 0:
        ledger_entries.append({'account_name': pay_account, 'debit': round(amount_paid, 2), 'credit': 0, 'narration': f"Payment received via {pmode}"})
    if amount_due > 0:
        ledger_entries.append({'account_name': 'Sundry Debtors', 'debit': round(amount_due, 2), 'credit': 0, 'narration': f"Amount due from {d.get('customer_name', 'Customer')}"})
    ledger_entries.append({'account_name': 'Sales Account', 'debit': 0, 'credit': round(subtotal, 2), 'narration': f"Sales revenue Bill {bill_no}"})

    if is_interstate:
        if igst_total > 0:
            ledger_entries.append({'account_name': 'IGST Payable', 'debit': 0, 'credit': round(igst_total, 2), 'narration': f"IGST on Bill {bill_no}"})
    else:
        if cgst_total > 0:
            ledger_entries.append({'account_name': 'CGST Payable', 'debit': 0, 'credit': round(cgst_total, 2), 'narration': f"CGST on Bill {bill_no}"})
        if sgst_total > 0:
            ledger_entries.append({'account_name': 'SGST Payable', 'debit': 0, 'credit': round(sgst_total, 2), 'narration': f"SGST on Bill {bill_no}"})

    tot_dr = sum(e['debit'] for e in ledger_entries)
    tot_cr = sum(e['credit'] for e in ledger_entries)
    diff = round(tot_dr - tot_cr, 2)
    if diff != 0:
        if diff > 0:
            ledger_entries.append({'account_name': 'Round Off', 'debit': 0, 'credit': diff, 'narration': 'Rounding difference'})
        else:
            ledger_entries.append({'account_name': 'Round Off', 'debit': abs(diff), 'credit': 0, 'narration': 'Rounding difference'})

    try:
        post_ledger_entry(
            conn,
            voucher_type='sales',
            voucher_no=bill_no,
            voucher_date=bill_date,
            entries=ledger_entries,
            reference_table='bills',
            reference_id=bill_id,
            created_by=session.get('username')
        )
    except Exception as e:
        conn.rollback()
        conn.close()
        return err(f"Ledger posting failed for Bill {bill_no}: {str(e)}", 500)

    conn.commit()
    bill       = conn.execute('SELECT * FROM bills WHERE id=?', (bill_id,)).fetchone()
    bill_items = conn.execute('SELECT * FROM bill_items WHERE bill_id=?', (bill_id,)).fetchall()
    payments   = conn.execute('SELECT * FROM bill_payments WHERE bill_id=? ORDER BY paid_at ASC', (bill_id,)).fetchall()
    conn.close()
    result = dict_row(bill); result['items'] = dict_rows(bill_items); result['payments'] = dict_rows(payments)
    log_activity('CREATE_BILL', f"Bill {bill_no} created — ₹{grand_total} for {d.get('customer_name','Walk-in')}", 'bills', bill_id)
    return ok(result, f"Bill {bill_no} created"), 201

@app.route('/api/bills/next-number', methods=['GET'])
@require_permission('billing.view')
def get_next_bill_number():
    conn = get_db()
    if session.get('user_role') == 'tester':
        n = int(get_setting('next_test_no', conn) or 1)
        conn.close()
        return ok({'next_bill_no': f"TEST-{n:05d}", 'is_test': True})

    n = int(get_setting('next_bill_no', conn) or 1)
    prefix = get_setting('bill_prefix', conn) or 'MPI'
    conn.close()
    return ok({'next_bill_no': f"{prefix}-{n:05d}", 'is_test': False})

@app.route('/api/bills/<int:bid>', methods=['DELETE'])
@require_permission('billing.void_bill')
def cancel_bill(bid):
    d = request.get_json() or {}
    reason = (d.get('reason') or '').strip()
    if not reason:
        return err("Cancellation / rejection reason is required")

    conn = get_db()
    bill = conn.execute('SELECT * FROM bills WHERE id=?', (bid,)).fetchone()
    if not bill: conn.close(); return err("Not found", 404)
    if bill['status'] == 'cancelled': conn.close(); return err("Already cancelled")
    items = conn.execute('SELECT * FROM bill_items WHERE bill_id=?', (bid,)).fetchall()
    for it in items:
        if it['product_id']:
            prod = conn.execute('SELECT conversion_factor FROM products WHERE id=?', (it['product_id'],)).fetchone()
            conv = float(prod['conversion_factor'] or 1.0) if prod else 1.0
            qty_purchase = round(float(it['quantity']) / conv, 4)
            update_stock(conn, it['product_id'], qty_purchase, 'in',
                         notes=f"Reversal of {bill['bill_no']} (Reason: {reason})",
                         ref=bill['bill_no'])
    conn.execute("UPDATE bills SET status='cancelled', cancel_reason=? WHERE id=?", (reason, bid))

    # Audit alert for Managing Director if Manager cancelled a bill
    if session.get('user_role') == 'manager':
        msg = f"Manager @{session.get('username')} ({session.get('full_name')}) cancelled Bill {bill['bill_no']} (Total: ₹{bill['grand_total']}). Reason: {reason}"
        conn.execute('INSERT INTO notifications (target_role, title, message) VALUES (?,?,?)',
                     ('admin', '⚠️ Bill Rejection / Cancellation by Manager', msg))

    conn.commit(); conn.close()
    return ok(message=f"Bill {bill['bill_no']} cancelled and stock restored")

@app.route('/api/bills/<int:bid>/payments', methods=['POST'])
@require_permission('billing.payment')
def add_bill_payment(bid):
    d = request.get_json()
    if d is None:
        return err("Invalid or missing JSON payload")
    try:
        amount = float(d.get('amount', 0))
    except (ValueError, TypeError):
        amount = 0.0

    if amount <= 0:
        return err("Valid payment amount > 0 is required")

    payment_mode = (d.get('payment_mode') or 'cash').strip()
    notes = (d.get('notes') or 'Subsequent payment').strip()

    conn = get_db()
    bill = conn.execute('SELECT * FROM bills WHERE id=?', (bid,)).fetchone()
    if not bill:
        conn.close()
        return err("Bill not found", 404)

    if bill['status'] == 'cancelled':
        conn.close()
        return err("Cannot record payment for a cancelled bill")

    if bill['status'] == 'paid' or float(bill['amount_due'] or 0) <= 0:
        conn.close()
        return err("Bill is already fully paid")

    cur_due = float(bill['amount_due'])
    if amount > cur_due + 0.01:
        conn.close()
        return err(f"Payment amount (₹{amount:.2f}) exceeds remaining due amount (₹{cur_due:.2f})")

    p_amount = min(amount, cur_due)
    new_due = round(cur_due - p_amount, 2)
    new_paid = round(float(bill['amount_paid']) + p_amount, 2)
    new_status = 'paid' if new_due == 0 else 'partial'

    conn.execute(
        'UPDATE bills SET amount_paid=?, amount_due=?, status=? WHERE id=?',
        (new_paid, new_due, new_status, bid)
    )
    cur_p = conn.execute(
        '''INSERT INTO bill_payments (bill_id, amount, payment_mode, received_by, notes)
           VALUES (?,?,?,?,?)''',
        (bid, p_amount, payment_mode, session.get('username'), notes)
    )
    pmt_id = cur_p.lastrowid

    # ── Double-Entry Ledger Posting for Bill Payment Received ───────────────
    pmode = payment_mode.strip().lower()
    pay_account = 'Cash' if pmode == 'cash' else 'Bank'
    pmt_entries = [
        {'account_name': pay_account, 'debit': round(p_amount, 2), 'credit': 0, 'narration': f"Bill payment received ({pmode})"},
        {'account_name': 'Sundry Debtors', 'debit': 0, 'credit': round(p_amount, 2), 'narration': f"Bill payment credited against dues for Bill {bill['bill_no']}"}
    ]
    try:
        post_ledger_entry(
            conn,
            voucher_type='payment_in',
            voucher_no=f"PAYIN-{pmt_id}",
            voucher_date=str(date.today()),
            entries=pmt_entries,
            reference_table='bill_payments',
            reference_id=pmt_id,
            created_by=session.get('username')
        )
    except Exception as e:
        conn.rollback()
        conn.close()
        return err(f"Ledger posting failed for payment: {str(e)}", 500)

    conn.commit()

    updated_bill = conn.execute('SELECT * FROM bills WHERE id=?', (bid,)).fetchone()
    items = conn.execute('SELECT * FROM bill_items WHERE bill_id=?', (bid,)).fetchall()
    payments = conn.execute('SELECT * FROM bill_payments WHERE bill_id=? ORDER BY paid_at ASC', (bid,)).fetchall()
    conn.close()

    res = dict_row(updated_bill)
    res['items'] = dict_rows(items)
    res['payments'] = dict_rows(payments)
    return ok(res, f"Payment of ₹{p_amount:.2f} recorded for Bill {bill['bill_no']}")

# ─── Credit Notes ─────────────────────────────────────────────────────────────

@app.route('/api/bills/<int:bid>/credit-note', methods=['POST'])
@require_permission('billing.credit_note')
def create_credit_note(bid):
    d = request.get_json()
    if d is None:
        return err("Invalid or missing JSON payload")

    reason = (d.get('reason') or '').strip()
    if not reason:
        return err("Reason for credit note is required")

    items = d.get('items', [])
    if not isinstance(items, list) or not items:
        return err("Credit note must contain at least one item")

    conn = get_db()
    bill = conn.execute('SELECT * FROM bills WHERE id=?', (bid,)).fetchone()
    if not bill:
        conn.close()
        return err("Bill not found", 404)

    if bill['status'] == 'cancelled':
        conn.close()
        return err("Cannot issue a credit note for a cancelled bill")

    # Fetch original bill items
    orig_items_rows = conn.execute('SELECT * FROM bill_items WHERE bill_id=?', (bid,)).fetchall()
    orig_items = {it['id']: dict_row(it) for it in orig_items_rows}

    # Query already credited quantity per bill_item_id
    credited_rows = conn.execute('''
        SELECT cni.bill_item_id, COALESCE(SUM(cni.quantity), 0) AS credited_qty
        FROM credit_note_items cni
        JOIN credit_notes cn ON cni.credit_note_id = cn.id
        WHERE cn.bill_id = ?
        GROUP BY cni.bill_item_id
    ''', (bid,)).fetchall()
    credited_map = {r['bill_item_id']: float(r['credited_qty']) for r in credited_rows}

    is_interstate = bool(bill['is_interstate'])
    cn_no = next_cn_no(conn)

    computed_items = []
    subtotal = 0.0
    cgst_total = 0.0
    sgst_total = 0.0
    igst_total = 0.0
    total = 0.0

    for item in items:
        b_item_id = item.get('bill_item_id')
        if not b_item_id or b_item_id not in orig_items:
            conn.close()
            return err(f"Invalid or non-existent bill item ID: {b_item_id}")

        try:
            ret_qty = float(item.get('quantity', 0))
        except (ValueError, TypeError):
            ret_qty = 0.0

        if ret_qty <= 0:
            conn.close()
            return err("Return quantity must be greater than 0")

        orig_item = orig_items[b_item_id]
        orig_qty = float(orig_item['quantity'])
        already_credited = credited_map.get(b_item_id, 0.0)

        if round(already_credited + ret_qty, 4) > round(orig_qty, 4):
            conn.close()
            return err(
                f"Requested return quantity ({ret_qty}) plus previously credited ({already_credited}) "
                f"exceeds original billed quantity ({orig_qty}) for item '{orig_item['product_name']}'"
            )

        unit_price = float(orig_item['unit_price'])
        gst_rate = float(orig_item['gst_rate'] or 0)
        item_discount_pct = float(orig_item['discount'] or 0)

        taxable_amt = round(ret_qty * unit_price * (1 - item_discount_pct / 100), 2)
        if is_interstate:
            cgst_amt = 0.0
            sgst_amt = 0.0
            igst_amt = round(taxable_amt * gst_rate / 100, 2)
        else:
            cgst_amt = round(taxable_amt * gst_rate / 2 / 100, 2)
            sgst_amt = round(taxable_amt * gst_rate / 2 / 100, 2)
            igst_amt = 0.0

        item_amount = round(taxable_amt + cgst_amt + sgst_amt + igst_amt, 2)

        subtotal += taxable_amt
        cgst_total += cgst_amt
        sgst_total += sgst_amt
        igst_total += igst_amt
        total += item_amount

        # Restore stock via update_stock
        if orig_item['product_id']:
            pid = orig_item['product_id']
            prod = conn.execute('SELECT conversion_factor FROM products WHERE id=?', (pid,)).fetchone()
            conv = float(prod['conversion_factor'] or 1.0) if prod else 1.0
            qty_purchase = round(ret_qty / conv, 4)
            update_stock(
                conn, pid, qty_purchase, 'in',
                unit_price=unit_price,
                ref=bill['bill_no'],
                notes=f"Credit note {cn_no} against {bill['bill_no']}"
            )

        computed_items.append({
            'bill_item_id': b_item_id,
            'product_id': orig_item['product_id'],
            'product_name': orig_item['product_name'],
            'hsn_code': orig_item.get('hsn_code', ''),
            'unit': orig_item.get('unit', 'kg'),
            'quantity': ret_qty,
            'unit_price': unit_price,
            'gst_rate': gst_rate,
            'taxable_amt': taxable_amt,
            'cgst_amt': cgst_amt,
            'sgst_amt': sgst_amt,
            'igst_amt': igst_amt,
            'amount': item_amount
        })

    subtotal = round(subtotal, 2)
    cgst_total = round(cgst_total, 2)
    sgst_total = round(sgst_total, 2)
    igst_total = round(igst_total, 2)
    total = round(total, 2)

    c = conn.execute('''
        INSERT INTO credit_notes
        (cn_no, bill_id, customer_id, customer_name, reason, subtotal, cgst, sgst, igst, total, status, created_by)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
    ''', (
        cn_no, bid, bill['customer_id'], bill['customer_name'], reason,
        subtotal, cgst_total, sgst_total, igst_total, total, 'issued', session.get('username')
    ))
    cn_id = c.lastrowid

    for item_data in computed_items:
        conn.execute('''
            INSERT INTO credit_note_items
            (credit_note_id, bill_item_id, product_id, product_name, hsn_code, unit, quantity,
             unit_price, gst_rate, taxable_amt, cgst_amt, sgst_amt, igst_amt, amount)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (
            cn_id, item_data['bill_item_id'], item_data['product_id'], item_data['product_name'],
            item_data['hsn_code'], item_data['unit'], item_data['quantity'],
            item_data['unit_price'], item_data['gst_rate'], item_data['taxable_amt'],
            item_data['cgst_amt'], item_data['sgst_amt'], item_data['igst_amt'], item_data['amount']
        ))

    # ── Double-Entry Ledger Posting for Credit Note ─────────────────────────
    orig_pmode = (bill['payment_mode'] or 'cash').strip().lower()
    if orig_pmode == 'cash':
        credit_account = 'Cash'
    elif orig_pmode in ('upi', 'card', 'bank_transfer'):
        credit_account = 'Bank'
    else:
        credit_account = 'Sundry Debtors'

    cn_entries = [
        {'account_name': 'Sales Account', 'debit': round(subtotal, 2), 'credit': 0, 'narration': f"Sales return for CN {cn_no}"}
    ]
    if is_interstate:
        if igst_total > 0:
            cn_entries.append({'account_name': 'IGST Payable', 'debit': round(igst_total, 2), 'credit': 0, 'narration': f"IGST reversal for CN {cn_no}"})
    else:
        if cgst_total > 0:
            cn_entries.append({'account_name': 'CGST Payable', 'debit': round(cgst_total, 2), 'credit': 0, 'narration': f"CGST reversal for CN {cn_no}"})
        if sgst_total > 0:
            cn_entries.append({'account_name': 'SGST Payable', 'debit': round(sgst_total, 2), 'credit': 0, 'narration': f"SGST reversal for CN {cn_no}"})

    cn_entries.append({'account_name': credit_account, 'debit': 0, 'credit': round(total, 2), 'narration': f"Credit Note refund/reversal for CN {cn_no}"})

    tot_dr = sum(e['debit'] for e in cn_entries)
    tot_cr = sum(e['credit'] for e in cn_entries)
    diff = round(tot_dr - tot_cr, 2)
    if diff != 0:
        if diff > 0:
            cn_entries.append({'account_name': 'Round Off', 'debit': 0, 'credit': diff, 'narration': 'Rounding difference'})
        else:
            cn_entries.append({'account_name': 'Round Off', 'debit': abs(diff), 'credit': 0, 'narration': 'Rounding difference'})

    try:
        post_ledger_entry(
            conn,
            voucher_type='credit_note',
            voucher_no=cn_no,
            voucher_date=str(date.today()),
            entries=cn_entries,
            reference_table='credit_notes',
            reference_id=cn_id,
            created_by=session.get('username')
        )
    except Exception as e:
        conn.rollback()
        conn.close()
        return err(f"Ledger posting failed for Credit Note {cn_no}: {str(e)}", 500)

    conn.commit()

    cn_row = conn.execute('''
        SELECT cn.*, b.bill_no
        FROM credit_notes cn
        LEFT JOIN bills b ON cn.bill_id = b.id
        WHERE cn.id=?
    ''', (cn_id,)).fetchone()
    cn_items = conn.execute('SELECT * FROM credit_note_items WHERE credit_note_id=?', (cn_id,)).fetchall()
    conn.close()

    res = dict_row(cn_row)
    res['items'] = dict_rows(cn_items)
    return ok(res, f"Credit note {cn_no} created"), 201


@app.route('/api/credit-notes', methods=['GET'])
@require_permission('billing.view')
def list_credit_notes():
    conn = get_db()
    date_from  = request.args.get('from', '')
    date_to    = request.args.get('to', '')
    bill_id    = request.args.get('bill_id', '')
    customer_q = request.args.get('q', '').strip()

    sql = '''
        SELECT cn.*, b.bill_no
        FROM credit_notes cn
        LEFT JOIN bills b ON cn.bill_id = b.id
        WHERE 1=1
    '''
    params = []
    if date_from:
        sql += ' AND date(cn.created_at) >= ?'
        params.append(date_from)
    if date_to:
        sql += ' AND date(cn.created_at) <= ?'
        params.append(date_to)
    if bill_id:
        sql += ' AND cn.bill_id = ?'
        params.append(bill_id)
    if customer_q:
        sql += ' AND (cn.cn_no LIKE ? OR cn.customer_name LIKE ?)'
        params += [f'%{customer_q}%', f'%{customer_q}%']

    sql += ' ORDER BY cn.created_at DESC, cn.id DESC'
    rows = conn.execute(sql, params).fetchall()
    conn.close()

    if wants_excel():
        excel_rows = []
        for r in rows:
            excel_rows.append([
                r['cn_no'] or '',
                r['created_at'][:10] if r['created_at'] else '',
                r['bill_no'] or '',
                r['customer_name'] or '',
                float(r['subtotal'] or 0),
                float(r['cgst'] or 0),
                float(r['sgst'] or 0),
                float(r['igst'] or 0),
                float(r['total'] or 0),
                r['reason'] or '',
                r['status'] or ''
            ])
        sheets = [{
            "sheet_name": "Credit Notes",
            "headers": ["CN No", "Date", "Original Bill No", "Customer", "Subtotal (₹)", "CGST (₹)", "SGST (₹)", "IGST (₹)", "Total (₹)", "Reason", "Status"],
            "rows": excel_rows
        }]
        return export_to_excel(sheets, "credit_notes_list")

    return ok(dict_rows(rows))


@app.route('/api/credit-notes/<int:cnid>', methods=['GET'])
@require_permission('billing.view')
def get_credit_note(cnid):
    conn = get_db()
    cn = conn.execute('''
        SELECT cn.*, b.bill_no
        FROM credit_notes cn
        LEFT JOIN bills b ON cn.bill_id = b.id
        WHERE cn.id = ?
    ''', (cnid,)).fetchone()
    if not cn:
        conn.close()
        return err("Credit note not found", 404)

    items = conn.execute('SELECT * FROM credit_note_items WHERE credit_note_id = ?', (cnid,)).fetchall()
    conn.close()

    res = dict_row(cn)
    res['items'] = dict_rows(items)
    return ok(res)


@app.route('/api/bills/<int:bid>/credit-notes', methods=['GET'])
@require_permission('billing.view')
def list_bill_credit_notes(bid):
    conn = get_db()
    bill = conn.execute('SELECT id FROM bills WHERE id=?', (bid,)).fetchone()
    if not bill:
        conn.close()
        return err("Bill not found", 404)

    cn_rows = conn.execute('''
        SELECT cn.*, b.bill_no
        FROM credit_notes cn
        LEFT JOIN bills b ON cn.bill_id = b.id
        WHERE cn.bill_id = ?
        ORDER BY cn.created_at DESC, cn.id DESC
    ''', (bid,)).fetchall()

    credit_notes = dict_rows(cn_rows)
    for cn in credit_notes:
        items = conn.execute('SELECT * FROM credit_note_items WHERE credit_note_id = ?', (cn['id'],)).fetchall()
        cn['items'] = dict_rows(items)

    summary_rows = conn.execute('''
        SELECT cni.bill_item_id, COALESCE(SUM(cni.quantity), 0) as total_credited_quantity
        FROM credit_note_items cni
        JOIN credit_notes cn ON cni.credit_note_id = cn.id
        WHERE cn.bill_id = ?
        GROUP BY cni.bill_item_id
    ''', (bid,)).fetchall()

    credited_summary = {row['bill_item_id']: row['total_credited_quantity'] for row in summary_rows}

    conn.close()
    return ok({
        "credit_notes": credit_notes,
        "credited_summary": credited_summary
    })

# ─── Purchase Orders ─────────────────────────────────────────────────────────

@app.route('/api/purchase-orders', methods=['GET'])
@require_permission('purchase.view')
def list_purchase_orders():
    conn = get_db()
    date_from    = request.args.get('from', '')
    date_to      = request.args.get('to', '')
    supplier_q   = request.args.get('q', '').strip()
    status_param = request.args.get('status', '').strip().lower()
    limit        = int(request.args.get('limit', 200))

    sql = 'SELECT * FROM purchase_orders WHERE 1=1'
    params = []
    if date_from:
        sql += ' AND date(date) >= ?'; params.append(date_from)
    if date_to:
        sql += ' AND date(date) <= ?'; params.append(date_to)
    if supplier_q:
        sql += ' AND (supplier_name LIKE ? OR po_no LIKE ?)'
        params += [f'%{supplier_q}%', f'%{supplier_q}%']
    if status_param:
        sql += ' AND LOWER(status) = ?'; params.append(status_param)

    sql += ' ORDER BY date DESC, id DESC'

    if wants_excel():
        rows = conn.execute(sql, params).fetchall()
        conn.close()
        excel_rows = []
        for r in rows:
            excel_rows.append([
                r['po_no'] or '',
                r['date'] or '',
                r['supplier_name'] or '',
                float(r['total'] or 0),
                float(r['amount_paid'] or 0),
                float(r['amount_due'] or 0),
                r['status'] or '',
                r['notes'] or ''
            ])
        sheets = [{
            "sheet_name": "Purchase Orders",
            "headers": ["PO No", "Date", "Supplier", "Total (₹)", "Amount Paid (₹)", "Amount Due (₹)", "Status", "Notes"],
            "rows": excel_rows
        }]
        return export_to_excel(sheets, "purchase_orders_list")

    rows = conn.execute(sql + f' LIMIT ?', params + [limit]).fetchall()
    conn.close()
    return ok(dict_rows(rows))

@app.route('/api/purchase-orders/<int:oid>', methods=['GET'])
@require_permission('purchase.view')
def get_purchase_order(oid):
    conn = get_db()
    po = conn.execute('SELECT * FROM purchase_orders WHERE id=?', (oid,)).fetchone()
    if not po: conn.close(); return err("Not found", 404)
    items = conn.execute('SELECT * FROM purchase_order_items WHERE order_id=?', (oid,)).fetchall()
    payments = conn.execute('SELECT * FROM po_payments WHERE order_id=? ORDER BY paid_at ASC', (oid,)).fetchall()
    conn.close()
    result = dict_row(po); result['items'] = dict_rows(items); result['payments'] = dict_rows(payments)
    return ok(result)

@app.route('/api/purchase-orders', methods=['POST'])
@require_permission('purchase.manage')
def create_purchase_order():
    d = request.get_json()
    if d is None:
        return err("Invalid or missing JSON payload")
    items = d.get('items', [])
    if not isinstance(items, list) or not items: return err("PO must have at least one item")
    conn = get_db()
    po_no    = next_po_no(conn)
    supplier = conn.execute('SELECT name FROM suppliers WHERE id=?', (d.get('supplier_id'),)).fetchone()
    subtotal = sum(float(it['quantity']) * float(it['unit_price']) for it in items)
    total    = round(subtotal, 2)
    amount_paid = float(d.get('amount_paid', total))
    amount_due  = max(round(total - amount_paid, 2), 0.0)

    c = conn.execute(
        '''INSERT INTO purchase_orders
           (po_no, supplier_id, supplier_name, subtotal, total, amount_paid, amount_due, status, notes)
           VALUES (?,?,?,?,?,?,?,?,?)''',
        (po_no, d.get('supplier_id'), supplier['name'] if supplier else '',
         round(subtotal, 2), total, amount_paid, amount_due,
         d.get('status', 'received'), d.get('notes', ''))
    )
    oid = c.lastrowid

    if amount_paid > 0:
        conn.execute(
            '''INSERT INTO po_payments (order_id, amount, payment_mode, recorded_by, notes)
               VALUES (?,?,?,?,?)''',
            (oid, amount_paid, d.get('payment_mode', 'bank_transfer'), session.get('username'), 'Initial payment at PO creation')
        )

    for it in items:
        qty = float(it['quantity']); price = float(it['unit_price']); amount = round(qty * price, 2)
        conn.execute(
            '''INSERT INTO purchase_order_items
               (order_id, product_id, product_name, quantity, unit_price, amount, expiry_date)
               VALUES (?,?,?,?,?,?,?)''',
            (oid, it.get('product_id'), it['product_name'], qty, price, amount, it.get('expiry_date'))
        )
        if it.get('product_id') and d.get('status', 'received') == 'received':
            update_stock(conn, it['product_id'], qty, 'in', price,
                         po_no, d.get('supplier_id'), it.get('expiry_date'), f"PO {po_no}")

    # ── Double-Entry Ledger Posting for Purchase Order ───────────────────────
    po_pmode = (d.get('payment_mode') or 'bank_transfer').strip().lower()
    pay_account = 'Cash' if po_pmode == 'cash' else 'Bank'

    po_entries = [
        {'account_name': 'Purchase Account', 'debit': round(total, 2), 'credit': 0, 'narration': f"Purchase Order {po_no}"}
    ]
    if amount_paid > 0:
        po_entries.append({'account_name': pay_account, 'debit': 0, 'credit': round(amount_paid, 2), 'narration': f"PO payment paid ({po_pmode})"})
    if amount_due > 0:
        po_entries.append({'account_name': 'Sundry Creditors', 'debit': 0, 'credit': round(amount_due, 2), 'narration': f"PO amount payable to {supplier['name'] if supplier else 'Supplier'}"})

    tot_dr = sum(e['debit'] for e in po_entries)
    tot_cr = sum(e['credit'] for e in po_entries)
    diff = round(tot_dr - tot_cr, 2)
    if diff != 0:
        if diff > 0:
            po_entries.append({'account_name': 'Round Off', 'debit': 0, 'credit': diff, 'narration': 'Rounding difference'})
        else:
            po_entries.append({'account_name': 'Round Off', 'debit': abs(diff), 'credit': 0, 'narration': 'Rounding difference'})

    try:
        post_ledger_entry(
            conn,
            voucher_type='purchase',
            voucher_no=po_no,
            voucher_date=str(date.today()),
            entries=po_entries,
            reference_table='purchase_orders',
            reference_id=oid,
            created_by=session.get('username')
        )
    except Exception as e:
        conn.rollback()
        conn.close()
        return err(f"Ledger posting failed for PO {po_no}: {str(e)}", 500)

    conn.commit()
    po       = conn.execute('SELECT * FROM purchase_orders WHERE id=?', (oid,)).fetchone()
    po_items = conn.execute('SELECT * FROM purchase_order_items WHERE order_id=?', (oid,)).fetchall()
    payments = conn.execute('SELECT * FROM po_payments WHERE order_id=? ORDER BY paid_at ASC', (oid,)).fetchall()
    conn.close()
    result = dict_row(po); result['items'] = dict_rows(po_items); result['payments'] = dict_rows(payments)
    return ok(result, f"Purchase order {po_no} created"), 201

@app.route('/api/purchase-orders/<int:oid>/payments', methods=['POST'])
@require_permission('purchase.manage')
def add_po_payment(oid):
    d = request.get_json()
    if d is None:
        return err("Invalid or missing JSON payload")
    try:
        amount = float(d.get('amount', 0))
    except (ValueError, TypeError):
        amount = 0.0

    if amount <= 0:
        return err("Valid payment amount > 0 is required")

    payment_mode = (d.get('payment_mode') or 'bank_transfer').strip()
    notes = (d.get('notes') or 'Follow-up supplier payment').strip()

    conn = get_db()
    po = conn.execute('SELECT * FROM purchase_orders WHERE id=?', (oid,)).fetchone()
    if not po:
        conn.close()
        return err("Purchase order not found", 404)

    if po['status'] == 'cancelled':
        conn.close()
        return err("Cannot record payment for a cancelled purchase order")

    cur_due = float(po['amount_due'] or 0)
    if cur_due <= 0:
        conn.close()
        return err("Purchase order is already fully paid")

    if amount > cur_due + 0.01:
        conn.close()
        return err(f"Payment amount (₹{amount:.2f}) exceeds remaining due amount (₹{cur_due:.2f})")

    p_amount = min(amount, cur_due)
    new_due = round(cur_due - p_amount, 2)
    new_paid = round(float(po['amount_paid'] or 0) + p_amount, 2)

    conn.execute(
        'UPDATE purchase_orders SET amount_paid=?, amount_due=? WHERE id=?',
        (new_paid, new_due, oid)
    )
    cur_pop = conn.execute(
        '''INSERT INTO po_payments (order_id, amount, payment_mode, recorded_by, notes)
           VALUES (?,?,?,?,?)''',
        (oid, p_amount, payment_mode, session.get('username'), notes)
    )
    pop_id = cur_pop.lastrowid

    # ── Double-Entry Ledger Posting for Purchase Payment Sent ───────────────
    pmode = payment_mode.strip().lower()
    pay_account = 'Cash' if pmode == 'cash' else 'Bank'
    pop_entries = [
        {'account_name': 'Sundry Creditors', 'debit': round(p_amount, 2), 'credit': 0, 'narration': f"Supplier payment for PO {po['po_no']}"},
        {'account_name': pay_account, 'debit': 0, 'credit': round(p_amount, 2), 'narration': f"Payment outflow ({pmode})"}
    ]
    try:
        post_ledger_entry(
            conn,
            voucher_type='payment_out',
            voucher_no=f"PAYOUT-{pop_id}",
            voucher_date=str(date.today()),
            entries=pop_entries,
            reference_table='po_payments',
            reference_id=pop_id,
            created_by=session.get('username')
        )
    except Exception as e:
        conn.rollback()
        conn.close()
        return err(f"Ledger posting failed for PO payment: {str(e)}", 500)

    conn.commit()

    updated_po = conn.execute('SELECT * FROM purchase_orders WHERE id=?', (oid,)).fetchone()
    items = conn.execute('SELECT * FROM purchase_order_items WHERE order_id=?', (oid,)).fetchall()
    payments = conn.execute('SELECT * FROM po_payments WHERE order_id=? ORDER BY paid_at ASC', (oid,)).fetchall()
    conn.close()

    res = dict_row(updated_po)
    res['items'] = dict_rows(items)
    res['payments'] = dict_rows(payments)
    return ok(res, f"Payment of ₹{p_amount:.2f} recorded for PO {po['po_no']}")

# ─── Expenses & Other Income ─────────────────────────────────────────────────

@app.route('/api/expenses', methods=['GET'])
@require_permission('expenses.view')
def list_expenses():
    conn = get_db()
    date_from    = request.args.get('from', '')
    date_to      = request.args.get('to', '')
    category_q   = request.args.get('category', '').strip()
    entry_type_q = request.args.get('entry_type', '').strip().lower()

    sql = 'SELECT * FROM expenses WHERE 1=1'
    params = []
    if date_from:
        sql += ' AND date >= ?'; params.append(date_from)
    if date_to:
        sql += ' AND date <= ?'; params.append(date_to)
    if category_q:
        sql += ' AND category LIKE ?'; params.append(f'%{category_q}%')
    if entry_type_q in ('expense', 'income'):
        sql += ' AND (LOWER(entry_type) = ? OR (entry_type IS NULL AND ? = "expense"))'
        params += [entry_type_q, entry_type_q]

    sql += ' ORDER BY date DESC, id DESC'
    rows = conn.execute(sql, params).fetchall()
    conn.close()

    result_rows = dict_rows(rows)
    for r in result_rows:
        if not r.get('entry_type'):
            r['entry_type'] = 'expense'
        if not r.get('payment_mode'):
            r['payment_mode'] = 'cash'

    if wants_excel():
        excel_rows = []
        for r in result_rows:
            excel_rows.append([
                (r.get('entry_type') or 'expense').upper(),
                r.get('date') or '',
                r.get('category') or '',
                r.get('description') or '',
                float(r.get('amount') or 0),
                (r.get('payment_mode') or 'cash').upper()
            ])
        sheets = [{
            "sheet_name": "Income & Expenses",
            "headers": ["Type", "Date", "Category", "Description", "Amount (₹)", "Payment Mode"],
            "rows": excel_rows
        }]
        return export_to_excel(sheets, "income_expenses_list")

    return ok(result_rows)

@app.route('/api/expenses', methods=['POST'])
@require_permission('expenses.manage')
def create_expense():
    d = request.get_json()
    if d is None:
        return err("Invalid or missing JSON payload")
    if not d.get('category') or 'amount' not in d:
        return err("Category and amount are required")
    try:
        exp_amount = float(d['amount'])
    except (ValueError, TypeError):
        return err("Valid amount required")
    if exp_amount <= 0:
        return err("Amount must be greater than 0")

    entry_type = (d.get('entry_type') or 'expense').strip().lower()
    if entry_type not in ('expense', 'income'):
        entry_type = 'expense'

    pmode = (d.get('payment_mode') or 'cash').strip().lower()
    pay_account = 'Cash' if pmode == 'cash' else 'Bank'

    conn = get_db()
    c = conn.execute(
        'INSERT INTO expenses (category, description, amount, date, entry_type, payment_mode) VALUES (?,?,?,?,?,?)',
        (d['category'], d.get('description',''), exp_amount,
         d.get('date', str(date.today())), entry_type, pmode)
    )
    exp_id = c.lastrowid

    if entry_type == 'income':
        voucher_type = 'payment_in'
        voucher_no = f"INC-{exp_id}"
        exp_entries = [
            {'account_name': pay_account, 'debit': round(exp_amount, 2), 'credit': 0, 'narration': f"Other Income ({pmode})"},
            {'account_name': 'Sales Account', 'debit': 0, 'credit': round(exp_amount, 2), 'narration': f"Other Income: {d['category']} ({d.get('description', '')})".strip()}
        ]
    else:
        voucher_type = 'expense'
        voucher_no = f"EXP-{exp_id}"
        exp_cat = (d['category'] or '').strip().lower()
        direct_keywords = ('freight', 'transport', 'cartage', 'raw material', 'packaging', 'direct')
        if any(k in exp_cat for k in direct_keywords):
            exp_account = 'Direct Expenses'
        else:
            exp_account = 'Indirect Expenses'

        exp_entries = [
            {'account_name': exp_account, 'debit': round(exp_amount, 2), 'credit': 0, 'narration': f"Expense: {d['category']} ({d.get('description', '')})".strip()},
            {'account_name': pay_account, 'debit': 0, 'credit': round(exp_amount, 2), 'narration': f"Expense paid via {pmode}"}
        ]

    try:
        post_ledger_entry(
            conn,
            voucher_type=voucher_type,
            voucher_no=voucher_no,
            voucher_date=d.get('date', str(date.today())),
            entries=exp_entries,
            reference_table='expenses',
            reference_id=exp_id,
            created_by=session.get('username')
        )
    except Exception as e:
        conn.rollback()
        conn.close()
        return err(f"Ledger posting failed for {entry_type.title()}: {str(e)}", 500)

    conn.commit()
    row = conn.execute('SELECT * FROM expenses WHERE id=?', (c.lastrowid,)).fetchone()
    conn.close()
    log_activity('ADD_EXPENSE', f"{entry_type.title()}: {d['category']} — ₹{exp_amount}", 'expenses', exp_id)
    return ok(dict_row(row)), 201

@app.route('/api/expenses/<int:eid>', methods=['DELETE'])
@require_permission('expenses.manage')
def delete_expense(eid):
    conn = get_db()
    exp = conn.execute('SELECT category, amount FROM expenses WHERE id=?', (eid,)).fetchone()
    conn.execute('DELETE FROM expenses WHERE id=?', (eid,))
    conn.commit(); conn.close()
    if exp:
        log_activity('DELETE_EXPENSE', f"Deleted {exp['category']} ₹{exp['amount']}", 'expenses', eid)
    return ok(message="Expense deleted")

# ─── Reports ─────────────────────────────────────────────────────────────────

@app.route('/api/reports/dashboard', methods=['GET'])
@require_permission('reports.view')
def dashboard():
    if is_counter_staff():
        return err("Counter Staff do not have access to reports and financial data.", 403)
    conn = get_db()
    today = str(date.today())
    def q1(sql, params=[]): return conn.execute(sql, params).fetchone()
    today_sales   = q1("SELECT COALESCE(SUM(grand_total),0) AS v FROM bills WHERE date(date)=? AND status='paid'", (today,))['v']
    today_bills   = q1("SELECT COUNT(*) AS v FROM bills WHERE date(date)=? AND status='paid'", (today,))['v']
    month_sales   = q1("SELECT COALESCE(SUM(grand_total),0) AS v FROM bills WHERE strftime('%Y-%m',date)=strftime('%Y-%m','now') AND status='paid'")['v']
    total_products= q1("SELECT COUNT(*) AS v FROM products WHERE active=1")['v']
    low_stock     = q1("SELECT COUNT(*) AS v FROM products WHERE active=1 AND current_stock <= min_stock")['v']
    total_customers = q1("SELECT COUNT(*) AS v FROM customers")['v']
    daily = conn.execute('''
        SELECT date(date) AS day, COALESCE(SUM(grand_total),0) AS total, COUNT(*) AS bills
        FROM bills WHERE status='paid' AND date(date) >= date('now','-6 days')
        GROUP BY day ORDER BY day
    ''').fetchall()
    top = conn.execute('''
        SELECT bi.product_name, SUM(bi.quantity) AS qty, SUM(bi.amount) AS revenue
        FROM bill_items bi JOIN bills b ON bi.bill_id=b.id
        WHERE strftime('%Y-%m',b.date)=strftime('%Y-%m','now') AND b.status='paid'
        GROUP BY bi.product_name ORDER BY revenue DESC LIMIT 5
    ''').fetchall()
    payment = conn.execute('''
        SELECT payment_mode, COALESCE(SUM(grand_total),0) AS total, COUNT(*) AS count
        FROM bills WHERE date(date)=? AND status='paid' GROUP BY payment_mode
    ''', (today,)).fetchall()
    today_expenses = q1("SELECT COALESCE(SUM(amount),0) AS v FROM expenses WHERE date=?", (today,))['v']
    conn.close()
    return ok({
        "today_sales": today_sales, "today_bills": today_bills,
        "today_expenses": today_expenses, "month_sales": month_sales,
        "total_products": total_products, "low_stock_count": low_stock,
        "total_customers": total_customers,
        "daily_sales": dict_rows(daily), "top_products": dict_rows(top),
        "payment_breakdown": dict_rows(payment),
    })

@app.route('/api/reports/sales', methods=['GET'])
@require_permission('reports.view')
def sales_report():
    conn = get_db()
    date_from = request.args.get('from', str(date.today()))
    date_to   = request.args.get('to', str(date.today()))
    ptype     = request.args.get('product_type', '').strip().lower()

    if ptype in ('perishable', 'general'):
        rows = conn.execute('''
            SELECT DISTINCT b.*, GROUP_CONCAT(bi.product_name) AS products
            FROM bills b
            JOIN bill_items bi ON b.id=bi.bill_id
            JOIN products p ON bi.product_id=p.id
            WHERE date(b.date) BETWEEN ? AND ? AND b.status='paid' AND p.product_type=?
            GROUP BY b.id ORDER BY b.date DESC
        ''', [date_from, date_to, ptype]).fetchall()

        totals = conn.execute('''
            SELECT COALESCE(SUM(bi.amount),0) AS total_revenue,
                   COALESCE(SUM(bi.cgst_amt),0) AS total_cgst, COALESCE(SUM(bi.sgst_amt),0) AS total_sgst,
                   COALESCE(SUM(bi.igst_amt),0) AS total_igst,
                   COALESCE(SUM(bi.discount),0) AS total_discount, COUNT(DISTINCT b.id) AS bill_count
            FROM bills b
            JOIN bill_items bi ON b.id=bi.bill_id
            JOIN products p ON bi.product_id=p.id
            WHERE date(b.date) BETWEEN ? AND ? AND b.status='paid' AND p.product_type=?
        ''', [date_from, date_to, ptype]).fetchone()
    else:
        rows = conn.execute('''
            SELECT b.*, GROUP_CONCAT(bi.product_name) AS products
            FROM bills b LEFT JOIN bill_items bi ON b.id=bi.bill_id
            WHERE date(b.date) BETWEEN ? AND ? AND b.status='paid'
            GROUP BY b.id ORDER BY b.date DESC
        ''', (date_from, date_to)).fetchall()
        totals = conn.execute('''
            SELECT COALESCE(SUM(grand_total),0) AS total_revenue,
                   COALESCE(SUM(cgst),0) AS total_cgst, COALESCE(SUM(sgst),0) AS total_sgst,
                   COALESCE(SUM(igst),0) AS total_igst,
                   COALESCE(SUM(discount_amount),0) AS total_discount, COUNT(*) AS bill_count
            FROM bills WHERE date(date) BETWEEN ? AND ? AND status='paid'
        ''', (date_from, date_to)).fetchone()

    # Product Type Breakdown
    pt_rows = conn.execute('''
        SELECT p.product_type,
               COALESCE(SUM(bi.amount), 0) AS total_sales,
               COALESCE(SUM(bi.quantity), 0) AS item_count,
               COALESCE(SUM(bi.quantity * COALESCE(bi.cost_price, p.purchase_price)), 0) AS total_cost
        FROM bill_items bi
        JOIN bills b ON bi.bill_id = b.id
        JOIN products p ON bi.product_id = p.id
        WHERE date(b.date) BETWEEN ? AND ? AND b.status = 'paid'
        GROUP BY p.product_type
    ''', (date_from, date_to)).fetchall()

    by_product_type = {
        "perishable": {"total_sales": 0.0, "item_count": 0.0, "gross_margin": 0.0},
        "general":    {"total_sales": 0.0, "item_count": 0.0, "gross_margin": 0.0}
    }
    for ptr in pt_rows:
        t = ptr['product_type'] or 'perishable'
        tsales = float(ptr['total_sales'] or 0)
        tcount = float(ptr['item_count'] or 0)
        tcost  = float(ptr['total_cost'] or 0)
        if t in by_product_type:
            by_product_type[t] = {
                "total_sales": round(tsales, 2),
                "item_count": round(tcount, 3),
                "gross_margin": round(tsales - tcost, 2)
            }

    conn.close()

    if wants_excel():
        excel_rows = []
        for b in rows:
            excel_rows.append([
                b['date'],
                b['bill_no'],
                b['customer_name'] or 'Walk-in',
                b['products'] or '',
                float(b['subtotal'] or 0),
                float(b['cgst'] or 0),
                float(b['sgst'] or 0),
                float(b['igst'] or 0),
                float(b['discount_amount'] or 0),
                float(b['grand_total'] or 0),
                b['payment_mode'] or '',
                b['status'] or ''
            ])
        sheets = [{
            "sheet_name": "Sales Register",
            "headers": ["Date", "Bill No", "Customer", "Items", "Subtotal", "CGST", "SGST", "IGST", "Discount", "Grand Total", "Payment Mode", "Status"],
            "rows": excel_rows
        }]
        return export_to_excel(sheets, f"sales_report_{date_from}_to_{date_to}")

    return ok({"bills": dict_rows(rows), "totals": dict_row(totals),
               "by_product_type": by_product_type,
               "from": date_from, "to": date_to})

@app.route('/api/reports/gst', methods=['GET'])
@require_permission('reports.view')
def gst_report():
    conn = get_db()
    date_from = request.args.get('from', str(date.today().replace(day=1)))
    date_to   = request.args.get('to', str(date.today()))
    rows = conn.execute('''
        SELECT bi.gst_rate, COALESCE(b.is_interstate, 0) AS is_interstate,
               SUM(bi.taxable_amt) AS taxable,
               SUM(bi.cgst_amt) AS cgst, SUM(bi.sgst_amt) AS sgst, SUM(bi.igst_amt) AS igst,
               SUM(bi.cgst_amt + bi.sgst_amt + bi.igst_amt) AS total_gst
        FROM bill_items bi JOIN bills b ON bi.bill_id=b.id
        WHERE date(b.date) BETWEEN ? AND ? AND b.status='paid'
        GROUP BY bi.gst_rate, COALESCE(b.is_interstate, 0) ORDER BY bi.gst_rate, is_interstate
    ''', (date_from, date_to)).fetchall()
    conn.close()

    if wants_excel():
        excel_rows = []
        for r in rows:
            excel_rows.append([
                float(r['gst_rate'] or 0),
                "Interstate (IGST)" if r['is_interstate'] else "Intrastate (CGST+SGST)",
                float(r['taxable'] or 0),
                float(r['cgst'] or 0),
                float(r['sgst'] or 0),
                float(r['igst'] or 0),
                float(r['total_gst'] or 0)
            ])
        sheets = [{
            "sheet_name": "GST Summary",
            "headers": ["GST Rate (%)", "Supply Type", "Taxable Amount (₹)", "CGST (₹)", "SGST (₹)", "IGST (₹)", "Total GST (₹)"],
            "rows": excel_rows
        }]
        return export_to_excel(sheets, f"gst_report_{date_from}_to_{date_to}")

    return ok({"gst_summary": dict_rows(rows), "from": date_from, "to": date_to})

@app.route('/api/reports/stock', methods=['GET'])
@require_permission('reports.view')
def stock_report():
    ptype = request.args.get('product_type', '').strip().lower()
    conn = get_db()
    sql = '''
        SELECT p.*, c.name AS category_name,
               (p.current_stock * p.purchase_price) AS stock_value
        FROM products p LEFT JOIN categories c ON p.category_id=c.id
        WHERE p.active=1
    '''
    params = []
    if ptype in ('perishable', 'general'):
        sql += ' AND p.product_type = ?'
        params.append(ptype)
    sql += ' ORDER BY c.name, p.name'

    rows = conn.execute(sql, params).fetchall()

    val_sql = 'SELECT COALESCE(SUM(current_stock * purchase_price),0) AS v FROM products WHERE active=1'
    val_params = []
    if ptype in ('perishable', 'general'):
        val_sql += ' AND product_type = ?'
        val_params.append(ptype)
    total_value = conn.execute(val_sql, val_params).fetchone()['v']
    conn.close()

    result = dict_rows(rows)
    for r in result:
        conv = float(r.get('conversion_factor') or 1.0)
        r['current_stock_sale_unit'] = round(float(r['current_stock']) * conv, 3)
    return ok({"products": result, "total_value": total_value})

@app.route('/api/reports/top-products', methods=['GET'])
@require_permission('reports.view')
def top_products():
    conn = get_db()
    date_from = request.args.get('from', str(date.today().replace(day=1)))
    date_to   = request.args.get('to', str(date.today()))
    rows = conn.execute('''
        SELECT bi.product_name, bi.unit, SUM(bi.quantity) AS total_qty,
               SUM(bi.amount) AS total_revenue, COUNT(DISTINCT bi.bill_id) AS bill_count
        FROM bill_items bi JOIN bills b ON bi.bill_id=b.id
        WHERE date(b.date) BETWEEN ? AND ? AND b.status='paid'
        GROUP BY bi.product_name ORDER BY total_revenue DESC LIMIT 20
    ''', (date_from, date_to)).fetchall()
    conn.close()
    return ok(dict_rows(rows))

@app.route('/api/reports/expiring-soon', methods=['GET'])
@require_permission('reports.view')
def report_expiring_soon():
    days_param = request.args.get('days', '3')
    try:
        days = int(days_param)
    except ValueError:
        days = 3

    product_type_param = request.args.get('product_type', '').strip().lower()

    conn = get_db()
    sql = '''
        SELECT sb.*, p.name AS product_name, p.code AS product_code, p.purchase_unit, s.name AS supplier_name,
               CAST(julianday(sb.expiry_date) - julianday('now') AS INTEGER) AS days_until_expiry
        FROM stock_batches sb
        JOIN products p ON sb.product_id = p.id
        LEFT JOIN suppliers s ON sb.supplier_id = s.id
        WHERE sb.quantity_remaining > 0
          AND sb.expiry_date IS NOT NULL
          AND sb.expiry_date != ''
          AND date(sb.expiry_date) <= date('now', '+' || ? || ' days')
    '''
    params = [days]
    if product_type_param in ('perishable', 'general'):
        sql += ' AND p.product_type = ?'
        params.append(product_type_param)
    sql += ' ORDER BY sb.expiry_date ASC, sb.id ASC'
    rows = conn.execute(sql, params).fetchall()
    conn.close()

    if wants_excel():
        excel_rows = []
        for r in rows:
            excel_rows.append([
                r['product_name'] or '',
                r['product_code'] or '',
                r['batch_no'] or '',
                float(r['quantity_remaining'] or 0),
                r['purchase_unit'] or '',
                r['expiry_date'] or '',
                int(r['days_until_expiry'] or 0),
                r['supplier_name'] or ''
            ])
        sheets = [{
            "sheet_name": "Expiring Batches",
            "headers": ["Product Name", "Code", "Batch No", "Qty Remaining", "Unit", "Expiry Date", "Days Until Expiry", "Supplier"],
            "rows": excel_rows
        }]
        return export_to_excel(sheets, f"expiring_in_{days}_days")

    return ok({"batches": dict_rows(rows), "days": days})

@app.route('/api/reports/wastage', methods=['GET'])
@require_permission('reports.view')
def wastage_report():
    conn = get_db()
    date_from = request.args.get('from', str(date.today().replace(day=1)))
    date_to   = request.args.get('to', str(date.today()))
    ptype     = request.args.get('product_type', '').strip().lower()

    sql = '''
        SELECT st.product_id, p.name AS product_name, p.code AS product_code, p.product_type,
               p.purchase_unit, p.sale_unit, p.conversion_factor, p.purchase_price,
               SUM(st.quantity) AS total_wastage_purchase_units,
               SUM(st.quantity * p.conversion_factor) AS total_wastage_sale_units,
               SUM(st.quantity * p.purchase_price) AS total_wastage_cost,
               COUNT(st.id) AS transaction_count,
               GROUP_CONCAT(DISTINCT st.notes) AS reasons
        FROM stock_transactions st
        JOIN products p ON st.product_id = p.id
        WHERE st.type = 'wastage' AND date(st.date) BETWEEN ? AND ?
    '''
    params = [date_from, date_to]
    if ptype in ('perishable', 'general'):
        sql += ' AND p.product_type = ?'
        params.append(ptype)
    sql += ' GROUP BY st.product_id ORDER BY total_wastage_cost DESC'

    rows = conn.execute(sql, params).fetchall()

    cost_sql = '''
        SELECT COALESCE(SUM(st.quantity * p.purchase_price), 0) AS total_cost
        FROM stock_transactions st
        JOIN products p ON st.product_id = p.id
        WHERE st.type = 'wastage' AND date(st.date) BETWEEN ? AND ?
    '''
    cost_params = [date_from, date_to]
    if ptype in ('perishable', 'general'):
        cost_sql += ' AND p.product_type = ?'
        cost_params.append(ptype)
    total_cost = conn.execute(cost_sql, cost_params).fetchone()['total_cost']

    # Product Type Breakdown for Wastage
    pt_rows = conn.execute('''
        SELECT p.product_type,
               COALESCE(SUM(st.quantity * p.purchase_price), 0) AS total_cost,
               COUNT(st.id) AS transaction_count
        FROM stock_transactions st
        JOIN products p ON st.product_id = p.id
        WHERE st.type = 'wastage' AND date(st.date) BETWEEN ? AND ?
        GROUP BY p.product_type
    ''', (date_from, date_to)).fetchall()

    by_product_type = {
        "perishable": {"total_cost": 0.0, "transaction_count": 0},
        "general":    {"total_cost": 0.0, "transaction_count": 0}
    }
    for ptr in pt_rows:
        t = ptr['product_type'] or 'perishable'
        if t in by_product_type:
            by_product_type[t] = {
                "total_cost": round(float(ptr['total_cost'] or 0), 2),
                "transaction_count": int(ptr['transaction_count'] or 0)
            }

    conn.close()

    if wants_excel():
        excel_rows = []
        for r in rows:
            excel_rows.append([
                r['product_name'] or '',
                r['product_code'] or '',
                r['product_type'] or '',
                r['purchase_unit'] or '',
                float(r['total_wastage_purchase_units'] or 0),
                float(r['total_wastage_sale_units'] or 0),
                float(r['total_wastage_cost'] or 0),
                int(r['transaction_count'] or 0),
                r['reasons'] or ''
            ])
        sheets = [{
            "sheet_name": "Wastage Summary",
            "headers": ["Product Name", "Product Code", "Product Type", "Purchase Unit", "Wastage Qty (Purchase Units)", "Wastage Qty (Sale Units)", "Wastage Cost (₹)", "Transaction Count", "Reasons"],
            "rows": excel_rows
        }]
        return export_to_excel(sheets, f"wastage_report_{date_from}_to_{date_to}")

    return ok({
        "wastage_summary": dict_rows(rows),
        "total_wastage_cost": round(total_cost, 2),
        "by_product_type": by_product_type,
        "from": date_from,
        "to": date_to
    })

@app.route('/api/reports/margin-by-category', methods=['GET'])
@require_permission('reports.view')
def margin_by_category():
    # Approximation Note: Cost is computed as (quantity * p.purchase_price) using the product's CURRENT purchase_price.
    # Future enhancement candidate: snapshot purchase_price onto bill_items at sale time for exact historical margin.
    date_from = request.args.get('from', str(date.today().replace(day=1)))
    date_to   = request.args.get('to', str(date.today()))
    ptype     = request.args.get('product_type', '').strip().lower()

    conn = get_db()
    all_cats = conn.execute('SELECT id, name, parent_category_id FROM categories').fetchall()
    cat_map = {c['id']: dict(c) for c in all_cats}

    def get_root_category(cid):
        if not cid or cid not in cat_map:
            return 0, 'Uncategorized'
        curr = cat_map[cid]
        visited = set()
        while curr.get('parent_category_id') and curr['parent_category_id'] in cat_map:
            if curr['id'] in visited:
                break
            visited.add(curr['id'])
            curr = cat_map[curr['parent_category_id']]
        return curr['id'], curr['name']

    sql = '''
        SELECT bi.product_id, bi.quantity, bi.amount, bi.cost_price, p.category_id, p.purchase_price, p.product_type
        FROM bill_items bi
        JOIN bills b ON bi.bill_id = b.id
        JOIN products p ON bi.product_id = p.id
        WHERE date(b.date) BETWEEN ? AND ? AND b.status = 'paid'
    '''
    params = [date_from, date_to]
    if ptype in ('perishable', 'general'):
        sql += ' AND p.product_type = ?'
        params.append(ptype)

    rows = conn.execute(sql, params).fetchall()
    conn.close()

    from collections import defaultdict
    cat_summary = defaultdict(lambda: {'quantity_sold': 0.0, 'revenue': 0.0, 'cost': 0.0})

    for r in rows:
        root_id, root_name = get_root_category(r['category_id'])
        qty = float(r['quantity'] or 0)
        rev = float(r['amount'] or 0)
        unit_cost = float(r['cost_price'] if r['cost_price'] is not None else (r['purchase_price'] or 0))
        cost = round(qty * unit_cost, 2)

        key = (root_id, root_name)
        cat_summary[key]['quantity_sold'] += qty
        cat_summary[key]['revenue'] += rev
        cat_summary[key]['cost'] += cost

    result_list = []
    for (cat_id, cat_name), data in cat_summary.items():
        rev = round(data['revenue'], 2)
        cost = round(data['cost'], 2)
        margin = round(rev - cost, 2)
        margin_pct = round((margin / rev) * 100, 2) if rev > 0 else 0.0

        result_list.append({
            'category_id': cat_id,
            'category_name': cat_name,
            'quantity_sold': round(data['quantity_sold'], 3),
            'revenue': rev,
            'cost': cost,
            'gross_margin': margin,
            'margin_percent': margin_pct
        })

    result_list.sort(key=lambda x: x['revenue'], reverse=True)

    if wants_excel():
        excel_rows = []
        for r in result_list:
            excel_rows.append([
                r['category_name'] or '',
                float(r['quantity_sold'] or 0),
                float(r['revenue'] or 0),
                float(r['cost'] or 0),
                float(r['gross_margin'] or 0),
                float(r['margin_percent'] or 0)
            ])
        sheets = [{
            "sheet_name": "Category Margins",
            "headers": ["Category Name", "Quantity Sold", "Revenue (₹)", "Cost (₹)", "Gross Margin (₹)", "Margin %"],
            "rows": excel_rows
        }]
        return export_to_excel(sheets, f"margin_by_category_{date_from}_to_{date_to}")

    return ok({
        "categories": result_list,
        "from": date_from,
        "to": date_to
    })

@app.route('/api/reports/dues', methods=['GET'])
@require_permission('reports.view')
def dues_report():
    conn = get_db()
    rows = conn.execute('''
        SELECT c.id AS customer_id, c.name AS customer_name, c.phone, c.email, c.gstin,
               SUM(b.amount_due) AS total_due, COUNT(b.id) AS due_bills_count,
               MAX(b.date) AS latest_bill_date
        FROM bills b
        JOIN customers c ON b.customer_id = c.id
        WHERE b.status != 'cancelled' AND b.amount_due > 0
        GROUP BY c.id
        ORDER BY total_due DESC
    ''').fetchall()

    total_outstanding = conn.execute(
        "SELECT COALESCE(SUM(amount_due), 0) FROM bills WHERE status != 'cancelled' AND amount_due > 0"
    ).fetchone()[0]

    conn.close()

    if wants_excel():
        excel_rows = []
        for r in rows:
            excel_rows.append([
                r['customer_name'] or '',
                r['phone'] or '',
                r['email'] or '',
                r['gstin'] or '',
                float(r['total_due'] or 0),
                int(r['due_bills_count'] or 0),
                r['latest_bill_date'] or ''
            ])
        sheets = [{
            "sheet_name": "Customer Dues",
            "headers": ["Customer Name", "Phone", "Email", "GSTIN", "Total Due (₹)", "Due Bills Count", "Oldest Due Bill Date"],
            "rows": excel_rows
        }]
        return export_to_excel(sheets, f"customer_dues_as_of_{date.today()}")

    return ok({
        "due_customers": dict_rows(rows),
        "total_outstanding_due": round(total_outstanding, 2)
    })

@app.route('/api/reports/payables', methods=['GET'])
@require_permission('reports.view')
def payables_report():
    conn = get_db()
    rows = conn.execute('''
        SELECT po.*, s.name AS supplier_name, s.phone AS supplier_phone, s.contact_person
        FROM purchase_orders po
        LEFT JOIN suppliers s ON po.supplier_id = s.id
        WHERE po.amount_due > 0 AND (po.status IS NULL OR po.status != 'cancelled')
        ORDER BY po.amount_due DESC
    ''').fetchall()

    total_payables = conn.execute(
        "SELECT COALESCE(SUM(amount_due), 0) FROM purchase_orders WHERE amount_due > 0 AND (status IS NULL OR status != 'cancelled')"
    ).fetchone()[0]

    conn.close()

    if wants_excel():
        excel_rows = []
        for r in rows:
            excel_rows.append([
                r['po_no'] or '',
                r['date'] or '',
                r['supplier_name'] or '',
                r['contact_person'] or '',
                r['supplier_phone'] or '',
                float(r['total_amount'] or 0),
                float(r['amount_paid'] or 0),
                float(r['amount_due'] or 0),
                r['status'] or ''
            ])
        sheets = [{
            "sheet_name": "Supplier Payables",
            "headers": ["PO Number", "Date", "Supplier Name", "Contact Person", "Phone", "Total Amount (₹)", "Amount Paid (₹)", "Amount Due (₹)", "Status"],
            "rows": excel_rows
        }]
        return export_to_excel(sheets, f"supplier_payables_as_of_{date.today()}")

    return ok({
        "payables": dict_rows(rows),
        "total_payables": round(total_payables, 2)
    })

# ─── GST Return Reports ──────────────────────────────────────────────────────

def build_gstr1_data(conn, date_from='', date_to=''):
    # Note: If credit_notes / credit_note_items tables exist, credit note subtotals & taxes
    # are subtracted from the matching bucket (matched by original bill's gstin/interstate/place of supply/rate)
    # so the figures reflect net amounts after returns.

    bill_sql = '''
        SELECT b.*, c.name AS customer_name_from_cust
        FROM bills b
        LEFT JOIN customers c ON b.customer_id = c.id
        WHERE b.status != 'cancelled'
    '''
    bill_params = []
    if date_from:
        bill_sql += ' AND date(b.date) >= ?'
        bill_params.append(date_from)
    if date_to:
        bill_sql += ' AND date(b.date) <= ?'
        bill_params.append(date_to)

    bills = conn.execute(bill_sql, bill_params).fetchall()

    bi_sql = '''
        SELECT bi.*, p.hsn_code AS product_hsn_code, b.customer_gstin, b.place_of_supply, b.is_interstate, b.grand_total, b.id AS bill_id
        FROM bill_items bi
        JOIN bills b ON bi.bill_id = b.id
        LEFT JOIN products p ON bi.product_id = p.id
        WHERE b.status != 'cancelled'
    '''
    bi_params = []
    if date_from:
        bi_sql += ' AND date(b.date) >= ?'
        bi_params.append(date_from)
    if date_to:
        bi_sql += ' AND date(b.date) <= ?'
        bi_params.append(date_to)

    bill_items = conn.execute(bi_sql, bi_params).fetchall()

    has_cn = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='credit_notes'").fetchone() is not None
    credit_notes = []
    credit_note_items = []

    if has_cn:
        cn_sql = '''
            SELECT cn.*, b.customer_gstin, b.place_of_supply, b.is_interstate, b.grand_total
            FROM credit_notes cn
            JOIN bills b ON cn.bill_id = b.id
            WHERE (cn.status IS NULL OR cn.status != 'cancelled')
        '''
        cn_params = []
        if date_from:
            cn_sql += ' AND date(cn.created_at) >= ?'
            cn_params.append(date_from)
        if date_to:
            cn_sql += ' AND date(cn.created_at) <= ?'
            cn_params.append(date_to)
        credit_notes = conn.execute(cn_sql, cn_params).fetchall()

        cni_sql = '''
            SELECT cni.*, p.hsn_code AS product_hsn_code, b.customer_gstin, b.place_of_supply, b.is_interstate, b.grand_total, cn.bill_id
            FROM credit_note_items cni
            JOIN credit_notes cn ON cni.credit_note_id = cn.id
            JOIN bills b ON cn.bill_id = b.id
            LEFT JOIN products p ON cni.product_id = p.id
            WHERE (cn.status IS NULL OR cn.status != 'cancelled')
        '''
        cni_params = []
        if date_from:
            cni_sql += ' AND date(cn.created_at) >= ?'
            cni_params.append(date_from)
        if date_to:
            cni_sql += ' AND date(cn.created_at) <= ?'
            cni_params.append(date_to)
        credit_note_items = conn.execute(cni_sql, cni_params).fetchall()

    # --- Section 1: B2B ---
    b2b_cn_totals = {}
    for cn in credit_notes:
        bid = cn['bill_id']
        if bid not in b2b_cn_totals:
            b2b_cn_totals[bid] = {'subtotal': 0.0, 'cgst': 0.0, 'sgst': 0.0, 'igst': 0.0, 'total': 0.0}
        b2b_cn_totals[bid]['subtotal'] += float(cn['subtotal'] or 0)
        b2b_cn_totals[bid]['cgst'] += float(cn['cgst'] or 0)
        b2b_cn_totals[bid]['sgst'] += float(cn['sgst'] or 0)
        b2b_cn_totals[bid]['igst'] += float(cn['igst'] or 0)
        b2b_cn_totals[bid]['total'] += float(cn['total'] or 0)

    b2b_groups = {}
    for b in bills:
        gstin = (b['customer_gstin'] or '').strip().upper()
        if not gstin:
            continue
        c_name = b['customer_name'] or b['customer_name_from_cust'] or 'Valued Customer'
        if gstin not in b2b_groups:
            b2b_groups[gstin] = {
                'gstin': gstin,
                'customer_name': c_name,
                'invoices': []
            }
        cn_tot = b2b_cn_totals.get(b['id'], {'subtotal': 0.0, 'cgst': 0.0, 'sgst': 0.0, 'igst': 0.0, 'total': 0.0})

        net_taxable = round(max(0.0, float(b['subtotal'] or 0) - cn_tot['subtotal']), 2)
        net_cgst = round(max(0.0, float(b['cgst'] or 0) - cn_tot['cgst']), 2)
        net_sgst = round(max(0.0, float(b['sgst'] or 0) - cn_tot['sgst']), 2)
        net_igst = round(max(0.0, float(b['igst'] or 0) - cn_tot['igst']), 2)
        net_inv_val = round(max(0.0, float(b['grand_total'] or 0) - cn_tot['total']), 2)

        b2b_groups[gstin]['invoices'].append({
            'bill_no': b['bill_no'],
            'date': str(b['date'])[:10] if b['date'] else '',
            'place_of_supply': b['place_of_supply'] or '',
            'taxable_value': net_taxable,
            'cgst': net_cgst,
            'sgst': net_sgst,
            'igst': net_igst,
            'invoice_value': net_inv_val
        })

    b2b_list = list(b2b_groups.values())

    def is_b2c_large(gstin, is_interstate, grand_total):
        has_no_gstin = not bool(gstin and gstin.strip())
        return has_no_gstin and bool(is_interstate) and float(grand_total or 0) > 250000

    # --- Section 2: B2C Large ---
    b2cl_groups = {}
    for bi in bill_items:
        gstin = bi['customer_gstin']
        is_interstate = bi['is_interstate']
        g_total = bi['grand_total']
        if is_b2c_large(gstin, is_interstate, g_total):
            pos = (bi['place_of_supply'] or '').strip()
            rate = float(bi['gst_rate'] or 0)
            key = (pos, rate)
            if key not in b2cl_groups:
                b2cl_groups[key] = {'taxable_value': 0.0, 'igst': 0.0, 'bill_ids': set()}
            b2cl_groups[key]['taxable_value'] += float(bi['taxable_amt'] or 0)
            b2cl_groups[key]['igst'] += float(bi['igst_amt'] or 0)
            b2cl_groups[key]['bill_ids'].add(bi['bill_id'])

    for cni in credit_note_items:
        gstin = cni['customer_gstin']
        is_interstate = cni['is_interstate']
        g_total = cni['grand_total']
        if is_b2c_large(gstin, is_interstate, g_total):
            pos = (cni['place_of_supply'] or '').strip()
            rate = float(cni['gst_rate'] or 0)
            key = (pos, rate)
            if key in b2cl_groups:
                b2cl_groups[key]['taxable_value'] -= float(cni['taxable_amt'] or 0)
                b2cl_groups[key]['igst'] -= float(cni['igst_amt'] or 0)

    b2c_large_list = []
    for (pos, rate), data in b2cl_groups.items():
        b2c_large_list.append({
            'place_of_supply': pos,
            'gst_rate': rate,
            'taxable_value': round(max(0.0, data['taxable_value']), 2),
            'igst': round(max(0.0, data['igst']), 2),
            'invoice_count': len(data['bill_ids'])
        })

    # --- Section 3: B2C Small ---
    b2cs_groups = {}
    for bi in bill_items:
        gstin = bi['customer_gstin']
        is_interstate = bi['is_interstate']
        g_total = bi['grand_total']
        has_no_gstin = not bool(gstin and gstin.strip())
        if has_no_gstin and not is_b2c_large(gstin, is_interstate, g_total):
            pos = (bi['place_of_supply'] or '').strip()
            rate = float(bi['gst_rate'] or 0)
            key = (pos, rate)
            if key not in b2cs_groups:
                b2cs_groups[key] = {'taxable_value': 0.0, 'cgst': 0.0, 'sgst': 0.0, 'igst': 0.0, 'bill_ids': set()}
            b2cs_groups[key]['taxable_value'] += float(bi['taxable_amt'] or 0)
            b2cs_groups[key]['cgst'] += float(bi['cgst_amt'] or 0)
            b2cs_groups[key]['sgst'] += float(bi['sgst_amt'] or 0)
            b2cs_groups[key]['igst'] += float(bi['igst_amt'] or 0)
            b2cs_groups[key]['bill_ids'].add(bi['bill_id'])

    for cni in credit_note_items:
        gstin = cni['customer_gstin']
        is_interstate = cni['is_interstate']
        g_total = cni['grand_total']
        has_no_gstin = not bool(gstin and gstin.strip())
        if has_no_gstin and not is_b2c_large(gstin, is_interstate, g_total):
            pos = (cni['place_of_supply'] or '').strip()
            rate = float(cni['gst_rate'] or 0)
            key = (pos, rate)
            if key in b2cs_groups:
                b2cs_groups[key]['taxable_value'] -= float(cni['taxable_amt'] or 0)
                b2cs_groups[key]['cgst'] -= float(cni['cgst_amt'] or 0)
                b2cs_groups[key]['sgst'] -= float(cni['sgst_amt'] or 0)
                b2cs_groups[key]['igst'] -= float(cni['igst_amt'] or 0)

    b2c_small_list = []
    for (pos, rate), data in b2cs_groups.items():
        b2c_small_list.append({
            'place_of_supply': pos,
            'gst_rate': rate,
            'taxable_value': round(max(0.0, data['taxable_value']), 2),
            'cgst': round(max(0.0, data['cgst']), 2),
            'sgst': round(max(0.0, data['sgst']), 2),
            'igst': round(max(0.0, data['igst']), 2),
            'invoice_count': len(data['bill_ids'])
        })

    # --- Section 4: HSN Summary ---
    hsn_groups = {}
    for bi in bill_items:
        hsn = (bi['hsn_code'] or bi['product_hsn_code'] or 'N/A').strip()
        rate = float(bi['gst_rate'] or 0)
        key = (hsn, rate)
        if key not in hsn_groups:
            hsn_groups[key] = {'quantity': 0.0, 'taxable_value': 0.0, 'cgst': 0.0, 'sgst': 0.0, 'igst': 0.0}
        hsn_groups[key]['quantity'] += float(bi['quantity'] or 0)
        hsn_groups[key]['taxable_value'] += float(bi['taxable_amt'] or 0)
        hsn_groups[key]['cgst'] += float(bi['cgst_amt'] or 0)
        hsn_groups[key]['sgst'] += float(bi['sgst_amt'] or 0)
        hsn_groups[key]['igst'] += float(bi['igst_amt'] or 0)

    for cni in credit_note_items:
        hsn = (cni['hsn_code'] or cni['product_hsn_code'] or 'N/A').strip()
        rate = float(cni['gst_rate'] or 0)
        key = (hsn, rate)
        if key not in hsn_groups:
            hsn_groups[key] = {'quantity': 0.0, 'taxable_value': 0.0, 'cgst': 0.0, 'sgst': 0.0, 'igst': 0.0}
        hsn_groups[key]['quantity'] -= float(cni['quantity'] or 0)
        hsn_groups[key]['taxable_value'] -= float(cni['taxable_amt'] or 0)
        hsn_groups[key]['cgst'] -= float(cni['cgst_amt'] or 0)
        hsn_groups[key]['sgst'] -= float(cni['sgst_amt'] or 0)
        hsn_groups[key]['igst'] -= float(cni['igst_amt'] or 0)

    hsn_summary_list = []
    for (hsn, rate), data in hsn_groups.items():
        hsn_summary_list.append({
            'hsn_code': hsn,
            'gst_rate': rate,
            'quantity': round(max(0.0, data['quantity']), 3),
            'taxable_value': round(max(0.0, data['taxable_value']), 2),
            'cgst': round(max(0.0, data['cgst']), 2),
            'sgst': round(max(0.0, data['sgst']), 2),
            'igst': round(max(0.0, data['igst']), 2)
        })

    return {
        'b2b': b2b_list,
        'b2c_large': b2c_large_list,
        'b2c_small': b2c_small_list,
        'hsn_summary': hsn_summary_list
    }


@app.route('/api/reports/gstr1', methods=['GET'])
@require_permission('reports.view')
def get_gstr1_report():
    date_from = request.args.get('from', '')
    date_to = request.args.get('to', '')
    conn = get_db()
    data = build_gstr1_data(conn, date_from, date_to)
    conn.close()
    return ok(data)


@app.route('/api/reports/gstr3b', methods=['GET'])
@require_permission('reports.view')
def get_gstr3b_report():
    date_from = request.args.get('from', '')
    date_to = request.args.get('to', '')

    conn = get_db()

    bill_sql = '''
        SELECT COALESCE(SUM(subtotal), 0) AS subtotal,
               COALESCE(SUM(cgst), 0) AS cgst,
               COALESCE(SUM(sgst), 0) AS sgst,
               COALESCE(SUM(igst), 0) AS igst
        FROM bills
        WHERE status != 'cancelled'
    '''
    bill_params = []
    if date_from:
        bill_sql += ' AND date(date) >= ?'
        bill_params.append(date_from)
    if date_to:
        bill_sql += ' AND date(date) <= ?'
        bill_params.append(date_to)

    b_row = conn.execute(bill_sql, bill_params).fetchone()

    has_cn = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='credit_notes'").fetchone() is not None
    cn_subtotal = 0.0
    cn_cgst = 0.0
    cn_sgst = 0.0
    cn_igst = 0.0

    if has_cn:
        cn_sql = '''
            SELECT COALESCE(SUM(subtotal), 0) AS subtotal,
                   COALESCE(SUM(cgst), 0) AS cgst,
                   COALESCE(SUM(sgst), 0) AS sgst,
                   COALESCE(SUM(igst), 0) AS igst
            FROM credit_notes
            WHERE (status IS NULL OR status != 'cancelled')
        '''
        cn_params = []
        if date_from:
            cn_sql += ' AND date(created_at) >= ?'
            cn_params.append(date_from)
        if date_to:
            cn_sql += ' AND date(created_at) <= ?'
            cn_params.append(date_to)
        cn_row = conn.execute(cn_sql, cn_params).fetchone()
        if cn_row:
            cn_subtotal = float(cn_row['subtotal'] or 0)
            cn_cgst = float(cn_row['cgst'] or 0)
            cn_sgst = float(cn_row['sgst'] or 0)
            cn_igst = float(cn_row['igst'] or 0)

    conn.close()

    taxable_val = round(max(0.0, float(b_row['subtotal'] or 0) - cn_subtotal), 2)
    tot_cgst = round(max(0.0, float(b_row['cgst'] or 0) - cn_cgst), 2)
    tot_sgst = round(max(0.0, float(b_row['sgst'] or 0) - cn_sgst), 2)
    tot_igst = round(max(0.0, float(b_row['igst'] or 0) - cn_igst), 2)
    tot_tax = round(tot_cgst + tot_sgst + tot_igst, 2)

    return ok({
        "outward_supplies": {
            "taxable_value": taxable_val,
            "cgst": tot_cgst,
            "sgst": tot_sgst,
            "igst": tot_igst,
            "total_tax": tot_tax
        },
        "input_tax_credit": {
            "cgst": 0,
            "sgst": 0,
            "igst": 0,
            "note": "ITC requires purchase-side GST capture on purchase_orders, not yet implemented — fill this manually from purchase invoices for now."
        }
    })


@app.route('/api/reports/gstr1/export', methods=['GET'])
@require_permission('reports.view')
def export_gstr1_report():
    date_from = request.args.get('from', '')
    date_to = request.args.get('to', '')

    conn = get_db()
    data = build_gstr1_data(conn, date_from, date_to)
    conn.close()

    wb = openpyxl.Workbook()

    # Sheet 1: B2B
    ws_b2b = wb.active
    ws_b2b.title = "B2B"
    ws_b2b.append([
        "GSTIN/UIN of Recipient", "Receiver Name", "Invoice Number", "Invoice Date",
        "Invoice Value", "Place Of Supply", "Reverse Charge", "Applicable % of Tax Rate",
        "Rate", "Taxable Value", "Cess Amount"
    ])
    for g in data.get('b2b', []):
        for inv in g.get('invoices', []):
            ws_b2b.append([
                g.get('gstin', ''),
                g.get('customer_name', ''),
                inv.get('bill_no', ''),
                inv.get('date', ''),
                inv.get('invoice_value', 0),
                inv.get('place_of_supply', ''),
                "N/A",
                "N/A",
                "N/A",
                inv.get('taxable_value', 0),
                0
            ])

    # Sheet 2: B2C Large
    ws_b2cl = wb.create_sheet(title="B2C Large")
    ws_b2cl.append([
        "Place Of Supply", "Applicable % of Tax Rate", "Rate",
        "Taxable Value", "Cess Amount"
    ])
    for item in data.get('b2c_large', []):
        ws_b2cl.append([
            item.get('place_of_supply', ''),
            "N/A",
            item.get('gst_rate', 0),
            item.get('taxable_value', 0),
            0
        ])

    # Sheet 3: B2C Small
    ws_b2cs = wb.create_sheet(title="B2C Small")
    ws_b2cs.append([
        "Type", "Place Of Supply", "Applicable % of Tax Rate", "Rate",
        "Taxable Value", "Cess Amount"
    ])
    for item in data.get('b2c_small', []):
        ws_b2cs.append([
            "E",
            item.get('place_of_supply', ''),
            "N/A",
            item.get('gst_rate', 0),
            item.get('taxable_value', 0),
            0
        ])

    # Sheet 4: HSN Summary
    ws_hsn = wb.create_sheet(title="HSN Summary")
    ws_hsn.append([
        "HSN", "Description", "UQC", "Total Quantity", "Taxable Value",
        "Integrated Tax Amount", "Central Tax Amount", "State/UT Tax Amount", "Cess Amount"
    ])
    for item in data.get('hsn_summary', []):
        ws_hsn.append([
            item.get('hsn_code', ''),
            "N/A",
            "N/A",
            item.get('quantity', 0),
            item.get('taxable_value', 0),
            item.get('igst', 0),
            item.get('cgst', 0),
            item.get('sgst', 0),
            0
        ])

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    f_from = date_from or "start"
    f_to = date_to or "end"
    filename = f"gstr1_{f_from}_{f_to}.xlsx"

    return send_file(
        out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ─── Double-Entry Ledger Chart of Accounts ───────────────────────────────────

@app.route('/api/ledger/accounts', methods=['GET'])
@require_permission('accounts.view_ledger')
def list_ledger_accounts():
    conn = get_db()
    rows = conn.execute('SELECT * FROM ledger_accounts ORDER BY account_group, name').fetchall()
    conn.close()
    return ok(dict_rows(rows))


@app.route('/api/ledger/accounts', methods=['POST'])
@require_permission('accounts.manage')
def create_ledger_account():
    d = request.get_json()
    if d is None:
        return err("Invalid or missing JSON payload")

    name = (d.get('name') or '').strip()
    if not name:
        return err("Account name is required")

    group = (d.get('account_group') or '').strip()
    valid_groups = ('Asset', 'Liability', 'Income', 'Expense', 'Equity')
    if group not in valid_groups:
        return err(f"Account group must be one of: {', '.join(valid_groups)}")

    acc_type = (d.get('account_type') or group).strip()

    try:
        op_bal = float(d.get('opening_balance', 0) or 0)
        if op_bal < 0:
            return err("Opening balance must be non-negative")
    except (ValueError, TypeError):
        return err("Invalid opening_balance")

    op_type = (d.get('opening_balance_type') or 'dr').strip().lower()
    if op_type not in ('dr', 'cr'):
        return err("Opening balance type must be 'dr' or 'cr'")

    conn = get_db()
    try:
        cur = conn.execute('''
            INSERT INTO ledger_accounts
            (name, account_group, account_type, opening_balance, opening_balance_type, is_system)
            VALUES (?, ?, ?, ?, ?, 0)
        ''', (name, group, acc_type, round(op_bal, 2), op_type))
        conn.commit()
        row = conn.execute('SELECT * FROM ledger_accounts WHERE id=?', (cur.lastrowid,)).fetchone()
        conn.close()
        return ok(dict_row(row), "Ledger account created"), 201
    except sqlite3.IntegrityError:
        conn.close()
        return err(f"Ledger account '{name}' already exists")


@app.route('/api/ledger/accounts/<int:acc_id>', methods=['PUT'])
@require_permission('accounts.manage')
def update_ledger_account(acc_id):
    d = request.get_json()
    if d is None:
        return err("Invalid or missing JSON payload")

    conn = get_db()
    account = conn.execute('SELECT * FROM ledger_accounts WHERE id=?', (acc_id,)).fetchone()
    if not account:
        conn.close()
        return err("Ledger account not found", 404)

    is_system = bool(account['is_system'])
    new_name = (d.get('name') or account['name']).strip()
    new_group = (d.get('account_group') or account['account_group']).strip()
    new_type = (d.get('account_type') or account['account_type']).strip()

    valid_groups = ('Asset', 'Liability', 'Income', 'Expense', 'Equity')
    if new_group not in valid_groups:
        conn.close()
        return err(f"Account group must be one of: {', '.join(valid_groups)}")

    if is_system and new_name != account['name']:
        conn.close()
        return err("Cannot rename system accounts to preserve automated ledger posting integrity")

    try:
        op_bal = float(d.get('opening_balance', account['opening_balance']) or 0)
        if op_bal < 0:
            conn.close()
            return err("Opening balance must be non-negative")
    except (ValueError, TypeError):
        conn.close()
        return err("Invalid opening_balance")

    op_type = (d.get('opening_balance_type') or account['opening_balance_type']).strip().lower()
    if op_type not in ('dr', 'cr'):
        conn.close()
        return err("Opening balance type must be 'dr' or 'cr'")

    try:
        if is_system:
            conn.execute('''
                UPDATE ledger_accounts
                SET opening_balance=?, opening_balance_type=?
                WHERE id=?
            ''', (round(op_bal, 2), op_type, acc_id))
        else:
            conn.execute('''
                UPDATE ledger_accounts
                SET name=?, account_group=?, account_type=?, opening_balance=?, opening_balance_type=?
                WHERE id=?
            ''', (new_name, new_group, new_type, round(op_bal, 2), op_type, acc_id))

        conn.commit()
        updated_row = conn.execute('SELECT * FROM ledger_accounts WHERE id=?', (acc_id,)).fetchone()
        conn.close()
        return ok(dict_row(updated_row), "Ledger account updated")
    except sqlite3.IntegrityError:
        conn.close()
        return err(f"Ledger account name '{new_name}' already exists")


# ─── Financial Statements & Ledger Statements ──────────────────────────────

@app.route('/api/reports/trial-balance', methods=['GET'])
@require_permission('accounts.view_ledger')
def trial_balance_report():
    as_of = request.args.get('as_of', str(date.today()))
    conn = get_db()
    accounts = conn.execute('SELECT * FROM ledger_accounts ORDER BY account_group, name').fetchall()

    result_accounts = []
    total_dr_closing = 0.0
    total_cr_closing = 0.0

    for acc in accounts:
        acc_id = acc['id']
        group = acc['account_group']
        op_bal = float(acc['opening_balance'] or 0)
        op_type = (acc['opening_balance_type'] or 'dr').lower()

        sums = conn.execute('''
            SELECT COALESCE(SUM(debit), 0) AS dr_sum, COALESCE(SUM(credit), 0) AS cr_sum
            FROM ledger_entries
            WHERE account_id = ? AND date(voucher_date) <= date(?)
        ''', (acc_id, as_of)).fetchone()

        dr_sum = float(sums['dr_sum'])
        cr_sum = float(sums['cr_sum'])

        total_dr = dr_sum + (op_bal if op_type == 'dr' else 0.0)
        total_cr = cr_sum + (op_bal if op_type == 'cr' else 0.0)

        if group in ('Asset', 'Expense'):
            net = total_dr - total_cr
            if net >= 0:
                side = 'dr'
                c_bal = round(net, 2)
            else:
                side = 'cr'
                c_bal = round(abs(net), 2)
        else:
            net = total_cr - total_dr
            if net >= 0:
                side = 'cr'
                c_bal = round(net, 2)
            else:
                side = 'dr'
                c_bal = round(abs(net), 2)

        if side == 'dr':
            total_dr_closing += c_bal
        else:
            total_cr_closing += c_bal

        result_accounts.append({
            'account_id': acc_id,
            'account_name': acc['name'],
            'account_group': group,
            'account_type': acc['account_type'],
            'debit_total': round(total_dr, 2),
            'credit_total': round(total_cr, 2),
            'closing_balance': c_bal,
            'balance_side': side
        })

    conn.close()

    total_dr_closing = round(total_dr_closing, 2)
    total_cr_closing = round(total_cr_closing, 2)
    is_balanced = abs(total_dr_closing - total_cr_closing) <= 0.01

    if wants_excel():
        excel_rows = [
            ["Status:", "BALANCED" if is_balanced else "UNBALANCED", "Total Debit Closing:", total_dr_closing, "Total Credit Closing:", total_cr_closing],
            []
        ]
        headers = ["Account Name", "Account Group", "Account Type", "Debit Total (₹)", "Credit Total (₹)", "Closing Balance (₹)", "Side"]
        rows_data = []
        for a in result_accounts:
            rows_data.append([
                a['account_name'],
                a['account_group'],
                a['account_type'],
                float(a['debit_total']),
                float(a['credit_total']),
                float(a['closing_balance']),
                a['balance_side'].upper()
            ])
        sheets = [{
            "sheet_name": "Trial Balance",
            "headers": headers,
            "rows": excel_rows + rows_data
        }]
        return export_to_excel(sheets, f"trial_balance_as_of_{as_of}")

    return ok({
        'as_of': as_of,
        'accounts': result_accounts,
        'total_debit': total_dr_closing,
        'total_credit': total_cr_closing,
        'is_balanced': is_balanced
    })


@app.route('/api/reports/profit-loss', methods=['GET'])
@require_permission('accounts.view_ledger')
def profit_loss_report():
    date_from = request.args.get('from', '')
    date_to = request.args.get('to', '')

    conn = get_db()
    accounts = conn.execute(
        "SELECT * FROM ledger_accounts WHERE account_group IN ('Income', 'Expense') ORDER BY name"
    ).fetchall()

    income_accounts = []
    expense_accounts = []
    total_income = 0.0
    total_expense = 0.0

    for acc in accounts:
        acc_id = acc['id']
        group = acc['account_group']
        op_bal = float(acc['opening_balance'] or 0)
        op_type = (acc['opening_balance_type'] or 'dr').lower()

        sql = 'SELECT COALESCE(SUM(debit), 0) AS dr_sum, COALESCE(SUM(credit), 0) AS cr_sum FROM ledger_entries WHERE account_id = ?'
        params = [acc_id]
        if date_from:
            sql += ' AND date(voucher_date) >= date(?)'
            params.append(date_from)
        if date_to:
            sql += ' AND date(voucher_date) <= date(?)'
            params.append(date_to)

        sums = conn.execute(sql, params).fetchone()
        dr_sum = float(sums['dr_sum'])
        cr_sum = float(sums['cr_sum'])

        if group == 'Income':
            inc_net = round(cr_sum - dr_sum + (op_bal if op_type == 'cr' else -op_bal), 2)
            income_accounts.append({
                'account_id': acc_id,
                'account_name': acc['name'],
                'amount': inc_net
            })
            total_income += inc_net
        else:
            exp_net = round(dr_sum - cr_sum + (op_bal if op_type == 'dr' else -op_bal), 2)
            expense_accounts.append({
                'account_id': acc_id,
                'account_name': acc['name'],
                'amount': exp_net
            })
            total_expense += exp_net

    conn.close()

    total_income = round(total_income, 2)
    total_expense = round(total_expense, 2)
    net_profit = round(total_income - total_expense, 2)

    if wants_excel():
        summary_sheet = {
            "sheet_name": "Summary",
            "headers": ["Metric", "Amount (₹)"],
            "rows": [
                ["Total Income", float(total_income)],
                ["Total Expense", float(total_expense)],
                ["Net Profit / (Loss)", float(net_profit)]
            ]
        }
        income_sheet = {
            "sheet_name": "Income",
            "headers": ["Account Name", "Amount (₹)"],
            "rows": [[r['account_name'], float(r['amount'])] for r in income_accounts]
        }
        expense_sheet = {
            "sheet_name": "Expenses",
            "headers": ["Account Name", "Amount (₹)"],
            "rows": [[r['account_name'], float(r['amount'])] for r in expense_accounts]
        }
        sheets = [summary_sheet, income_sheet, expense_sheet]
        f_from = date_from or "start"
        f_to = date_to or "end"
        return export_to_excel(sheets, f"profit_loss_{f_from}_to_{f_to}")

    return ok({
        'from': date_from,
        'to': date_to,
        'income_accounts': income_accounts,
        'expense_accounts': expense_accounts,
        'total_income': total_income,
        'total_expense': total_expense,
        'net_profit': net_profit
    })


@app.route('/api/reports/balance-sheet', methods=['GET'])
@require_permission('accounts.view_ledger')
def balance_sheet_report():
    as_of = request.args.get('as_of', str(date.today()))
    conn = get_db()

    accounts = conn.execute('SELECT * FROM ledger_accounts ORDER BY name').fetchall()

    assets = []
    liabilities = []
    equity = []

    inc_total_as_of = 0.0
    exp_total_as_of = 0.0

    for acc in accounts:
        acc_id = acc['id']
        group = acc['account_group']
        op_bal = float(acc['opening_balance'] or 0)
        op_type = (acc['opening_balance_type'] or 'dr').lower()

        sums = conn.execute('''
            SELECT COALESCE(SUM(debit), 0) AS dr_sum, COALESCE(SUM(credit), 0) AS cr_sum
            FROM ledger_entries
            WHERE account_id = ? AND date(voucher_date) <= date(?)
        ''', (acc_id, as_of)).fetchone()

        dr_sum = float(sums['dr_sum'])
        cr_sum = float(sums['cr_sum'])

        if group == 'Asset':
            bal = round((op_bal if op_type == 'dr' else -op_bal) + dr_sum - cr_sum, 2)
            assets.append({'account_id': acc_id, 'account_name': acc['name'], 'account_type': acc['account_type'], 'amount': bal})
        elif group == 'Liability':
            bal = round((op_bal if op_type == 'cr' else -op_bal) + cr_sum - dr_sum, 2)
            liabilities.append({'account_id': acc_id, 'account_name': acc['name'], 'account_type': acc['account_type'], 'amount': bal})
        elif group == 'Equity':
            bal = round((op_bal if op_type == 'cr' else -op_bal) + cr_sum - dr_sum, 2)
            equity.append({'account_id': acc_id, 'account_name': acc['name'], 'account_type': acc['account_type'], 'amount': bal})
        elif group == 'Income':
            inc_net = (op_bal if op_type == 'cr' else -op_bal) + cr_sum - dr_sum
            inc_total_as_of += inc_net
        elif group == 'Expense':
            exp_net = (op_bal if op_type == 'dr' else -op_bal) + dr_sum - cr_sum
            exp_total_as_of += exp_net

    retained_profit = round(inc_total_as_of - exp_total_as_of, 2)
    equity.append({
        'account_id': None,
        'account_name': 'Retained Earnings / Current Profit',
        'account_type': 'Retained Earnings',
        'amount': retained_profit
    })

    conn.close()

    tot_assets = round(sum(a['amount'] for a in assets), 2)
    tot_liab = round(sum(l['amount'] for l in liabilities), 2)
    tot_equity = round(sum(e['amount'] for e in equity), 2)
    tot_liab_eq = round(tot_liab + tot_equity, 2)
    is_balanced = abs(tot_assets - tot_liab_eq) <= 0.01

    if wants_excel():
        summary_sheet = {
            "sheet_name": "Summary",
            "headers": ["Metric", "Amount (₹)"],
            "rows": [
                ["Total Assets", float(tot_assets)],
                ["Total Liabilities", float(tot_liab)],
                ["Total Equity", float(tot_equity)],
                ["Total Liabilities & Equity", float(tot_liab_eq)],
                ["Balance Check Status", "BALANCED" if is_balanced else "UNBALANCED"]
            ]
        }
        assets_sheet = {
            "sheet_name": "Assets",
            "headers": ["Account Name", "Account Type", "Amount (₹)"],
            "rows": [[r['account_name'], r['account_type'], float(r['amount'])] for r in assets]
        }
        liab_sheet = {
            "sheet_name": "Liabilities",
            "headers": ["Account Name", "Account Type", "Amount (₹)"],
            "rows": [[r['account_name'], r['account_type'], float(r['amount'])] for r in liabilities]
        }
        equity_sheet = {
            "sheet_name": "Equity",
            "headers": ["Account Name", "Account Type", "Amount (₹)"],
            "rows": [[r['account_name'], r['account_type'], float(r['amount'])] for r in equity]
        }
        sheets = [summary_sheet, assets_sheet, liab_sheet, equity_sheet]
        return export_to_excel(sheets, f"balance_sheet_as_of_{as_of}")

    return ok({
        'as_of': as_of,
        'assets': assets,
        'liabilities': liabilities,
        'equity': equity,
        'retained_earnings': retained_profit,
        'total_assets': tot_assets,
        'total_liabilities': tot_liab,
        'total_equity': tot_equity,
        'total_liabilities_equity': tot_liab_eq,
        'is_balanced': is_balanced
    })


@app.route('/api/ledger/accounts/<int:acc_id>/statement', methods=['GET'])
@require_permission('accounts.view_ledger')
def get_ledger_account_statement(acc_id):
    date_from = request.args.get('from', '')
    date_to = request.args.get('to', '')

    conn = get_db()
    account = conn.execute('SELECT * FROM ledger_accounts WHERE id=?', (acc_id,)).fetchone()
    if not account:
        conn.close()
        return err("Ledger account not found", 404)

    group = account['account_group']
    op_bal = float(account['opening_balance'] or 0)
    op_type = (account['opening_balance_type'] or 'dr').lower()

    prior_dr = 0.0
    prior_cr = 0.0
    if date_from:
        prior_sums = conn.execute('''
            SELECT COALESCE(SUM(debit), 0) AS dr_sum, COALESCE(SUM(credit), 0) AS cr_sum
            FROM ledger_entries
            WHERE account_id = ? AND date(voucher_date) < date(?)
        ''', (acc_id, date_from)).fetchone()
        prior_dr = float(prior_sums['dr_sum'])
        prior_cr = float(prior_sums['cr_sum'])

    if group in ('Asset', 'Expense'):
        running = (op_bal if op_type == 'dr' else -op_bal) + prior_dr - prior_cr
    else:
        running = (op_bal if op_type == 'cr' else -op_bal) + prior_cr - prior_dr

    opening_balance = round(running, 2)

    sql = 'SELECT * FROM ledger_entries WHERE account_id = ?'
    params = [acc_id]
    if date_from:
        sql += ' AND date(voucher_date) >= date(?)'
        params.append(date_from)
    if date_to:
        sql += ' AND date(voucher_date) <= date(?)'
        params.append(date_to)

    sql += ' ORDER BY date(voucher_date) ASC, id ASC'
    rows = conn.execute(sql, params).fetchall()
    conn.close()

    statement_entries = []
    for r in dict_rows(rows):
        dr = float(r['debit'] or 0)
        cr = float(r['credit'] or 0)
        if group in ('Asset', 'Expense'):
            running = running + dr - cr
        else:
            running = running + cr - dr

        r['running_balance'] = round(running, 2)
        statement_entries.append(r)

    return ok({
        'account': dict_row(account),
        'from': date_from,
        'to': date_to,
        'opening_balance': opening_balance,
        'closing_balance': round(running, 2),
        'entries': statement_entries
    })


# ─── Backup / Restore ────────────────────────────────────────────────────────
# ─── Tally Prime XML Export Routes ───────────────────────────────────────────

def get_tally_parent_group(acc_name, acc_group, acc_type):
    name = (acc_name or '').strip()
    grp = (acc_group or '').strip()
    typ = (acc_type or '').strip()

    name_lower = name.lower()
    typ_lower = typ.lower()

    if name_lower == 'cash':
        return 'Cash-in-Hand'
    if name_lower == 'bank':
        return 'Bank Accounts'
    if name_lower in ('sundry debtors', 'debtors'):
        return 'Sundry Debtors'
    if name_lower in ('sundry creditors', 'creditors'):
        return 'Sundry Creditors'
    if name_lower in ('cgst payable', 'sgst payable', 'igst payable') or typ_lower == 'duties & taxes':
        return 'Duties & Taxes'
    if name_lower == 'sales account':
        return 'Sales Accounts'
    if name_lower == 'purchase account':
        return 'Purchase Accounts'
    if name_lower == 'capital account' or typ_lower == 'capital':
        return 'Capital Account'

    if typ_lower in ('direct expense', 'direct expenses'):
        return 'Direct Expenses'
    if typ_lower in ('indirect expense', 'indirect expenses') or name_lower in ('discount allowed', 'round off'):
        return 'Indirect Expenses'

    if grp == 'Asset':
        return 'Current Assets'
    if grp == 'Liability':
        return 'Current Liabilities'
    if grp == 'Income':
        return 'Sales Accounts'
    if grp == 'Expense':
        return 'Indirect Expenses'
    if grp == 'Equity':
        return 'Capital Account'

    return 'Primary'


@app.route('/api/ledger/accounts/export-tally-masters', methods=['GET'])
@require_permission('accounts.view_ledger')
def export_tally_masters():
    conn = get_db()
    accounts = conn.execute('SELECT * FROM ledger_accounts ORDER BY account_group, name').fetchall()
    conn.close()

    import xml.etree.ElementTree as ET

    envelope = ET.Element('ENVELOPE')

    header = ET.SubElement(envelope, 'HEADER')
    tally_req = ET.SubElement(header, 'TALLYREQUEST')
    tally_req.text = 'Import Data'

    body = ET.SubElement(envelope, 'BODY')
    import_data = ET.SubElement(body, 'IMPORTDATA')

    req_desc = ET.SubElement(import_data, 'REQUESTDESC')
    rep_name = ET.SubElement(req_desc, 'REPORTNAME')
    rep_name.text = 'All Masters'
    static_vars = ET.SubElement(req_desc, 'STATICVARIABLES')
    company_var = ET.SubElement(static_vars, 'SVCURRENTCOMPANY')
    company_var.text = 'Meat Products of India'

    req_data = ET.SubElement(import_data, 'REQUESTDATA')

    for acc in accounts:
        parent_group = get_tally_parent_group(acc['name'], acc['account_group'], acc['account_type'])

        op_bal = float(acc['opening_balance'] or 0)
        op_type = (acc['opening_balance_type'] or 'dr').lower()
        if op_type == 'dr':
            op_str = f"-{op_bal:.2f}" if op_bal > 0 else "0.00"
        else:
            op_str = f"{op_bal:.2f}"

        msg = ET.SubElement(req_data, 'TALLYMESSAGE', {'xmlns:UDF': 'TallyUDF'})
        ledger_node = ET.SubElement(msg, 'LEDGER', {'NAME': acc['name'], 'ACTION': 'Create'})

        name_node = ET.SubElement(ledger_node, 'NAME')
        name_node.text = acc['name']

        parent_node = ET.SubElement(ledger_node, 'PARENT')
        parent_node.text = parent_group

        opbal_node = ET.SubElement(ledger_node, 'OPENINGBALANCE')
        opbal_node.text = op_str

    xml_str = ET.tostring(envelope, encoding='utf-8', xml_declaration=True)
    out = BytesIO(xml_str)

    return send_file(
        out,
        mimetype='application/xml',
        as_attachment=True,
        download_name='tally_masters.xml'
    )


@app.route('/api/ledger/export-tally-vouchers', methods=['GET'])
@require_permission('accounts.view_ledger')
def export_tally_vouchers():
    date_from = request.args.get('from', '')
    date_to = request.args.get('to', '')
    vch_type_filter = (request.args.get('voucher_type') or '').strip().lower()

    conn = get_db()
    sql = '''
        SELECT le.*, la.name AS account_name
        FROM ledger_entries le
        JOIN ledger_accounts la ON le.account_id = la.id
        WHERE 1=1
    '''
    params = []
    if date_from:
        sql += ' AND date(le.voucher_date) >= date(?)'
        params.append(date_from)
    if date_to:
        sql += ' AND date(le.voucher_date) <= date(?)'
        params.append(date_to)
    if vch_type_filter:
        sql += ' AND le.voucher_type = ?'
        params.append(vch_type_filter)

    sql += ' ORDER BY date(le.voucher_date) ASC, le.id ASC'
    rows = conn.execute(sql, params).fetchall()
    conn.close()

    vch_map = {
        'sales': 'Sales',
        'credit_note': 'Credit Note',
        'purchase': 'Purchase',
        'payment_in': 'Receipt',
        'payment_out': 'Payment',
        'expense': 'Payment',
        'journal': 'Journal'
    }

    from collections import defaultdict
    grouped_vouchers = defaultdict(list)
    for r in rows:
        key = (r['voucher_type'], r['voucher_no'])
        grouped_vouchers[key].append(r)

    import xml.etree.ElementTree as ET

    envelope = ET.Element('ENVELOPE')

    header = ET.SubElement(envelope, 'HEADER')
    tally_req = ET.SubElement(header, 'TALLYREQUEST')
    tally_req.text = 'Import Data'

    body = ET.SubElement(envelope, 'BODY')
    import_data = ET.SubElement(body, 'IMPORTDATA')

    req_desc = ET.SubElement(import_data, 'REQUESTDESC')
    rep_name = ET.SubElement(req_desc, 'REPORTNAME')
    rep_name.text = 'Vouchers'
    static_vars = ET.SubElement(req_desc, 'STATICVARIABLES')
    company_var = ET.SubElement(static_vars, 'SVCURRENTCOMPANY')
    company_var.text = 'Meat Products of India'

    req_data = ET.SubElement(import_data, 'REQUESTDATA')

    for (v_type, v_no), entries in grouped_vouchers.items():
        vch_name = vch_map.get(v_type, 'Journal')

        raw_date = entries[0]['voucher_date'] or str(date.today())
        try:
            d_obj = datetime.strptime(raw_date[:10], '%Y-%m-%d')
            tally_date = d_obj.strftime('%Y%m%d')
        except Exception:
            tally_date = datetime.now().strftime('%Y%m%d')

        msg = ET.SubElement(req_data, 'TALLYMESSAGE', {'xmlns:UDF': 'TallyUDF'})
        vch_node = ET.SubElement(msg, 'VOUCHER', {'VCHTYPE': vch_name, 'ACTION': 'Create'})

        date_node = ET.SubElement(vch_node, 'DATE')
        date_node.text = tally_date

        vtype_node = ET.SubElement(vch_node, 'VOUCHERTYPENAME')
        vtype_node.text = vch_name

        vno_node = ET.SubElement(vch_node, 'VOUCHERNUMBER')
        vno_node.text = str(v_no or '')

        view_node = ET.SubElement(vch_node, 'PERSISTEDVIEW')
        view_node.text = 'Accounting Voucher View'

        for entry in entries:
            dr = float(entry['debit'] or 0)
            cr = float(entry['credit'] or 0)

            is_deemed_pos = 'Yes' if dr > 0 else 'No'
            amt_val = -dr if dr > 0 else cr

            entry_list = ET.SubElement(vch_node, 'ALLLEDGERENTRIES.LIST')

            lname = ET.SubElement(entry_list, 'LEDGERNAME')
            lname.text = entry['account_name']

            pos_node = ET.SubElement(entry_list, 'ISDEEMEDPOSITIVE')
            pos_node.text = is_deemed_pos

            amt_node = ET.SubElement(entry_list, 'AMOUNT')
            amt_node.text = f"{amt_val:.2f}"

    xml_str = ET.tostring(envelope, encoding='utf-8', xml_declaration=True)
    out = BytesIO(xml_str)

    f_from = date_from or "start"
    f_to = date_to or "end"
    filename = f"tally_vouchers_{f_from}_{f_to}.xml"

    return send_file(
        out,
        mimetype='application/xml',
        as_attachment=True,
        download_name=filename
    )
@app.route('/api/backup', methods=['GET'])
@require_permission('backup.manage')
def backup():
    from database import DB_PATH
    if not os.path.exists(DB_PATH):
        return err("No database found", 404)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_name = f"meatshop_backup_{ts}.db"
    return send_file(DB_PATH, as_attachment=True, download_name=backup_name)

@app.route('/api/backup/cloud-now', methods=['POST'])
@require_permission('backup.manage')
def backup_cloud_now():
    success, msg = run_cloud_backup_job()
    if success:
        return ok(data={'message': msg}, message=f"Cloud Backup Success: {msg}")
    return err(f"Cloud Backup Failed: {msg}", 400)

# ─── Main ────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    init_db()
    try:
        start_cloud_backup_scheduler()
    except Exception as e:
        print(f"Cloud backup scheduler init notice: {e}")

    print("=" * 60)
    print("  Meat Products of India — Billing & Inventory App")
    print("  Server: http://localhost:5000")
    print("=" * 60)
    app.run(debug=False, host='127.0.0.1', port=5000, use_reloader=False)
